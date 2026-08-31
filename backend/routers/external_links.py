"""External Links - the company-wide app/link directory (rebuilt Aug 2026 from
the old start.greensglobal.com launchpad, which is company-wide, unstyled, and
department-scoped only by a raw dropdown). This owns models.ExternalLink end
to end.

NOTE for whoever touches backend/routers/assets.py next: that file still has a
leftover stub /external-links GET/POST/click trio from before this module was
built out. This router is the real implementation (full CRUD, department
scoping, ordering, audit trail) and registers its own routes - do not restore
the assets.py stub, and remove it there next time that file is touched (it is
Ankush's file, not this module's, so it's left alone here rather than edited
in place).

Read is baseline for every employee (App.jsx treats external-links as a
baseline module like sop). Managing entries (create/edit/delete/reorder)
requires manager+ globally OR an Access Group grant of "external-links" at
editor+ (full+ to delete), mirroring items.py's require_items_admin pattern.
"""
import asyncio
import html as html_lib
import io
import ipaddress
import json
import os
import re
import socket
import uuid
from datetime import datetime, timezone
from urllib.parse import urlparse
import httpx
from fastapi import APIRouter, Depends, HTTPException, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional

import models
from database import get_db
from auth import get_current_user, require_level_or_module

router = APIRouter(prefix="/external-links", tags=["External Links"], dependencies=[Depends(get_current_user)])

_ROLE_LEVEL = {"employee": 1, "supervisor": 2, "manager": 3, "administrator": 4, "owner": 5}
require_links_admin  = require_level_or_module(_ROLE_LEVEL["manager"],       "external-links", "editor")
require_links_delete = require_level_or_module(_ROLE_LEVEL["administrator"], "external-links", "full")


# IT service areas, for the Ticket module's intake form. A ticket raised
# against an app copies its link's area and is triaged by it, so this is the
# desk's own taxonomy rather than the (user-facing, admin-editable) category
# list above - which is why it lives as a constant and not in
# ExternalLinkTaxonomy. Mirrors SERVICE_AREAS in frontend/src/tickets/
# ticketMeta.js - keep the two in step.
SERVICE_AREA_KEYS = (
    "email", "collab", "tasks", "files", "knowledge", "storageops", "finance",
    "hr", "assets", "network", "security", "web", "hardware", "general",
)


def _clean_service_area(value: str) -> str:
    """An unknown key reads as "general" rather than 422-ing. A link is a
    launcher first: refusing to save one because a service area was dropped
    from the list would break the screen that owns it over a field the Ticket
    module merely borrows."""
    v = (value or "").strip().lower()
    return v if v in SERVICE_AREA_KEYS else ("general" if v else "")


class ExternalLinkCreate(BaseModel):
    name: str
    url: str
    categories: list[str] = []  # at least one required - validated below, not by Pydantic, so the 422 detail can be specific
    description: str = ""
    departments: list[str] = []  # [] = shown to every department (company-wide)
    company: str = ""         # "" = shown to every company; else an HrEntity.id
    icon: str = "Link2"       # lucide-react icon key, resolved client-side
    is_pinned: bool = False
    service_area: str = ""    # SERVICE_AREA_KEYS; "" reads as "general" at intake


class ExternalLinkUpdate(BaseModel):
    name: Optional[str] = None
    url: Optional[str] = None
    categories: Optional[list[str]] = None
    description: Optional[str] = None
    departments: Optional[list[str]] = None
    company: Optional[str] = None
    icon: Optional[str] = None
    is_pinned: Optional[bool] = None
    service_area: Optional[str] = None


def _clean_list(values: list[str]) -> list[str]:
    """Trim, drop blanks, de-dupe while preserving order - shared by create
    and update so "  Banking , Banking" doesn't become two entries."""
    seen, out = set(), []
    for v in values:
        v = (v or "").strip()
        if v and v not in seen:
            seen.add(v)
            out.append(v)
    return out


class ReorderEntry(BaseModel):
    id: int
    sort_order: int


class ImportRow(BaseModel):
    name: str = ""
    url: str = ""
    categories: list[str] = []
    departments: list[str] = []
    company: str = ""
    description: str = ""
    icon: str = "Link2"
    is_pinned: bool = False


class ImportRequest(BaseModel):
    rows: list[ImportRow]


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _audit(db: Session, user: dict, action: str, link: "models.ExternalLink", extra: dict | None = None):
    details = {"name": link.name, "url": link.url, "categories": link.categories, "departments": link.departments}
    if extra:
        details.update(extra)
    db.add(models.AuditLog(
        timestamp=_now(), user_email=user["email"], user_role=user.get("role", ""),
        action=action, resource_type="external_links", resource_id=str(link.id),
        details=json.dumps(details), ip_address="",
    ))


@router.get("")
def list_external_links(db: Session = Depends(get_db)):
    """Every employee can read the full directory - the department/category
    filtering is a client-side UX affordance, not an access boundary (these
    are just launch links, same as the old company-wide start page)."""
    return (
        db.query(models.ExternalLink)
        .order_by(models.ExternalLink.is_pinned.desc(), models.ExternalLink.sort_order.asc(), models.ExternalLink.name.asc())
        .all()
    )


@router.get("/meta")
def external_links_meta(db: Session = Depends(get_db)):
    """Distinct departments/categories currently in use, for the filter
    dropdowns - avoids a second hardcoded list drifting from real data. Each
    link can carry several of each now (Aug 14, "add multiple checkbox
    option in departments and category"), so this flattens every row's
    array client-side rather than a plain SQL DISTINCT (no portable way to
    dedupe across JSON array elements in one query across both engines)."""
    departments, categories = set(), set()
    for (d,) in db.query(models.ExternalLink.departments).all():
        departments.update(d or [])
    for (c,) in db.query(models.ExternalLink.categories).all():
        categories.update(c or [])
    return {"departments": sorted(departments), "categories": sorted(categories)}


# ── Taxonomy (admin-managed Department/Category picker options) ────────────

_TAXONOMY_KINDS = ("department", "category")


def _taxonomy_dict(row: "models.ExternalLinkTaxonomy") -> dict:
    return {"id": row.id, "kind": row.kind, "name": row.name}


@router.get("/taxonomy")
def list_taxonomy(db: Session = Depends(get_db)):
    """Every employee can read this (it only feeds picker/filter dropdowns,
    same posture as list_external_links above) - only add/rename/remove is
    admin-gated below."""
    rows = db.query(models.ExternalLinkTaxonomy).order_by(
        models.ExternalLinkTaxonomy.kind, models.ExternalLinkTaxonomy.sort_order, models.ExternalLinkTaxonomy.name
    ).all()
    return {
        "departments": [_taxonomy_dict(r) for r in rows if r.kind == "department"],
        "categories": [_taxonomy_dict(r) for r in rows if r.kind == "category"],
    }


class TaxonomyCreate(BaseModel):
    kind: str
    name: str


@router.post("/taxonomy")
def create_taxonomy(body: TaxonomyCreate, user: dict = Depends(require_links_admin), db: Session = Depends(get_db)):
    if body.kind not in _TAXONOMY_KINDS:
        raise HTTPException(status_code=422, detail="kind must be 'department' or 'category'.")
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=422, detail="Name can't be empty.")
    if len(name) > 80:
        raise HTTPException(status_code=422, detail="Name must be 80 characters or fewer.")
    exists = db.query(models.ExternalLinkTaxonomy).filter(
        models.ExternalLinkTaxonomy.kind == body.kind, models.ExternalLinkTaxonomy.name == name).first()
    if exists:
        raise HTTPException(status_code=409, detail=f'"{name}" already exists.')
    max_pos = db.query(models.ExternalLinkTaxonomy).filter(models.ExternalLinkTaxonomy.kind == body.kind).count()
    row = models.ExternalLinkTaxonomy(
        id=str(uuid.uuid4()), kind=body.kind, name=name, sort_order=max_pos, created_at=_now())
    db.add(row)
    db.commit()
    return _taxonomy_dict(row)


class TaxonomyRename(BaseModel):
    name: str


@router.patch("/taxonomy/{taxonomy_id}")
def rename_taxonomy(taxonomy_id: str, body: TaxonomyRename, user: dict = Depends(require_links_admin), db: Session = Depends(get_db)):
    """Renaming bulk-updates every ExternalLink row currently tagged with
    the old string, in the same transaction - see ExternalLinkTaxonomy's
    docstring for why that matters (department/category aren't a FK)."""
    row = db.query(models.ExternalLinkTaxonomy).filter(models.ExternalLinkTaxonomy.id == taxonomy_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Not found.")
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=422, detail="Name can't be empty.")
    if len(name) > 80:
        raise HTTPException(status_code=422, detail="Name must be 80 characters or fewer.")
    dupe = db.query(models.ExternalLinkTaxonomy).filter(
        models.ExternalLinkTaxonomy.kind == row.kind, models.ExternalLinkTaxonomy.name == name,
        models.ExternalLinkTaxonomy.id != taxonomy_id).first()
    if dupe:
        raise HTTPException(status_code=409, detail=f'"{name}" already exists.')

    old_name = row.name
    row.name = name
    if old_name != name:
        # A link can hold several departments/categories now, so this can't
        # be a single portable SQL UPDATE (no cross-engine way to replace
        # one element inside a JSON array column) - fetch every row that has
        # the old name anywhere in its list and rewrite that one element in
        # Python instead.
        field_name = "departments" if row.kind == "department" else "categories"
        field = getattr(models.ExternalLink, field_name)
        candidates = db.query(models.ExternalLink).filter(field.isnot(None)).all()
        for link in candidates:
            values = getattr(link, field_name) or []
            if old_name in values:
                setattr(link, field_name, _clean_list([name if v == old_name else v for v in values]))
    db.commit()
    return _taxonomy_dict(row)


@router.delete("/taxonomy/{taxonomy_id}")
def delete_taxonomy(taxonomy_id: str, user: dict = Depends(require_links_admin), db: Session = Depends(get_db)):
    """Removes the option from the curated picker only - links already
    tagged with this name keep it (free text, same as Category always
    was), they just won't see it as a suggested/fixed choice going forward."""
    row = db.query(models.ExternalLinkTaxonomy).filter(models.ExternalLinkTaxonomy.id == taxonomy_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Not found.")
    db.delete(row)
    db.commit()
    return {"ok": True}


def _normalize_url(url: str) -> str:
    """Duplicate-link detection (Add Link / Add Personal Link, Aug 13) -
    collapses the differences that would otherwise let the same site get
    added twice (http vs https, www. vs not, a trailing slash, mixed case)
    without masking genuinely different pages on the same host. Mirrors
    normalizeUrl in ExternalLinks.jsx - keep both in sync. This is a
    server-side backstop behind the client-side check (which is what a user
    actually sees); it exists for the API path directly, not for UX."""
    try:
        parsed = urlparse(url if re.match(r"^https?://", url, re.I) else f"https://{url}")
        host = (parsed.hostname or "").lower()
        if host.startswith("www."):
            host = host[4:]
        path = parsed.path.rstrip("/")
        return f"{host}{path}"
    except ValueError:
        return url.strip().lower()


_META_DESC_RE = re.compile(
    r'<meta[^>]+(?:name|property)=["\'](?:description|og:description)["\'][^>]+content=["\']([^"\']*)["\']', re.I
)
_META_DESC_RE_REV = re.compile(
    r'<meta[^>]+content=["\']([^"\']*)["\'][^>]+(?:name|property)=["\'](?:description|og:description)["\']', re.I
)
_TITLE_RE = re.compile(r"<title[^>]*>([^<]*)</title>", re.I)


def _resolves_to_public_ip(hostname: str) -> bool:
    """SSRF guard for /preview below - an admin-supplied URL must not be able
    to make the backend fetch internal/private infrastructure (RFC1918,
    loopback, link-local incl. the 169.254.169.254 cloud metadata endpoint).
    ip.is_global is False for all of those."""
    try:
        infos = socket.getaddrinfo(hostname, None)
    except socket.gaierror:
        return False
    for info in infos:
        try:
            ip = ipaddress.ip_address(info[4][0])
        except ValueError:
            return False
        if not ip.is_global:
            return False
    return True


_CATEGORY_MODEL = "claude-haiku-4-5-20251001"  # cheap: one tiny classification call per Add Link
_ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "").strip()
_MAX_CATEGORY_WORDS = 3
_MAX_CATEGORY_LEN = 30


def _classify_category(hostname: str, title: str, meta_description: str) -> str:
    """Neil, Aug 14: the auto-filled description was reading as a full
    marketing sentence pulled verbatim off the site ("F&M Bank is a local
    Southern California community bank with more than 100 years...") -
    wanted instead as a one-to-three-word category tag (Banking, AI Tool,
    Shopping Platform). No amount of truncating the meta description gets
    there - it needs an actual classification, not a substring - so this
    asks Claude Haiku for the short label instead of scraping one. Same
    call pattern as every other AI feature in this backend (raw httpx to
    the Anthropic Messages API, no SDK dependency - see construction_ai.py's
    docstring for why). Best-effort: returns "" on a missing key or any
    failure, same as the rest of this endpoint - this is a convenience
    prefill, never something that should block Add Link."""
    if not _ANTHROPIC_API_KEY:
        return ""
    context = f"Website: {hostname}"
    if title:
        context += f"\nPage title: {title}"
    if meta_description:
        context += f"\nMeta description: {meta_description[:400]}"
    try:
        resp = httpx.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": _ANTHROPIC_API_KEY,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": _CATEGORY_MODEL,
                "max_tokens": 20,
                "system": (
                    "You label a website with the single best short category for what kind of "
                    "application or tool it is. Reply with ONLY 1-3 words, no punctuation, no "
                    "explanation - just the category. Examples: Banking, AI Tool, Shopping Platform, "
                    "Video Streaming, Project Management, Team Chat, Cloud Storage, Payroll, "
                    "Expense Management, Social Media."
                ),
                "messages": [{"role": "user", "content": f"{context}\n\nShort category:"}],
            },
            timeout=8.0,
        )
        resp.raise_for_status()
        text = "".join(b.get("text", "") for b in resp.json().get("content", []) if b.get("type") == "text").strip()
    except (httpx.HTTPError, ValueError, KeyError):
        return ""
    # The model is asked for 1-3 words but is never trusted to actually stop
    # there - cap hard so a verbose reply can never regress back into a full
    # sentence, which is the exact thing this feature exists to avoid.
    words = text.strip(" .").split()
    return " ".join(words[:_MAX_CATEGORY_WORDS])[:_MAX_CATEGORY_LEN]


def _fetch_link_preview(url: str) -> dict:
    """Runs in a worker thread (see asyncio.to_thread in the route below) -
    this does blocking DNS + network I/O, which must never sit on the async
    event loop. Best-effort only: any failure (bad scheme, private IP, DNS
    failure, timeout, non-2xx, no meta description) returns {} rather than
    raising, since this is a convenience prefill for the Add Link modal, not
    something that should ever block or error out the save."""
    parsed = urlparse(url)
    if parsed.scheme not in ("http", "https") or not parsed.hostname:
        return {}
    if not _resolves_to_public_ip(parsed.hostname):
        return {}
    try:
        with httpx.Client(
            follow_redirects=True, timeout=6.0,
            headers={"User-Agent": "Mozilla/5.0 (compatible; NexusLinkPreview/1.0)"},
        ) as client:
            with client.stream("GET", url) as resp:
                final_host = urlparse(str(resp.url)).hostname
                if not final_host or not _resolves_to_public_ip(final_host):
                    return {}  # redirected off to a private host mid-request
                if resp.status_code >= 400:
                    return {}
                chunks, total = [], 0
                for chunk in resp.iter_text():
                    chunks.append(chunk)
                    total += len(chunk)
                    if total > 300_000:  # cap - only the <head> is needed
                        break
                page = "".join(chunks)
    except httpx.HTTPError:
        return {}
    desc_match = _META_DESC_RE.search(page) or _META_DESC_RE_REV.search(page)
    title_match = _TITLE_RE.search(page)
    meta_description = html_lib.unescape(desc_match.group(1)).strip() if desc_match else ""
    title = html_lib.unescape(title_match.group(1)).strip() if title_match else ""
    # No fallback to the raw meta description - a multi-sentence marketing
    # blurb is exactly what the 1-3 word category exists to replace (Aug 13:
    # "i said i want max 3 word description"). If classification is
    # unavailable (no ANTHROPIC_API_KEY configured, or the call failed),
    # leave the field blank rather than ever showing the long sentence -
    # blank is a safe, honest default the admin can fill in by hand.
    category = _classify_category(parsed.hostname, title, meta_description)
    return {"description": category[:300], "title": title[:120]}


@router.get("/preview")
async def preview_external_link(url: str, user: dict = Depends(get_current_user)):
    """Add Link modal auto-fill (Aug 13), used by both the Company Links Add
    Link modal (manager+) and the Personal Links Add modal (every employee -
    self-service, no admin gate). Any signed-in user is enough here: fetches
    the given URL server-side (a client-side fetch would be blocked by both
    CORS and CSP's connect-src, which has no reason to allowlist arbitrary
    third-party sites) and pulls its <meta name="description">/og:description
    to prefill the description field - see _resolves_to_public_ip for the
    SSRF guard, which is what actually needs to hold here since the URL is
    user-supplied. Always 200s with whatever it found, possibly {}."""
    return await asyncio.to_thread(_fetch_link_preview, url)


@router.post("", status_code=201)
def create_external_link(link: ExternalLinkCreate, user: dict = Depends(require_links_admin), db: Session = Depends(get_db)):
    target = _normalize_url(link.url)
    existing = next(
        (l for l in db.query(models.ExternalLink.name, models.ExternalLink.url).all() if _normalize_url(l.url) == target),
        None,
    )
    if existing:
        raise HTTPException(status_code=409, detail=f'This link is already added as "{existing.name}".')
    categories = _clean_list(link.categories)
    if not categories:
        raise HTTPException(status_code=422, detail="Pick at least one category.")
    departments = _clean_list(link.departments)
    now = _now()
    data = link.model_dump(exclude={"categories", "departments", "service_area"})
    data["service_area"] = _clean_service_area(link.service_area)
    db_link = models.ExternalLink(
        **data, categories=categories, departments=departments,
        # Legacy singular columns kept in sync on write, best-effort, purely
        # so nothing that still reads them (there's nothing left in this
        # codebase that does, but the columns are NOT NULL on `category`)
        # sees a stale/empty value - first pick wins, same as any "primary"
        # tag would.
        category=categories[0], department=departments[0] if departments else "",
        created_by=user["email"], created_at=now, updated_at=now,
    )
    db.add(db_link)
    db.flush()
    _audit(db, user, "Created external link", db_link)
    db.commit()
    db.refresh(db_link)
    return db_link


@router.patch("/reorder")
def reorder_external_links(entries: list[ReorderEntry], user: dict = Depends(require_links_admin), db: Session = Depends(get_db)):
    """Bulk drag-reorder within a category - one round trip instead of N PATCHes.
    MUST stay registered before PATCH /{link_id} below: Starlette matches
    routes in registration order, and /external-links/{link_id} would
    otherwise greedily match /external-links/reorder (link_id="reorder") and
    422 on the int parse before this route is ever tried."""
    by_id = {e.id: e.sort_order for e in entries}
    rows = db.query(models.ExternalLink).filter(models.ExternalLink.id.in_(by_id.keys())).all()
    for row in rows:
        row.sort_order = by_id[row.id]
    db.commit()
    return {"ok": True, "updated": len(rows)}


@router.patch("/{link_id}")
def update_external_link(link_id: int, patch: ExternalLinkUpdate, user: dict = Depends(require_links_admin), db: Session = Depends(get_db)):
    db_link = db.query(models.ExternalLink).filter(models.ExternalLink.id == link_id).first()
    if not db_link:
        raise HTTPException(status_code=404, detail="Link not found")
    changes = patch.model_dump(exclude_unset=True)
    if "categories" in changes:
        changes["categories"] = _clean_list(changes["categories"])
        if not changes["categories"]:
            raise HTTPException(status_code=422, detail="Pick at least one category.")
    if "departments" in changes:
        changes["departments"] = _clean_list(changes["departments"])
    if "service_area" in changes:
        changes["service_area"] = _clean_service_area(changes["service_area"])
    for field, value in changes.items():
        setattr(db_link, field, value)
    # Legacy singular columns kept in sync - see create_external_link's
    # comment for why.
    if "categories" in changes:
        db_link.category = changes["categories"][0]
    if "departments" in changes:
        db_link.department = changes["departments"][0] if changes["departments"] else ""
    db_link.updated_at = _now()
    _audit(db, user, "Updated external link", db_link, {"changed_fields": list(changes.keys())})
    db.commit()
    db.refresh(db_link)
    return db_link


@router.post("/{link_id}/refresh-description")
async def refresh_link_description(link_id: int, user: dict = Depends(require_links_admin), db: Session = Depends(get_db)):
    """Re-runs the same auto-fill /preview uses against an EXISTING link's
    URL and persists the result to the row (Aug 14). The short-category
    autofill only ever applied going forward, to new Add Link submissions -
    a link added before that shipped, or via CSV import (which never set a
    description at all), keeps whatever it was originally saved with until
    someone explicitly asks for a refresh, here or via the bulk version
    below. Best-effort: if the fetch/classification comes back empty, the
    existing description is left alone rather than blanked out."""
    db_link = db.query(models.ExternalLink).filter(models.ExternalLink.id == link_id).first()
    if not db_link:
        raise HTTPException(status_code=404, detail="Link not found")
    preview = await asyncio.to_thread(_fetch_link_preview, db_link.url)
    if preview.get("description"):
        db_link.description = preview["description"]
        db_link.updated_at = _now()
        db.commit()
        db.refresh(db_link)
    return db_link


@router.post("/refresh-descriptions")
async def refresh_all_link_descriptions(user: dict = Depends(require_links_admin), db: Session = Depends(get_db)):
    """Bulk version of the single-link refresh above - one pass over every
    Company Link, best-effort per row (a site that fails to fetch/classify
    just keeps its current description). Sequential, not parallel - this is
    an occasional admin action over a small directory, not a hot path worth
    the complexity of fanning the requests out concurrently."""
    links = db.query(models.ExternalLink).all()
    updated = 0
    for link in links:
        preview = await asyncio.to_thread(_fetch_link_preview, link.url)
        if preview.get("description") and preview["description"] != link.description:
            link.description = preview["description"]
            link.updated_at = _now()
            updated += 1
    db.commit()
    return {"updated": updated, "total": len(links)}


@router.delete("/{link_id}")
def delete_external_link(link_id: int, user: dict = Depends(require_links_delete), db: Session = Depends(get_db)):
    db_link = db.query(models.ExternalLink).filter(models.ExternalLink.id == link_id).first()
    if not db_link:
        raise HTTPException(status_code=404, detail="Link not found")
    _audit(db, user, "Deleted external link", db_link)
    db.delete(db_link)
    db.commit()
    return {"ok": True}


@router.patch("/{link_id}/click")
def increment_click(link_id: int, db: Session = Depends(get_db)):
    """Every employee can fire this (it's how "Most Used" ordering works) - no
    admin gate, just needs to be signed in."""
    link = db.query(models.ExternalLink).filter(models.ExternalLink.id == link_id).first()
    if not link:
        raise HTTPException(status_code=404, detail="Link not found")
    link.clicks += 1
    db.commit()
    db.refresh(link)
    return link


# ── Site logos, cached on our own origin ─────────────────────────────────────
# Every tile renders the real brand mark rather than a generic glyph, which
# means resolving a logo per domain. Doing that from the BROWSER meant 50+
# simultaneous requests to icon.horse per page view, per employee, all from one
# office egress IP - and it rate-limits per IP, so most tiles got HTTP 429 and
# fell back to the lucide glyph (Charmi/Neil, Aug 31: "are each of these
# supposed to have an image?"). Resolved here instead: once per domain, ever.
#
# There is no SSRF surface. The link's own URL is never fetched - the hostname
# is handed to two FIXED third-party resolvers, which do the fetching. The
# domain must already appear on a link in this Nexus, so this can't be driven
# as a general-purpose favicon proxy either.
_ICON_MAX_BYTES = 512 * 1024
_ICON_TTL_DAYS = 45          # a re-brand shouldn't need a deploy to show up
_ICON_RETRY_DAYS = 7         # but a domain with NO logo is retried far less often


def _icon_domain_candidates(domain: str) -> list:
    """The hostname, then its parent. A login subdomain usually publishes no
    favicon of its own (auth.hostinger.com, accounts.intuit.com,
    tagmanager.google.com) while the brand's own domain obviously does - and the
    brand mark is what makes a tile recognizable, which is the entire point of
    showing a logo. Only one level up, and never for a bare IP."""
    out = [domain]
    labels = domain.split(".")
    if len(labels) > 2 and not re.fullmatch(r"[\d.]+", domain):
        parent = ".".join(labels[-2:])
        if parent != domain:
            out.append(parent)
    return out


def _icon_resolvers(domain: str):
    """Ordered logo sources, across the hostname and its parent. icon.horse
    serves a site's real high-res mark; Google's faviconV2 is the second pass
    for domains it doesn't have. `fallback_opts` is deliberately empty so an
    unknown domain 404s rather than returning Google's generic globe glyph as
    though it were a real logo."""
    out = []
    for d in _icon_domain_candidates(domain):
        out.append(("icon.horse", f"https://icon.horse/icon/{d}"))
        out.append(("google", "https://t1.gstatic.com/faviconV2?client=SOCIAL&type=FAVICON"
                              f"&fallback_opts=&url=https://{d}&size=128"))
    return out


def _known_icon_domains(db: Session) -> set:
    """Hostnames that appear on a link somebody has actually saved."""
    out = set()
    for (url,) in db.query(models.ExternalLink.url).all():
        try:
            h = (urlparse(url).hostname or "").lower().strip(".")
        except ValueError:
            continue
        if h:
            out.add(h)
    return out


def _is_icon_placeholder(r) -> bool:
    """icon.horse answers a domain it has NO logo for with HTTP 200 and a
    GENERATED letter-avatar, not a 404 - so status alone can't tell a real brand
    mark from a stand-in, and the stand-in is what made unrecognized links look
    like they were "missing their image". It does distinguish them in one place:
    a real logo is served s-maxage=2592000 (30 days), a generated fallback
    s-maxage=300, because it wants to re-check the moment the site publishes a
    real favicon. Verified across both sets (Aug 31).

    Deliberately conservative - only a value we positively parsed AND that is
    short counts as a placeholder. If they ever drop or rename the header we go
    back to showing the stand-in, which is the status quo; the opposite default
    would silently strip real logos off every tile."""
    m = re.search(r"s-maxage=(\d+)", r.headers.get("cache-control", ""), re.I)
    return bool(m) and int(m.group(1)) < 86400


def _fetch_icon(domain: str) -> dict:
    """Try each resolver in turn. Returns the first real image, or a 'none'
    result that is cached so a dead lookup isn't retried on every render."""
    last_err = ""
    for source, url in _icon_resolvers(domain):
        try:
            with httpx.Client(timeout=8.0, follow_redirects=True) as c:
                r = c.get(url)
            if r.status_code != 200:
                last_err = f"{source}: HTTP {r.status_code}"
                continue
            ctype = (r.headers.get("content-type") or "").split(";")[0].strip().lower()
            if not ctype.startswith("image/") or not r.content:
                last_err = f"{source}: not an image ({ctype or 'no content-type'})"
                continue
            if len(r.content) > _ICON_MAX_BYTES:
                last_err = f"{source}: {len(r.content)} bytes is too large"
                continue
            if source == "icon.horse" and _is_icon_placeholder(r):
                # A generated stand-in, not this brand's mark. Fall through, and
                # if nothing better turns up the client draws its own curated
                # glyph - which at least reads as deliberate.
                last_err = f"{source}: generated placeholder, not a real logo"
                continue
            return {"data": r.content, "content_type": ctype, "source": source,
                    "size_bytes": len(r.content), "error": ""}
        except Exception as e:                    # noqa: BLE001 - any transport failure is just "try the next one"
            last_err = f"{source}: {type(e).__name__}"
    return {"data": None, "content_type": "", "source": "none", "size_bytes": 0, "error": last_err[:300]}


def _icon_is_fresh(row, now: datetime) -> bool:
    try:
        age = now - datetime.strptime((row.fetched_at or "")[:19], "%Y-%m-%dT%H:%M:%S").replace(tzinfo=timezone.utc)
    except (TypeError, ValueError):
        return False
    return age.days < (_ICON_RETRY_DAYS if row.source == "none" else _ICON_TTL_DAYS)


@router.get("/icon")
def link_icon(d: str, db: Session = Depends(get_db)):
    """The logo for one domain, from our cache. 404 when the domain genuinely
    has no logo anywhere - that is what lets the client fall through to its own
    curated glyph instead of showing a stand-in globe."""
    domain = (d or "").lower().strip().strip(".")
    if not domain or len(domain) > 253 or not re.fullmatch(r"[a-z0-9.\-]+", domain):
        raise HTTPException(400, "Bad domain")
    if domain not in _known_icon_domains(db):
        raise HTTPException(404, "Not a domain on any link here")

    now = datetime.now(timezone.utc)
    row = db.query(models.LinkIcon).filter(models.LinkIcon.domain == domain).first()
    if row is None or not _icon_is_fresh(row, now):
        got = _fetch_icon(domain)
        if row is None:
            row = models.LinkIcon(domain=domain)
            db.add(row)
        # A refresh that failed keeps the logo we already had - a resolver having
        # a bad afternoon must not blank a tile that was working yesterday.
        if got["data"] or row.data is None:
            row.data = got["data"]
            row.content_type = got["content_type"]
            row.source = got["source"]
            row.size_bytes = got["size_bytes"]
        row.error = got["error"]
        row.fetched_at = now.strftime("%Y-%m-%dT%H:%M:%S")
        db.commit()

    if not row.data:
        raise HTTPException(404, "No logo found for that domain")
    return Response(
        content=row.data,
        media_type=row.content_type or "image/png",
        # Long but not immutable: the row can be refreshed after _ICON_TTL_DAYS,
        # and a browser holding it forever would never see a re-brand.
        headers={"Cache-Control": "public, max-age=604800"},
    )


@router.get("/import-template")
def import_template(user: dict = Depends(require_links_admin), db: Session = Depends(get_db)):
    """Excel import template (Manage > Import > Export Template, Aug 14 -
    "department and category column in excel but multi selectable and in
    dropdown"). A single Excel cell can't hold a native multi-select
    dropdown without an embedded VBA macro (.xlsm, triggers Excel's "Enable
    Content" security prompt, and some IT policies block macros outright) -
    so instead this gives 3 Department columns and 3 Category columns, each
    a plain single-select Data Validation dropdown sourced from the live
    taxonomy. Picking more than one department/category for a link is just
    filling more than one of those columns; import_external_links below
    already unions them (via _clean_list) same as it would 3 typed values.
    Options list off a hidden sheet, not an inline comma list, since Excel's
    inline-list validation formula is capped around 255 characters and the
    taxonomy will outgrow that."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    rows = db.query(models.ExternalLinkTaxonomy).order_by(
        models.ExternalLinkTaxonomy.kind, models.ExternalLinkTaxonomy.sort_order, models.ExternalLinkTaxonomy.name
    ).all()
    departments = [r.name for r in rows if r.kind == "department"] or ["General"]
    categories = [r.name for r in rows if r.kind == "category"] or ["Imported"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Links"
    headers = ["Name", "URL", "Company", "Description", "Pinned",
               "Department 1", "Department 2", "Department 3",
               "Category 1", "Category 2", "Category 3"]
    hfill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
    ws.append(["ADP", "https://adp.com", "Greens", "Payroll processing", False, "Accounting", "", "", "Finance", "", ""])
    ws.append(["Slack", "https://slack.com", "", "Team chat", False, "", "", "", "", "", ""])
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    lists = wb.create_sheet("Lists")
    lists.sheet_state = "hidden"
    lists["A1"] = "Department"
    lists["B1"] = "Category"
    for i, d in enumerate(departments, start=2):
        lists.cell(row=i, column=1, value=d)
    for i, c in enumerate(categories, start=2):
        lists.cell(row=i, column=2, value=c)

    max_rows = 500
    dept_dv = DataValidation(type="list", formula1=f"=Lists!$A$2:$A${1 + len(departments)}", allow_blank=True)
    cat_dv = DataValidation(type="list", formula1=f"=Lists!$B$2:$B${1 + len(categories)}", allow_blank=True)
    ws.add_data_validation(dept_dv)
    ws.add_data_validation(cat_dv)
    for col in ("F", "G", "H"):
        dept_dv.add(f"{col}2:{col}{max_rows}")
    for col in ("I", "J", "K"):
        cat_dv.add(f"{col}2:{col}{max_rows}")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=external-links-import-template.xlsx"},
    )


@router.post("/import")
def import_external_links(payload: ImportRequest, user: dict = Depends(require_links_admin), db: Session = Depends(get_db)):
    """Batch CSV import (Manage > Import) for standing up a department's
    directory in one paste instead of one Add Link modal at a time. Best-
    effort: valid rows are created, invalid ones are reported back by 1-based
    row number and skipped - a typo in row 12 shouldn't lose rows 1-11.

    Department/category arrive as arrays now (Aug 14, "select multiple
    department or category by selecting" - the import preview table picks
    these per row with the same checkbox dropdown as Add/Edit, not typed
    CSV text), same shape `update_external_link` takes. A row with no
    category defaults to "Imported" so it still groups with its siblings
    instead of failing import - icon is still left off the sheet on purpose
    (Neil, Aug 12): it has to match an internal key, not something to fill
    in by hand. `company` is typed as a company NAME (e.g. "Greens India"),
    not the HrEntity.id a human has no reason to know - resolved to the id
    here, case-insensitively; an unmatched name is left blank (company-wide)
    rather than failing the whole row."""
    now = _now()
    company_by_name = {e.name.strip().lower(): e.id for e in db.query(models.HrEntity).all()}
    created, errors = [], []
    for i, row in enumerate(payload.rows, start=1):
        name, url = row.name.strip(), row.url.strip()
        if not name or not url:
            errors.append({"row": i, "name": name or f"Row {i}", "reason": "Name and URL are required"})
            continue
        if not re.match(r"^https?://", url, re.I):
            url = f"https://{url}"
        categories = _clean_list(row.categories) or ["Imported"]
        departments = _clean_list(row.departments)
        db_link = models.ExternalLink(
            name=name, url=url, category=categories[0], description=row.description.strip(),
            department=departments[0] if departments else "", categories=categories, departments=departments,
            company=company_by_name.get(row.company.strip().lower(), ""),
            icon=row.icon or "Link2", is_pinned=row.is_pinned,
            created_by=user["email"], created_at=now, updated_at=now,
        )
        db.add(db_link)
        db.flush()
        created.append(db_link)
    if created:
        db.add(models.AuditLog(
            timestamp=now, user_email=user["email"], user_role=user.get("role", ""),
            action="Bulk-imported external links", resource_type="external_links", resource_id="",
            details=json.dumps({"count": len(created), "names": [c.name for c in created]}), ip_address="",
        ))
    db.commit()
    for c in created:
        db.refresh(c)
    return {"created": created, "errors": errors}


# ─────────────────────────────────────────────────────────────────────────────
# Personal Links (Aug 12) - an employee's own day-to-day shortcuts, separate
# from the curated directory above. Private by construction: every query
# below filters on owner_email == the signed-in user, and update/delete
# additionally re-check ownership before touching a row - there is no
# manager/admin bypass, so even an owner-level account can't see or edit
# someone else's personal links through this API. No require_links_admin
# gate anywhere here on purpose: this is self-service, not a managed module.
# ─────────────────────────────────────────────────────────────────────────────
personal_router = APIRouter(prefix="/personal-links", tags=["Personal Links"], dependencies=[Depends(get_current_user)])


class PersonalLinkCreate(BaseModel):
    name: str
    url: str
    description: str = ""
    icon: str = "Link2"
    vault_cred_id: str = ""
    department: str = ""
    category: str = ""


class PersonalLinkUpdate(BaseModel):
    name: Optional[str] = None
    url: Optional[str] = None
    description: Optional[str] = None
    icon: Optional[str] = None
    vault_cred_id: Optional[str] = None
    department: Optional[str] = None
    category: Optional[str] = None


class PersonalReorderEntry(BaseModel):
    id: int
    sort_order: int


def _own_personal_link(link_id: int, user: dict, db: Session) -> models.PersonalLink:
    link = (
        db.query(models.PersonalLink)
        .filter(models.PersonalLink.id == link_id, models.PersonalLink.owner_email == user["email"])
        .first()
    )
    if not link:
        raise HTTPException(status_code=404, detail="Link not found")
    return link


@personal_router.get("")
def list_personal_links(user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    return (
        db.query(models.PersonalLink)
        .filter(models.PersonalLink.owner_email == user["email"])
        .order_by(models.PersonalLink.sort_order.asc(), models.PersonalLink.name.asc())
        .all()
    )


def _owned_vault_cred_id(vault_cred_id: str, user: dict, db: Session) -> str:
    """Guards against pointing a link at someone else's (or a deleted)
    Credential Vault personal credential - id is client-supplied, so this is
    the only thing standing between a typo/tamper and reveal() being asked
    to decrypt a row that was never this user's. Silently drops an invalid
    id rather than 400ing the whole save; the link itself is still valid
    without a credential attached."""
    if not vault_cred_id:
        return ""
    owns = (
        db.query(models.VaultPersonalCredential.id)
        .filter(models.VaultPersonalCredential.id == vault_cred_id, models.VaultPersonalCredential.owner_email == user["email"])
        .first()
    )
    return vault_cred_id if owns else ""


@personal_router.post("", status_code=201)
def create_personal_link(link: PersonalLinkCreate, user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    target = _normalize_url(link.url)
    existing = next(
        (
            l for l in db.query(models.PersonalLink.name, models.PersonalLink.url)
            .filter(models.PersonalLink.owner_email == user["email"]).all()
            if _normalize_url(l.url) == target
        ),
        None,
    )
    if existing:
        raise HTTPException(status_code=409, detail=f'This is already in your Personal Links as "{existing.name}".')
    now = _now()
    data = link.model_dump()
    data["vault_cred_id"] = _owned_vault_cred_id(data.get("vault_cred_id", ""), user, db)
    db_link = models.PersonalLink(**data, owner_email=user["email"], created_at=now, updated_at=now)
    db.add(db_link)
    db.commit()
    db.refresh(db_link)
    return db_link


@personal_router.patch("/reorder")
def reorder_personal_links(entries: list[PersonalReorderEntry], user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """MUST stay registered before PATCH /{link_id} below - same route-order
    gotcha as the main reorder endpoint above (see its comment)."""
    by_id = {e.id: e.sort_order for e in entries}
    rows = (
        db.query(models.PersonalLink)
        .filter(models.PersonalLink.id.in_(by_id.keys()), models.PersonalLink.owner_email == user["email"])
        .all()
    )
    for row in rows:
        row.sort_order = by_id[row.id]
    db.commit()
    return {"ok": True, "updated": len(rows)}


@personal_router.patch("/{link_id}")
def update_personal_link(link_id: int, patch: PersonalLinkUpdate, user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    db_link = _own_personal_link(link_id, user, db)
    changes = patch.model_dump(exclude_unset=True)
    if "vault_cred_id" in changes:
        changes["vault_cred_id"] = _owned_vault_cred_id(changes["vault_cred_id"] or "", user, db)
    for field, value in changes.items():
        setattr(db_link, field, value)
    db_link.updated_at = _now()
    db.commit()
    db.refresh(db_link)
    return db_link


@personal_router.delete("/{link_id}")
def delete_personal_link(link_id: int, user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    db_link = _own_personal_link(link_id, user, db)
    db.delete(db_link)
    db.commit()
    return {"ok": True}


@personal_router.patch("/{link_id}/click")
def click_personal_link(link_id: int, user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    db_link = _own_personal_link(link_id, user, db)
    db_link.clicks += 1
    db.commit()
    db.refresh(db_link)
    return db_link
