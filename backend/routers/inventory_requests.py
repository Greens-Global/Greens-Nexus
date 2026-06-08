import io
import os
import threading
import uuid
from datetime import datetime, timezone
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import text, or_, func
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional
import httpx
from database import get_db
from auth import get_current_user, require_level_or_module
from models import InventoryRequest, InventoryItem, NexusRole, AuditLog

# Valid predecessor statuses for each target status — guards against illegal
# jumps like pending -> allocated or re-approving an already-resolved request.
_VALID_TRANSITIONS = {
    "approved":  {"pending"},
    "rejected":  {"pending"},
    "allocated": {"approved"},
    "returned":  {"allocated"},
    "cancelled": {"pending", "approved"},
}

# Mirrors ROLE_LEVEL in routers/roles.py — duplicated here to avoid a cross-router
# import; both must stay in sync if role names ever change.
_ROLE_LEVEL = {"employee": 1, "supervisor": 2, "manager": 3, "administrator": 4, "owner": 5}

# Catalogue management, report export, and the audit log: Manager+ can always
# do these, OR an Access Group granting "inventory" at Editor level or above
# unlocks them for its members — without elevating anyone to Manager anywhere
# else in the app (mirrors a folder's "Editor: Download/Edit/Upload" grant).
require_inventory_admin = require_level_or_module(_ROLE_LEVEL["manager"], "inventory", "editor")

# Deletion stays at Owner by default, OR an Access Group granting "inventory"
# at Full level or above (mirrors "Full: Download/Edit/Upload/Delete").
require_inventory_delete = require_level_or_module(_ROLE_LEVEL["owner"], "inventory", "full")


def _title_case_email(email: str) -> str:
    """Fallback display name for accounts with no display_name on file yet —
    derives 'Sai Malladi' from 'sai.malladi@...' to match the org's
    firstname.lastname@ convention. Self-heals once Access Manager captures
    the real Microsoft Graph name on the next role assignment."""
    local = email.split("@", 1)[0]
    return " ".join(part.capitalize() for part in local.replace("_", ".").split(".") if part)

_DAMAGE_KEYWORDS = ("damaged", "broken", "cracked", "lost", "destroyed", "unusable", "retired")


def _reserve_stock(db: Session, item_id: str, qty: int) -> bool:
    """Atomically decrement available_qty, but only if enough stock remains.
    A single UPDATE...WHERE is atomic at the row level in Postgres — no explicit
    locking needed, and it can't race two simultaneous allocations into negative
    stock the way a SELECT-then-UPDATE would."""
    result = db.execute(
        text("UPDATE inventory_items SET available_qty = available_qty - :qty, "
             "last_updated = :now WHERE id = :id AND available_qty >= :qty"),
        {"qty": qty, "id": item_id, "now": datetime.now(timezone.utc).isoformat()},
    )
    return result.rowcount > 0


def _release_stock(db: Session, item_id: str, qty: int, damaged: bool) -> None:
    """Atomically restore stock on return. A damaged/lost/retired unit never
    re-enters circulation — it comes off both available_qty and total_qty."""
    now = datetime.now(timezone.utc).isoformat()
    if damaged:
        db.execute(
            text("UPDATE inventory_items SET total_qty = GREATEST(total_qty - :qty, 0), "
                 "last_updated = :now WHERE id = :id"),
            {"qty": qty, "id": item_id, "now": now},
        )
    else:
        db.execute(
            text("UPDATE inventory_items SET "
                 "available_qty = LEAST(available_qty + :qty, total_qty), "
                 "last_updated = :now WHERE id = :id"),
            {"qty": qty, "id": item_id, "now": now},
        )

_SUPABASE_URL = os.getenv("SUPABASE_URL", "")
_SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY", "")


def _post_inventory_event(request_id: str, status: str, affected_email: str) -> None:
    try:
        httpx.post(
            f"{_SUPABASE_URL}/rest/v1/inventory_events",
            headers={
                "apikey": _SUPABASE_SERVICE_KEY,
                "Authorization": f"Bearer {_SUPABASE_SERVICE_KEY}",
                "Content-Type": "application/json",
                "Prefer": "return=minimal",
            },
            json={
                "request_id": request_id,
                "status": status,
                "affected_email": affected_email,
            },
            timeout=5.0,
        )
    except Exception:
        pass


def _fire_inventory_event(request_id: str, status: str, affected_email: str) -> None:
    """Insert a row into inventory_events so Supabase Realtime notifies clients.
    Fire-and-forget: runs on a background thread so a slow/down Supabase never
    holds the request thread (and its checked-out DB connection) hostage for
    up to 5 seconds — that was compounding pool contention under load."""
    if not _SUPABASE_URL or not _SUPABASE_SERVICE_KEY:
        return
    threading.Thread(
        target=_post_inventory_event,
        args=(request_id, status, affected_email),
        daemon=True,
    ).start()

router = APIRouter(prefix="/inventory-requests", tags=["inventory-requests"], dependencies=[Depends(get_current_user)])


class RequestIn(BaseModel):
    id:                  str
    item_id:             str
    item_name:           str
    requested_by:        str
    requested_by_email:  str = ""
    raised_by:           str
    department:          str
    quantity:            int = 1
    days:                int = 1
    reason:              str = ""


class StatusUpdate(BaseModel):
    status:                    str
    resolved_by:               Optional[str] = ""
    reject_reason:             Optional[str] = ""
    assigned_allocator_email:  Optional[str] = ""
    assigned_allocator_name:   Optional[str] = ""
    allocated_by:              Optional[str] = ""
    return_photo_name:         Optional[str] = ""
    return_photo_url:          Optional[str] = ""
    condition_note:            Optional[str] = ""


def _to_dict(r: InventoryRequest) -> dict:
    return {
        "id":                r.id,
        "itemId":            r.item_id,
        "itemName":          r.item_name,
        "requestedBy":       r.requested_by,
        "requestedByEmail":  r.requested_by_email,
        "raisedBy":          r.raised_by,
        "department":        r.department,
        "quantity":          r.quantity,
        "days":              r.days,
        "reason":            r.reason,
        "status":            r.status,
        "createdAt":         r.created_at,
        "resolvedAt":        r.resolved_at  or None,
        "resolvedBy":        r.resolved_by  or None,
        "rejectReason":      r.reject_reason or None,
        "assignedAllocatorEmail": r.assigned_allocator_email or None,
        "assignedAllocatorName":  r.assigned_allocator_name  or None,
        "allocatedAt":       r.allocated_at  or None,
        "allocatedBy":       r.allocated_by  or None,
        "returnedAt":        r.returned_at   or None,
        "returnPhotoName":   r.return_photo_name or None,
        "returnPhotoUrl":    r.return_photo_url  or None,
        "conditionNote":     r.condition_note    or None,
    }


def _item_to_dict(i: InventoryItem) -> dict:
    return {
        "id":         i.id,
        "name":       i.name,
        "category":   i.category,
        "department": i.department,
        "location":   i.location,
        "available":  i.available_qty,
        "total":      i.total_qty,
    }


@router.get("/items")
def list_items(db: Session = Depends(get_db)):
    """Live stock levels — the single source of truth for available/total counts.
    Replaces the static mock data the frontend used to carry locally."""
    rows = db.query(InventoryItem).order_by(InventoryItem.department, InventoryItem.name).all()
    return [_item_to_dict(i) for i in rows]


class ItemImportRow(BaseModel):
    name:       str
    category:   Optional[str] = ""
    department: Optional[str] = ""
    location:   Optional[str] = ""
    total_qty:  int = 0


class ItemImportRequest(BaseModel):
    items: list[ItemImportRow]


@router.post("/items/import")
def import_items(body: ItemImportRequest, user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """Bulk-creates or updates catalogue items from a parsed spreadsheet.
    Matches existing items by name (case-insensitive): an existing item's
    total_qty is replaced and available_qty is shifted by the same delta, so
    units already out on loan stay accounted for. Unmatched rows become new
    items with available_qty seeded to the imported total."""
    now = datetime.now(timezone.utc).isoformat()
    created = updated = skipped = 0

    for row in body.items:
        name = row.name.strip()
        if not name:
            skipped += 1
            continue
        total = max(row.total_qty, 0)
        category   = (row.category or "").strip()
        department = (row.department or "").strip()
        location   = (row.location or "").strip()

        existing = db.query(InventoryItem).filter(func.lower(InventoryItem.name) == name.lower()).first()
        if existing:
            delta = total - existing.total_qty
            existing.total_qty     = total
            existing.available_qty = min(max(existing.available_qty + delta, 0), total)
            if category:   existing.category   = category
            if department: existing.department = department
            if location:   existing.location   = location
            existing.last_updated  = now
            updated += 1
        else:
            db.add(InventoryItem(
                id=str(uuid.uuid4()),
                name=name,
                category=category,
                department=department,
                location=location,
                total_qty=total,
                available_qty=total,
                last_updated=now,
            ))
            created += 1

    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped}


class ItemUpdate(BaseModel):
    name:       Optional[str] = None
    category:   Optional[str] = None
    department: Optional[str] = None
    location:   Optional[str] = None
    total_qty:  Optional[int] = None


class ItemCreate(BaseModel):
    name:       str
    category:   Optional[str] = ""
    department: Optional[str] = ""
    location:   Optional[str] = ""
    total_qty:  int = 0


@router.post("/items", status_code=201)
def create_item(body: ItemCreate, user: dict = Depends(require_inventory_admin), db: Session = Depends(get_db)):
    """Manually add a new catalogue item — the friendly counterpart to bulk CSV
    import, for managers who just need to add one or two new things. Same
    duplicate-name guard as update_item so the catalogue can't end up with two
    entries for the same thing via the two different creation paths."""
    name = body.name.strip()
    if not name:
        raise HTTPException(400, "Name cannot be empty")
    dupe = db.query(InventoryItem).filter(func.lower(InventoryItem.name) == name.lower()).first()
    if dupe:
        raise HTTPException(409, f"An item named '{name}' already exists")

    total = max(body.total_qty, 0)
    now = datetime.now(timezone.utc).isoformat()
    item = InventoryItem(
        id=str(uuid.uuid4()),
        name=name,
        category=(body.category or "").strip(),
        department=(body.department or "").strip(),
        location=(body.location or "").strip(),
        total_qty=total,
        available_qty=total,
        last_updated=now,
    )
    db.add(item)
    db.commit()
    return _item_to_dict(item)


@router.patch("/items/{item_id}")
def update_item(item_id: str, body: ItemUpdate, user: dict = Depends(require_inventory_admin), db: Session = Depends(get_db)):
    """Edit a catalogue item's name/category/department/location/stock total.
    Manager and above — same bar as approving requests, since this affects what
    the whole org sees and can request. Changing total_qty shifts available_qty
    by the same delta, mirroring the bulk-import logic so units on loan stay
    accounted for."""
    item = db.query(InventoryItem).filter(InventoryItem.id == item_id).first()
    if not item:
        raise HTTPException(404, "Item not found")

    if body.name is not None:
        name = body.name.strip()
        if not name:
            raise HTTPException(400, "Name cannot be empty")
        dupe = db.query(InventoryItem).filter(
            func.lower(InventoryItem.name) == name.lower(), InventoryItem.id != item_id
        ).first()
        if dupe:
            raise HTTPException(409, f"Another item is already named '{name}'")
        item.name = name
    if body.category is not None:
        item.category = body.category.strip()
    if body.department is not None:
        item.department = body.department.strip()
    if body.location is not None:
        item.location = body.location.strip()
    if body.total_qty is not None:
        total = max(body.total_qty, 0)
        delta = total - item.total_qty
        item.total_qty     = total
        item.available_qty = min(max(item.available_qty + delta, 0), total)

    item.last_updated = datetime.now(timezone.utc).isoformat()
    db.commit()
    return _item_to_dict(item)


@router.delete("/items/{item_id}")
def delete_item(item_id: str, user: dict = Depends(require_inventory_delete), db: Session = Depends(get_db)):
    """Permanently remove a catalogue item — Global Admin only, matching the
    'Permanently Delete Records' policy. Blocked while any request against the
    item is still pending/approved/allocated, so history never points at a void."""
    item = db.query(InventoryItem).filter(InventoryItem.id == item_id).first()
    if not item:
        raise HTTPException(404, "Item not found")

    active = db.query(InventoryRequest).filter(
        InventoryRequest.item_id == item_id,
        InventoryRequest.status.in_(["pending", "approved", "allocated"]),
    ).count()
    if active:
        raise HTTPException(409, "Can't delete an item with pending or active requests against it")

    db.delete(item)
    db.commit()
    return {"ok": True}


@router.get("/allocators")
def list_allocators(user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """Everyone at supervisor level or above — the pool a manager can hand an
    approved request off to for physical allocation. Manager-accessible (not
    gated behind require_administrator like GET /roles, which is too high a
    bar for the person who actually needs this list)."""
    if user["level"] < _ROLE_LEVEL["manager"]:
        raise HTTPException(403, "Manager or above required to view the allocator list")
    rows = db.query(NexusRole).filter(NexusRole.role.in_(
        [role for role, level in _ROLE_LEVEL.items() if level >= _ROLE_LEVEL["supervisor"]]
    )).order_by(NexusRole.email).all()
    return [
        {"email": r.email, "name": r.display_name or _title_case_email(r.email), "role": r.role}
        for r in rows
    ]


@router.get("")
def list_requests(email: Optional[str] = None, user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    q = db.query(InventoryRequest).order_by(InventoryRequest.created_at.desc())
    if user["level"] < 3:
        # Non-managers see their own requests, plus anything assigned to them
        # to physically allocate — both are "their business" regardless of
        # who raised the request.
        q = q.filter(or_(
            InventoryRequest.requested_by_email == user["email"],
            InventoryRequest.assigned_allocator_email == user["email"],
        ))
    elif email:
        # Managers can optionally filter by a specific email
        q = q.filter(InventoryRequest.requested_by_email == email.lower())
    return [_to_dict(r) for r in q.all()]


@router.post("")
def create_request(body: RequestIn, db: Session = Depends(get_db)):
    if db.query(InventoryRequest).filter(InventoryRequest.id == body.id).first():
        raise HTTPException(409, "A request with this id already exists")

    now = datetime.now(timezone.utc).isoformat()
    row = InventoryRequest(
        id=body.id,
        item_id=body.item_id,
        item_name=body.item_name,
        requested_by=body.requested_by,
        requested_by_email=body.requested_by_email.lower(),
        raised_by=body.raised_by,
        department=body.department,
        quantity=body.quantity,
        days=body.days,
        reason=body.reason,
        status="pending",
        created_at=now,
    )
    db.add(row)
    db.commit()
    _fire_inventory_event(row.id, "pending", row.requested_by_email or "")
    return _to_dict(row)


@router.patch("/{req_id}")
def update_request(req_id: str, body: StatusUpdate, user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    row = db.query(InventoryRequest).filter(InventoryRequest.id == req_id).first()
    if not row:
        return {"ok": False, "error": "not found"}

    if body.status in ("approved", "rejected") and user["level"] < 3:
        raise HTTPException(403, "Manager or above required to approve or reject requests")
    if body.status == "approved" and not (body.assigned_allocator_email or "").strip():
        raise HTTPException(400, "Pick who should allocate this item before approving")
    if body.status == "allocated":
        # Only the person the manager assigned can hand the item over physically —
        # managers/admins keep an override for when the assignee is unavailable.
        is_assignee = row.assigned_allocator_email and row.assigned_allocator_email.lower() == user["email"]
        if not is_assignee and user["level"] < 3:
            raise HTTPException(403, "Only the assigned allocator (or a manager) can mark this as allocated")
    if body.status == "returned" and user["level"] < 2 and row.requested_by_email.lower() != user["email"]:
        raise HTTPException(403, "You can only return your own items")
    if body.status == "cancelled" and row.requested_by_email.lower() != user["email"]:
        raise HTTPException(403, "You can only cancel your own requests")

    valid_predecessors = _VALID_TRANSITIONS.get(body.status)
    if valid_predecessors is not None and row.status not in valid_predecessors:
        raise HTTPException(409, f"Cannot move a '{row.status}' request to '{body.status}'")

    now = datetime.now(timezone.utc).isoformat()

    if body.status in ("approved", "rejected"):
        row.resolved_at = now
        row.resolved_by = body.resolved_by or ""
        if body.status == "approved":
            row.assigned_allocator_email = (body.assigned_allocator_email or "").lower().strip()
            row.assigned_allocator_name  = (body.assigned_allocator_name  or "").strip()
        if body.status == "rejected":
            row.reject_reason = body.reject_reason or ""

    elif body.status == "cancelled":
        row.resolved_at = now
        row.resolved_by = body.resolved_by or ""

    elif body.status == "allocated":
        # Reserve stock atomically before flipping status — if there isn't enough
        # available, fail the whole request rather than allocating phantom units.
        if not _reserve_stock(db, row.item_id, row.quantity):
            db.rollback()
            raise HTTPException(409, "Not enough stock available to allocate this request")
        row.allocated_at = now
        row.allocated_by = body.allocated_by or ""

    elif body.status == "returned":
        note = (body.condition_note or "").lower()
        damaged = any(k in note for k in _DAMAGE_KEYWORDS)
        _release_stock(db, row.item_id, row.quantity, damaged)
        row.returned_at       = now
        row.return_photo_name = body.return_photo_name or ""
        row.return_photo_url  = body.return_photo_url  or ""
        row.condition_note    = body.condition_note    or ""

    row.status = body.status
    db.commit()
    _fire_inventory_event(req_id, row.status, row.requested_by_email or "")
    return _to_dict(row)


_REPORT_HEADERS = [
    "Item", "Category", "Department", "Location", "Requested By", "Requested By Email",
    "Quantity", "Status", "Requested Date", "Allocated Date", "Allocated By",
    "Returned Date", "Condition Note",
]


def _report_rows(db: Session, *, user_email, location, department, category, status):
    """Joins requests to their catalogue item (loose string-key coupling, same as
    elsewhere in this router) so location/category — which only live on the item —
    can be filtered and displayed alongside per-request details."""
    q = db.query(InventoryRequest, InventoryItem).outerjoin(
        InventoryItem, InventoryRequest.item_id == InventoryItem.id
    ).order_by(InventoryRequest.created_at.desc())

    if user_email:
        q = q.filter(InventoryRequest.requested_by_email == user_email.lower().strip())
    if department:
        q = q.filter(InventoryRequest.department == department)
    if status:
        q = q.filter(InventoryRequest.status == status)
    if location:
        q = q.filter(func.lower(InventoryItem.location) == location.lower().strip())
    if category:
        q = q.filter(InventoryItem.category == category)

    rows = []
    for r, item in q.all():
        rows.append([
            r.item_name, item.category if item else "", r.department, item.location if item else "",
            r.requested_by, r.requested_by_email, r.quantity, r.status,
            r.created_at[:10] if r.created_at else "",
            r.allocated_at[:10] if r.allocated_at else "", r.allocated_by or "",
            r.returned_at[:10] if r.returned_at else "", r.condition_note or "",
        ])
    return rows


@router.get("/report")
def export_inventory_report(
    format: str = "excel",
    user_email: Optional[str] = None,
    location: Optional[str] = None,
    department: Optional[str] = None,
    category: Optional[str] = None,
    status: Optional[str] = None,
    user: dict = Depends(require_inventory_admin),
    db: Session = Depends(get_db),
):
    """Filterable asset-status report, exportable as Excel or PDF — manager and
    above, matching the bar for the catalogue-wide CSV export."""
    rows = _report_rows(db, user_email=user_email, location=location, department=department,
                        category=category, status=status)
    stamp = datetime.utcnow().strftime("%Y%m%d")

    if format == "pdf":
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
        except ImportError:
            raise HTTPException(500, "reportlab not installed — run: pip install reportlab")

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title="Inventory Asset Status Report")
        table = Table([_REPORT_HEADERS] + rows, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A1A2E")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 6.5),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5FA")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        doc.build([table])
        buf.seek(0)
        filename = f"inventory_report_{stamp}.pdf"
        return StreamingResponse(buf, media_type="application/pdf",
                                 headers={"Content-Disposition": f"attachment; filename={filename}"})

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed — run: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Asset Status Report"

    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(_REPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"inventory_report_{stamp}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/audit-log")
def inventory_audit_log(
    q: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
    user: dict = Depends(require_inventory_admin),
    db: Session = Depends(get_db),
):
    """Manager-accessible, inventory-scoped slice of the audit log — searchable
    by item name/id, user, or action. Deliberately separate from the admin-only
    system-wide GET /audit-logs (level >= 4) so managers can see catalogue
    history without exposing other modules' audit data."""
    query = db.query(AuditLog).filter(AuditLog.resource_type == "inventory-requests")
    if q:
        needle = q.strip()
        query = query.filter(or_(
            AuditLog.details.contains(needle),
            AuditLog.user_email.contains(needle.lower()),
            AuditLog.action.contains(needle),
            AuditLog.resource_id.contains(needle),
        ))
    total = query.count()
    rows = query.order_by(AuditLog.timestamp.desc()).offset(offset).limit(limit).all()
    return {
        "total": total,
        "rows": [
            {
                "id":            r.id,
                "timestamp":     r.timestamp,
                "user_email":    r.user_email,
                "user_role":     r.user_role,
                "action":        r.action,
                "resource_id":   r.resource_id,
                "details":       r.details,
            }
            for r in rows
        ],
    }
