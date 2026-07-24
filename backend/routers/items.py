import io
import json
import os
import re
import threading
import uuid
from collections import defaultdict, deque
from datetime import datetime, timezone
from fastapi import APIRouter, Depends, HTTPException, Response
from fastapi.responses import StreamingResponse
from sqlalchemy import or_, func, text as sa_text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional
import httpx
from database import get_db
from auth import get_current_user, require_level_or_module
from models import Item, ItemCheckout, ItemCartEntry, ItemAssignment, ItemCustomField, ItemType, NexusRole, NexusNotification, AuditLog, NexusEmployee

_VALID_TRANSITIONS = {
    "approved":         {"pending"},
    "rejected":         {"pending", "approved"},
    "pending_receipt":  {"approved"},
    "allocated":        {"approved", "pending_receipt"},
    "returned":         {"allocated"},
    # P1-3: a stuck pending_receipt (allocator handed over, employee never confirmed
    # receipt) previously had NO exit — allow it to be cancelled (manager-guarded in
    # update_checkout / the reconcile-state endpoint).
    "cancelled":        {"pending", "approved", "rejected", "pending_receipt"},
}

_ROLE_LEVEL = {"employee": 1, "supervisor": 2, "manager": 3, "administrator": 4, "owner": 5}

# Controlled list of item types — add/edit is dropdown-only, and an unrecognised
# type on IMPORT now lands in "Other" for cleanup rather than spawning a junk type
# (Neil). To add a real new type, add it here (a deliberate, rare change).
_ITEM_TYPES    = ["Computer", "Peripheral", "Networking", "Server", "Storage",
                  "IP Camera", "Devices", "Tools", "Equipment", "Vehicles",
                  "Furniture", "Keys", "Other"]
_ITEM_STATUSES = ["available", "checked_out", "permanently_assigned", "retired"]
# Operational status (Neil) — what condition/deployment state the unit is in. SEPARATE
# from the lifecycle `status` above (which is auto-driven by checkouts/assignments).
# '' = unset. Mirror _OP_STATUSES on the frontend if you change this list.
_OP_STATUSES   = ["deployed", "in_storage", "in_repair", "needs_replacement", "retired", "lost"]
# Op statuses that are declared AGAINST a person — they capture an op_status_person
# and notify that person. Anything else clears the person. Mirror _OP_STATUS_PERSON
# on the frontend (OP_STATUS_PERSON) if you change this set.
_OP_STATUS_PERSON = {"lost", "retired"}
_CUSTOM_FIELD_TYPES = ["text", "number", "date", "select", "boolean", "url"]
# Soft-deleted items are restorable for this many days, then purged for good.
_RECYCLE_BIN_DAYS = 30

_TYPE_DEFAULT_OWNER = {
    "Devices":   "IT",
    "Tools":     "Construction (MCD)",
    "Vehicles":  "Construction",
    "Equipment": "",
    "Keys":      "Operations (Oversite)",
    "Other":     "",
}

# P1-13: return condition is now an explicit enum on the return payload instead of
# sniffing the free-text note for damage keywords (which false-matched "undamaged",
# "not broken", etc. and wrongly retired items). Only these values alter op_status.
_RETURN_CONDITIONS = ("ok", "damaged", "lost")

require_items_admin  = require_level_or_module(_ROLE_LEVEL["manager"], "inventory", "editor")
require_items_delete = require_level_or_module(_ROLE_LEVEL["owner"],   "inventory", "full")


def _is_items_manager(user: dict, db) -> bool:
    """Inline version of require_items_admin for checks INSIDE handlers:
    global manager+ OR an Access-Group grant of inventory at editor or above.
    The raw `user["level"] >= 3` checks this replaces ignored module grants, so
    an Employee-tier person with 'Item Management: Full' was blocked from
    approve/allocate/manage despite the grant (Jul 24 — India Admin Team)."""
    from auth import _module_level, _MODULE_LEVEL_RANK
    return (user.get("level", 1) >= _ROLE_LEVEL["manager"]
            or _module_level(user["email"], "inventory", db) >= _MODULE_LEVEL_RANK["editor"])

_SUPABASE_URL         = os.getenv("SUPABASE_URL", "")
_SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY", "")

# Valid photo URL prefix — only allow URLs pointing to our own Supabase storage
# so fake/external evidence cannot be submitted.
_STORAGE_PREFIX = f"{_SUPABASE_URL}/storage/v1/object/public/" if _SUPABASE_URL else None


def _validate_photo_url(url: Optional[str], field: str) -> None:
    """Raise 400 if url is non-empty and does not originate from our storage bucket."""
    if not url or not url.strip():
        return
    if _STORAGE_PREFIX and not url.startswith(_STORAGE_PREFIX):
        raise HTTPException(400, f"{field} must be a Supabase storage URL")


def _notify(db: Session, *, type: str, recipient: str, title: str, body: str,
            ref_id: str = "", item_name: str = "", requested_by: str = "") -> None:
    row = NexusNotification(
        id=str(uuid.uuid4()),
        type=type,
        recipient=recipient.lower() if recipient else "",
        title=title,
        body=body,
        ref_id=ref_id,
        item_name=item_name,
        requested_by=requested_by,
        action="",
        actioned=False,
        read_by="",
        created_at=datetime.now(timezone.utc).isoformat(),
    )
    db.add(row)


def _notify_op_status_declaration(db: Session, *, op_status: str, person_email: str,
                                  person_name: str, item_name: str) -> None:
    """Tell the person an op_status was declared against. Lost carries a friendly
    keep-it-safe disclaimer; other person-statuses (retired) get a neutral note.
    Server-side only (employees can't POST notifications) — one per declaration."""
    email = (person_email or "").lower().strip()
    if not email or op_status not in _OP_STATUS_PERSON:
        return
    if op_status == "lost":
        title = "An item was reported lost under your name"
        body  = (f"{item_name} has been recorded as Lost by you. Please take extra "
                 "care to keep your assigned equipment safe going forward — and if "
                 "this was logged by mistake, let your manager know.")
    else:  # retired
        title = "An item was retired under your name"
        body  = (f"{item_name} has been retired under your name and taken out of "
                 "active service. No action needed.")
    _notify(db, type=f"op_status_{op_status}", recipient=email,
            title=title, body=body, item_name=item_name)


def _title_case_email(email: str) -> str:
    local = email.split("@", 1)[0]
    return " ".join(p.capitalize() for p in local.replace("_", ".").split(".") if p)


def _post_item_event(checkout_id: str, status: str, affected_email: str) -> None:
    try:
        httpx.post(
            f"{_SUPABASE_URL}/rest/v1/inventory_events",
            headers={
                "apikey": _SUPABASE_SERVICE_KEY,
                "Authorization": f"Bearer {_SUPABASE_SERVICE_KEY}",
                "Content-Type": "application/json",
                "Prefer": "return=minimal",
            },
            # affected_email deliberately blank: inventory_events is anon-readable
            # for realtime pings, so nothing personal may be written into it.
            # Clients never used the field — they refetch via the authed API.
            json={"request_id": checkout_id, "status": status, "affected_email": ""},
            timeout=5.0,
        )
    except Exception:
        pass


def _fire_item_event(checkout_id: str, status: str, affected_email: str) -> None:
    if not _SUPABASE_URL or not _SUPABASE_SERVICE_KEY:
        return
    threading.Thread(
        target=_post_item_event,
        args=(checkout_id, status, affected_email),
        daemon=True,
    ).start()


router = APIRouter(prefix="/items", tags=["items"], dependencies=[Depends(get_current_user)])


# ── Serialisers ───────────────────────────────────────────────────────────────

def _item_to_dict(i: Item) -> dict:
    return {
        "id":            i.id,
        "serialNumber":  i.serial_number or "",
        "name":          i.name,
        "itemType":      i.item_type,
        "make":          i.make,
        "model":         i.model,
        "year":          i.year,
        "department":    i.department,
        "defaultOwner":  i.default_owner,
        "ownershipType": i.ownership_type,
        "status":        i.status,
        "location":      i.location,
        "photoUrl":      i.photo_url,
        "createdBy":     i.created_by,
        "createdAt":     i.created_at,
        "assignedToEmail": i.assigned_to_email or "",
        "assignedToName":  i.assigned_to_name  or "",
        "assignedToLocation": i.assigned_to_location or "",
        "assignedAt":      i.assigned_at       or "",
        # NULL (pre-migration rows) must read as True — photos required by default
        "pictureRequired": True if i.picture_required is None else bool(i.picture_required),
        "assetValue":      float(i.asset_value or 0),
        "opStatus":        i.op_status or "",
        "opStatusPersonEmail": i.op_status_person_email or "",
        "opStatusPersonName":  i.op_status_person_name or "",
        "customFields":    i.custom_fields if isinstance(i.custom_fields, dict) else {},
        "deletedAt":       i.deleted_at or "",
        "deletedBy":       i.deleted_by or "",
        "deletedLocation": i.deleted_location or "",
    }


def _custom_field_to_dict(f: ItemCustomField) -> dict:
    return {
        "id":            f.id,
        "fieldKey":      f.field_key,
        "label":         f.label,
        "fieldType":     f.field_type,
        "options":       f.options if isinstance(f.options, list) else [],
        "appliesToType": f.applies_to_type or "",
        "sortOrder":     f.sort_order or 0,
    }


def _checkout_to_dict(c: ItemCheckout) -> dict:
    return {
        "id":                      c.id,
        "itemId":                  c.item_id,
        "itemName":                c.item_name,
        "itemType":                c.item_type,
        "requestedBy":             c.requested_by,
        "requestedByEmail":        c.requested_by_email,
        "raisedBy":                c.raised_by,
        "department":              c.department,
        "days":                    c.days,
        "reason":                  c.reason,
        "status":                  c.status,
        "createdAt":               c.created_at,
        "resolvedAt":              c.resolved_at              or None,
        "resolvedBy":              c.resolved_by              or None,
        "rejectReason":            c.reject_reason            or None,
        "assignedAllocatorEmail":  c.assigned_allocator_email or None,
        "assignedAllocatorName":   c.assigned_allocator_name  or None,
        "allocatedAt":             c.allocated_at             or None,
        "allocatedBy":             c.allocated_by             or None,
        "checkoutPhotoUrl":        c.checkout_photo_url       or None,
        "checkoutPhotoName":       c.checkout_photo_name      or None,
        "returnedAt":              c.returned_at              or None,
        "returnPhotoUrl":          c.return_photo_url         or None,
        "returnPhotoName":         c.return_photo_name        or None,
        "conditionNote":           c.condition_note           or None,
        "orderId":                 c.order_id                 or "",
        "handoverPhotoBy":         c.handover_photo_by        or "",
        "handoverBatch":           bool(c.handover_batch)     if c.handover_batch is not None else False,
        "receiptPhotoUrl":         c.receipt_photo_url        or None,
        "receiptPhotoName":        c.receipt_photo_name       or None,
        "handedOverAt":            c.handed_over_at           or None,
        "extensionDays":           c.extension_days           or 0,
        "extensionReason":         c.extension_reason         or "",
        "extensionStatus":         c.extension_status         or "",
        "approverEmail":           c.approver_email           or "",
        "approverName":            c.approver_name            or "",
    }


# ── Item CRUD ─────────────────────────────────────────────────────────────────

class ItemCreate(BaseModel):
    name:           str
    item_type:      Optional[str] = "Other"
    make:           Optional[str] = ""
    model:          Optional[str] = ""
    year:           Optional[str] = ""
    department:     Optional[str] = ""
    default_owner:  Optional[str] = ""
    ownership_type: Optional[str] = "transient"
    location:       Optional[str] = ""
    photo_url:      Optional[str] = ""
    picture_required: Optional[bool]  = True
    asset_value:      Optional[float] = 0
    op_status:        Optional[str]  = ""
    op_status_person_email: Optional[str] = ""
    op_status_person_name:  Optional[str] = ""
    custom_fields:    Optional[dict] = None


class ItemUpdate(BaseModel):
    name:           Optional[str] = None
    item_type:      Optional[str] = None
    make:           Optional[str] = None
    model:          Optional[str] = None
    year:           Optional[str] = None
    department:     Optional[str] = None
    default_owner:  Optional[str] = None
    ownership_type: Optional[str] = None
    # P1-3: lifecycle `status` is intentionally NOT writable here — it is derived
    # solely from checkout/assignment transitions (a raw write could strand an idle
    # item as "checked_out" and make it permanently un-requestable). Use the
    # reconcile-state endpoint to re-derive it. op_status (operational) stays writable.
    location:       Optional[str] = None
    photo_url:      Optional[str] = None
    picture_required: Optional[bool]  = None
    asset_value:      Optional[float] = None
    op_status:        Optional[str]  = None
    op_status_person_email: Optional[str] = None
    op_status_person_name:  Optional[str] = None
    custom_fields:    Optional[dict] = None


class ItemImportRow(BaseModel):
    name:           str
    serial_number:  Optional[str] = ""
    item_type:      Optional[str] = "Other"
    make:           Optional[str] = ""
    model:          Optional[str] = ""
    year:           Optional[str] = ""
    department:     Optional[str] = ""
    default_owner:  Optional[str] = ""
    ownership_type: Optional[str] = "transient"
    location:       Optional[str] = ""
    custom_fields:  Optional[dict] = None


class ItemImportRequest(BaseModel):
    items: list[ItemImportRow]


@router.get("")
def list_items(
    department: Optional[str] = None,
    item_type:  Optional[str] = None,
    status:     Optional[str] = None,
    db: Session = Depends(get_db),
):
    # Soft-deleted items (deleted_at set) are hidden from the normal catalogue —
    # they live in GET /deleted until restored or purged (Ankush). NULL = legacy row.
    q = db.query(Item).filter(or_(Item.deleted_at.is_(None), Item.deleted_at == "")) \
          .order_by(Item.department, Item.item_type, Item.name)
    if department:
        q = q.filter(Item.department == department)
    if item_type:
        q = q.filter(Item.item_type == item_type)
    if status:
        q = q.filter(Item.status == status)
    items = q.all()

    # Enrich each item with live checkout activity so ALL users (not just
    # managers who see every checkout) know when an item is taken or under
    # review — prevents submitting a cart that will 409 at the server.
    active_cos = db.query(
        ItemCheckout.item_id,
        ItemCheckout.requested_by,
        ItemCheckout.status,
        ItemCheckout.created_at,
        ItemCheckout.days,
        ItemCheckout.allocated_at,
        ItemCheckout.handed_over_at,
    ).filter(
        ItemCheckout.status.in_(["pending", "approved", "pending_receipt", "allocated"])
    ).all()
    active_map = {row.item_id: row for row in active_cos}

    result = []
    for i in items:
        d = _item_to_dict(i)
        co = active_map.get(i.id)
        # hasActiveRequest: true when a checkout blocks new requests (not yet allocated)
        d["hasActiveRequest"] = co is not None and co.status in ("pending", "approved", "pending_receipt")
        # activeRequestedBy / activeDueDate: for "In Use — [Name] — available in X days".
        # The clock starts at physical handover (allocated/handed-over), not at the
        # request — approval delay must not eat into the employee's checkout days.
        d["activeRequestedBy"] = co.requested_by if co else None
        if co and co.days:
            try:
                from datetime import timedelta
                start_iso = co.allocated_at or co.handed_over_at or co.created_at
                start = datetime.fromisoformat(start_iso.replace("Z", "+00:00"))
                due = start + timedelta(days=int(co.days))
                d["activeDueDate"] = due.isoformat()
            except Exception:
                d["activeDueDate"] = None
        else:
            d["activeDueDate"] = None
        result.append(d)
    return result


# ── Serial numbers ────────────────────────────────────────────────────────────
# Each physical unit carries a static, Nexus-assigned serial (GG-#####). It is the
# identity the CSV import upserts on — names are NOT unique (10 identical laptops),
# so keying on name silently merged distinct units into one row (Sai, Jun 2026).

_SERIAL_RE = re.compile(r"^GG-(\d+)$", re.IGNORECASE)


def _serial_start(db: Session) -> int:
    """Next free auto-serial number. Scanned ONCE per import — autoflush=False means
    rows added during the loop are not visible to a re-query, so callers keep a local
    counter and never re-scan mid-batch."""
    top = 0
    for (s,) in db.query(Item.serial_number).all():
        m = _SERIAL_RE.match((s or "").strip())
        if m:
            top = max(top, int(m.group(1)))
    return top + 1


def _fmt_serial(n: int) -> str:
    return f"GG-{n:05d}"


# Spreadsheets use "N/A", "-", "none" etc. to mean "no value" — store them as blank
# so they don't render as real data (e.g. a Model column showing "... N/A").
_NA_TOKENS = {"", "n/a", "na", "n.a.", "n.a", "none", "null", "nil", "-", "–", "—"}


def _clean_field(v) -> str:
    s = (v or "").strip()
    return "" if s.lower() in _NA_TOKENS else s


# A loose "key" for a type so spelling/case/plural variants collapse to ONE type —
# "Office", "office", "OFFICES " all key to "office" (Neil: be smart about it). The
# valid set is the manager-curated item_types table; a static fallback canon covers
# the rare case where the table can't be read.
def _type_key(s) -> str:
    s = re.sub(r"\s+", " ", (s or "").strip().lower())
    if len(s) > 3 and s.endswith("ies"):                                  # batteries → battery
        return s[:-3] + "y"
    if len(s) > 4 and s.endswith(("ches", "shes", "ses", "xes", "zes")):  # boxes → box
        return s[:-2]
    if len(s) > 2 and s.endswith("s") and not s.endswith("ss"):           # cameras → camera
        return s[:-1]
    return s

def _build_canon(names) -> dict:
    return {_type_key(t): t for t in names}

_TYPE_CANON = _build_canon(_ITEM_TYPES)           # static fallback


def _seed_types_if_empty(db: Session) -> None:
    if db.query(ItemType).first():
        return
    now = _now_iso()
    for i, t in enumerate(_ITEM_TYPES):
        db.add(ItemType(name=t, sort_order=i, created_by="system", created_at=now))
    db.commit()


def _type_names(db: Session) -> list:
    _seed_types_if_empty(db)
    return [r.name for r in db.query(ItemType).order_by(ItemType.sort_order, ItemType.name).all()]


def _type_canon(db: Session) -> dict:
    return _build_canon(_type_names(db))


_AI_TYPE_MODEL = "claude-opus-4-8"

def _ai_match_types(new_labels: list, existing: list) -> dict:
    """Map each NEW label to an EXISTING type when it's the same kind of item — a
    spelling/abbreviation/plural/case variant (e.g. "IP Cams" → "IP Camera") — else
    None (it becomes a new type). Distinct items stay distinct (Laptop ≠ Computer,
    CCTV ≠ IP Camera). Returns {} with no API key or on error, so the caller just
    creates new types. One call per import keeps it cheap."""
    key = os.getenv("ANTHROPIC_API_KEY", "")
    if not key or not new_labels:
        return {}
    prompt = (
        "You normalise item-TYPE labels for a physical asset inventory.\n\n"
        f"EXISTING TYPES: {json.dumps(existing)}\n"
        f"NEW LABELS: {json.dumps(new_labels)}\n\n"
        "For each NEW LABEL decide if it is merely a spelling / abbreviation / plural / "
        "case variant of an EXISTING TYPE — i.e. the SAME kind of physical item.\n"
        'SAME (map it): "IP Cams" = "IP Camera"; "Laptops" = "Laptop"; "cctv camera" = "IP Camera" only if no better match.\n'
        'DIFFERENT — DO NOT MERGE: "Laptop" vs "Computer"; "CCTV" vs "IP Camera"; "Server" vs "Storage". '
        "These are genuinely different items even if related.\n"
        "Be conservative: only map when it is clearly the SAME item. Return ONLY a JSON object "
        "mapping each new label to the EXACT existing-type text it matches, or null if it is new. No prose."
    )
    try:
        with httpx.Client(timeout=30) as client:
            r = client.post(
                "https://api.anthropic.com/v1/messages",
                headers={"x-api-key": key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
                json={"model": _AI_TYPE_MODEL, "max_tokens": 600, "messages": [{"role": "user", "content": prompt}]},
            )
            r.raise_for_status()
            data = r.json()
        txt = "".join(b.get("text", "") for b in data.get("content", []) if b.get("type") == "text").strip()
        m = re.search(r"\{.*\}", txt, re.S)   # tolerate fences / stray text
        out = json.loads(m.group(0)) if m else {}
        return {k: v for k, v in out.items() if isinstance(k, str)}
    except Exception as e:  # noqa: BLE001
        print(f"[items] AI type match failed: {e}")
        return {}


def _normalize_type(raw, canon=None) -> str:
    s = _clean_field(raw)
    if not s:
        return "Other"
    # Match by the loose key so office/offices/OFFICE all land on the same type.
    # Anything with no match funnels to "Other" (the Add/Edit dropdowns can't produce
    # unknowns; CSV import auto-creates them — see import_items).
    return (canon or _TYPE_CANON).get(_type_key(s), "Other")


def _content_sig(name, item_type, make, model, year, department, location, ownership) -> tuple:
    """Identity of a serial-less row by its descriptive content (case-insensitive).
    Re-importing the same file matches each row to one existing unit by this sig, so
    it updates in place instead of duplicating — while two genuinely identical units
    still occupy two slots in the multiset and never collapse."""
    return tuple((x or "").strip().lower()
                 for x in (name, item_type, make, model, year, department, location, ownership))


@router.post("", status_code=201)
def create_item(body: ItemCreate, response: Response, user: dict = Depends(require_items_admin), db: Session = Depends(get_db)):
    name = body.name.strip()
    if not name:
        raise HTTPException(400, "Name cannot be empty")
    # P1-14: catalogue photos must come from our Supabase storage — same guard the
    # checkout/assignment evidence paths use — so external URLs can't land in every
    # user's catalogue.
    _validate_photo_url(body.photo_url, "photo_url")
    now = datetime.now(timezone.utc).isoformat()
    # Retry the rare serial race: two simultaneous adds can compute the same next
    # GG-##### and collide on the unique index. Recompute and retry a few times.
    for attempt in range(5):
        item = Item(
            id=str(uuid.uuid4()),
            serial_number=_fmt_serial(_serial_start(db)),
            name=name,
            item_type=_normalize_type(body.item_type, _type_canon(db)),
            make=(body.make or "").strip(),
            model=(body.model or "").strip(),
            year=(body.year or "").strip(),
            department=(body.department or "").strip(),
            default_owner=(body.default_owner or "").strip(),
            ownership_type=(body.ownership_type or "transient").strip(),
            # Permanent items start AVAILABLE too — "permanently_assigned" only
            # happens via the assignment flow once a real person accepts it.
            # Auto-stamping it at creation made unassigned items show as assigned.
            status="available",
            location=(body.location or "").strip(),
            photo_url=(body.photo_url or "").strip(),
            created_by=user["email"],
            created_at=now,
            picture_required=True if body.picture_required is None else bool(body.picture_required),
            asset_value=float(body.asset_value or 0),
            op_status=(body.op_status or "").strip() if (body.op_status or "").strip() in _OP_STATUSES else "",
            op_status_person_email=(body.op_status_person_email or "").lower().strip() if (body.op_status or "").strip() in _OP_STATUS_PERSON else "",
            op_status_person_name=(body.op_status_person_name or "").strip() if (body.op_status or "").strip() in _OP_STATUS_PERSON else "",
            custom_fields=body.custom_fields if isinstance(body.custom_fields, dict) else {},
        )
        db.add(item)
        try:
            db.commit()
            if (item.op_status or "") in _OP_STATUS_PERSON and item.op_status_person_email:
                _notify_op_status_declaration(db, op_status=item.op_status, person_email=item.op_status_person_email,
                                              person_name=item.op_status_person_name or "", item_name=item.name)
                db.commit()
            # Stamp the new id so the audit middleware can record WHICH item was
            # added (the path has no id on a POST) — lets the audit log thread an
            # item's full history by id, not just by name.
            response.headers["X-Created-Id"] = item.id
            return _item_to_dict(item)
        except IntegrityError:
            db.rollback()
    raise HTTPException(409, "Could not assign a unique serial number — please try again.")


@router.post("/import")
def import_items(body: ItemImportRequest, user: dict = Depends(require_items_admin), db: Session = Depends(get_db)):
    now = datetime.now(timezone.utc).isoformat()
    created = updated = skipped = 0

    # Matching order: (1) SERIAL — the stable identity. (2) For rows with a BLANK
    # serial, a CONTENT signature matched as a MULTISET: re-importing the same file
    # maps each row onto one existing unit (so it updates in place instead of
    # duplicating), yet two genuinely identical units keep two slots and never
    # collapse. A blank-serial row only creates a NEW unit when no unclaimed match
    # is left — then it gets the next Nexus-assigned GG-#####.
    by_serial: dict[str, Item] = {}
    sig_pool: dict[tuple, deque] = defaultdict(deque)
    for it in db.query(Item).all():
        s = (it.serial_number or "").strip().lower()
        if s:
            by_serial.setdefault(s, it)
        sig_pool[_content_sig(it.name, it.item_type, it.make, it.model, it.year,
                              it.department, it.location, it.ownership_type)].append(it)

    claimed: set[str] = set()          # existing items already matched this import
    next_serial = _serial_start(db)    # scanned once — see _serial_start
    _seed_types_if_empty(db)
    _type_rows = db.query(ItemType).all()
    canon = _build_canon([r.name for r in _type_rows])   # built once, extended below
    next_type_order = max([r.sort_order for r in _type_rows], default=0) + 1
    added_types: list = []             # new types this import created (manager-initiated)
    # Resolve any new type labels ONCE, up front. The AI maps same-item variants onto
    # an existing type (IP Cams → IP Camera) but keeps distinct items distinct; whatever
    # it doesn't match becomes a brand-new type. After this, the row loop just normalises.
    _raw_types  = {_clean_field(r.item_type) for r in body.items if _clean_field(r.item_type)}
    _new_labels = [t for t in _raw_types if _type_key(t) not in canon and len(t) <= 40]
    _ai_map = _ai_match_types(_new_labels, [r.name for r in _type_rows]) if _new_labels else {}
    for label in _new_labels:
        k = _type_key(label)
        if k in canon:
            continue
        match = _ai_map.get(label)
        match_key = _type_key(match) if isinstance(match, str) and match.strip() else None
        if match_key and match_key in canon:
            canon[k] = canon[match_key]    # variant of an existing type → reuse it
        else:
            db.add(ItemType(name=label, sort_order=next_type_order, created_by=user["email"], created_at=now))
            canon[k] = label
            next_type_order += 1
            added_types.append(label)

    def _claim_by_sig(sig):
        pool = sig_pool.get(sig)
        while pool:
            cand = pool.popleft()
            if cand.id not in claimed:
                return cand
        return None

    for row in body.items:
        name = (row.name or "").strip()
        if not name:
            skipped += 1
            continue
        ownership = (row.ownership_type or "transient").strip().lower()
        if ownership not in ("permanent", "transient"):
            ownership = "transient"
        item_type = _normalize_type(row.item_type, canon)   # canon resolved above
        make       = _clean_field(row.make)
        model      = _clean_field(row.model)
        year       = _clean_field(row.year)
        department = _clean_field(row.department)
        location   = _clean_field(row.location)
        default_owner = _clean_field(row.default_owner) or _TYPE_DEFAULT_OWNER.get(item_type, "")

        # Custom-field values supplied on the row (keyed by field_key). Blanks are
        # ignored so a sparse spreadsheet never wipes an existing value.
        row_cf = {k: v for k, v in (row.custom_fields or {}).items() if v not in (None, "")}

        serial = (row.serial_number or "").strip()
        if serial:
            existing = by_serial.get(serial.lower())   # same serial = same unit (repeats collapse)
        else:
            existing = _claim_by_sig(_content_sig(name, item_type, make, model, year, department, location, ownership))

        if existing is not None:
            # Update descriptive fields in place. Serial, status, photo, and the
            # assignment lifecycle are deliberately preserved.
            claimed.add(existing.id)   # so a later blank-serial row can't re-match this unit by content
            existing.name          = name
            existing.item_type     = item_type
            existing.make          = make
            existing.model         = model
            existing.year          = year
            existing.department    = department
            existing.default_owner = default_owner
            existing.ownership_type = ownership
            existing.location      = location
            if row_cf:
                existing.custom_fields = {**(existing.custom_fields or {}), **row_cf}
            updated += 1
        else:
            # Honour a serial the CSV supplied; otherwise assign the next GG-#####,
            # skipping any already taken in the DB or earlier in this same file.
            if not serial:
                while _fmt_serial(next_serial).lower() in by_serial:
                    next_serial += 1
                serial = _fmt_serial(next_serial)
                next_serial += 1
            new_item = Item(
                id=str(uuid.uuid4()),
                serial_number=serial,
                name=name,
                item_type=item_type,
                make=make,
                model=model,
                year=year,
                department=department,
                default_owner=default_owner,
                ownership_type=ownership,
                status="available",  # assignment flow sets permanently_assigned once accepted
                location=location,
                photo_url="",
                custom_fields=row_cf,
                created_by=user["email"],
                created_at=now,
            )
            db.add(new_item)
            by_serial[serial.lower()] = new_item  # a repeated explicit serial within the file collapses
            # NOT added to sig_pool: two identical rows in one file are two distinct
            # new units, so the second must not match the first.
            created += 1
    db.commit()
    # Report any types this import created so the UI can refresh the type list.
    return {"created": created, "updated": updated, "skipped": skipped,
            "added_types": sorted(set(added_types))}


# ── Persistent cart (must be before /{item_id} wildcards) ────────────────────

class CartAddBody(BaseModel):
    item_id:   str
    item_name: str
    item_type: str = "Other"

@router.get("/cart")
def get_cart(user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    email = (user.get("preferred_username") or user.get("email") or "").lower()
    rows = db.query(ItemCartEntry).filter(ItemCartEntry.user_email == email).order_by(ItemCartEntry.added_at).all()
    return [{"id": r.id, "itemId": r.item_id, "itemName": r.item_name, "itemType": r.item_type, "addedAt": r.added_at} for r in rows]

@router.post("/cart")
def add_to_cart(body: CartAddBody, user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    email = (user.get("preferred_username") or user.get("email") or "").lower()
    # Validate item exists and is requestable
    item = db.query(Item).filter(Item.id == body.item_id).first()
    if not item:
        raise HTTPException(404, "Item not found")
    if item.ownership_type != "transient":
        raise HTTPException(400, "Only transient items can be added to cart")
    if item.status != "available":
        raise HTTPException(409, f'"{item.name}" is not currently available')
    existing = db.query(ItemCartEntry).filter(ItemCartEntry.user_email == email, ItemCartEntry.item_id == body.item_id).first()
    if existing:
        return {"id": existing.id, "itemId": existing.item_id, "itemName": existing.item_name, "itemType": existing.item_type, "addedAt": existing.added_at}
    entry = ItemCartEntry(
        id=str(uuid.uuid4()), user_email=email,
        item_id=body.item_id, item_name=body.item_name, item_type=body.item_type,
        added_at=datetime.now(timezone.utc).isoformat(),
    )
    db.add(entry)
    db.commit()
    return {"id": entry.id, "itemId": entry.item_id, "itemName": entry.item_name, "itemType": entry.item_type, "addedAt": entry.added_at}

@router.delete("/cart/{item_id}")
def remove_from_cart(item_id: str, user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    email = (user.get("preferred_username") or user.get("email") or "").lower()
    db.query(ItemCartEntry).filter(ItemCartEntry.user_email == email, ItemCartEntry.item_id == item_id).delete()
    db.commit()
    return {"ok": True}

@router.delete("/cart")
def clear_cart(user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    email = (user.get("preferred_username") or user.get("email") or "").lower()
    db.query(ItemCartEntry).filter(ItemCartEntry.user_email == email).delete()
    db.commit()
    return {"ok": True}


@router.patch("/{item_id}")
def update_item(item_id: str, body: ItemUpdate, user: dict = Depends(require_items_admin), db: Session = Depends(get_db)):
    item = db.query(Item).filter(Item.id == item_id).first()
    if not item:
        raise HTTPException(404, "Item not found")
    # P1-14: validate any new catalogue photo URL originates from our storage bucket.
    if body.photo_url is not None:
        _validate_photo_url(body.photo_url, "photo_url")
    if body.name is not None:
        n = body.name.strip()
        if not n:
            raise HTTPException(400, "Name cannot be empty")
        item.name = n
    if body.item_type  is not None:
        item.item_type = _normalize_type(body.item_type, _type_canon(db))
    if body.make           is not None: item.make           = body.make.strip()
    if body.model          is not None: item.model          = body.model.strip()
    if body.year           is not None: item.year           = body.year.strip()
    if body.department     is not None: item.department     = body.department.strip()
    if body.default_owner  is not None: item.default_owner  = body.default_owner.strip()
    if body.ownership_type is not None:
        ot = body.ownership_type.strip().lower()
        if ot and ot not in ("permanent", "transient"):
            raise HTTPException(400, "ownership_type must be 'permanent' or 'transient'")
        item.ownership_type = ot
    # P1-3: lifecycle `status` is derived-only — no raw write here (see ItemUpdate and
    # POST /items/{id}/reconcile-state). op_status below is the operational field.
    if body.location  is not None: item.location  = body.location.strip()
    if body.photo_url is not None: item.photo_url = body.photo_url.strip()
    if body.picture_required is not None: item.picture_required = bool(body.picture_required)
    if body.asset_value      is not None: item.asset_value      = float(body.asset_value)
    # Op status + the person it's declared against (lost/retired). Capture a fresh
    # declaration so we notify that person exactly once (on status or person change).
    notify_decl = None
    if body.op_status is not None:
        op = body.op_status.strip()
        if op and op not in _OP_STATUSES:
            raise HTTPException(400, f"Invalid op_status. Must be one of: {', '.join(_OP_STATUSES)}")
        prev_op    = item.op_status or ""
        prev_email = (item.op_status_person_email or "").lower()
        item.op_status = op
        if op in _OP_STATUS_PERSON:
            if body.op_status_person_email is not None:
                item.op_status_person_email = body.op_status_person_email.lower().strip()
            if body.op_status_person_name is not None:
                item.op_status_person_name = body.op_status_person_name.strip()
            new_email = (item.op_status_person_email or "").lower()
            if new_email and (op != prev_op or new_email != prev_email):
                notify_decl = (op, new_email, item.op_status_person_name or "", item.name)
        else:
            # status is no longer person-bound — drop any stale declaration person
            item.op_status_person_email = ""
            item.op_status_person_name  = ""
    if body.custom_fields is not None:
        # Merge: only the keys sent are changed; a key set to '' / None clears that field.
        merged = dict(item.custom_fields or {})
        for k, v in body.custom_fields.items():
            if v in (None, ""):
                merged.pop(k, None)
            else:
                merged[k] = v
        item.custom_fields = merged
    if notify_decl:
        _notify_op_status_declaration(db, op_status=notify_decl[0], person_email=notify_decl[1],
                                      person_name=notify_decl[2], item_name=notify_decl[3])
    db.commit()
    return _item_to_dict(item)


def _soft_delete(item: Item, user_email: str) -> None:
    """Mark an item deleted instead of dropping the row, so it can be restored and
    carries a "Deleted In" (location at deletion time) — Ankush. The serial/index
    stay reserved until the row is restored or purged."""
    item.deleted_at = datetime.now(timezone.utc).isoformat()
    item.deleted_by = user_email
    item.deleted_location = item.location or ""


@router.delete("/{item_id}")
def delete_item(item_id: str, user: dict = Depends(require_items_delete), db: Session = Depends(get_db)):
    item = db.query(Item).filter(Item.id == item_id).first()
    if not item or item.deleted_at:
        raise HTTPException(404, "Item not found")
    active = db.query(ItemCheckout).filter(
        ItemCheckout.item_id == item_id,
        ItemCheckout.status.in_(["pending", "approved", "pending_receipt", "allocated"]),
    ).count()
    if active:
        raise HTTPException(409, "Cannot delete an item with an active checkout against it")
    live_assignment = db.query(ItemAssignment).filter(
        ItemAssignment.item_id == item_id,
        ItemAssignment.status.in_(["pending_acceptance", "active", "return_initiated"]),
    ).count()
    if live_assignment:
        raise HTTPException(409, "Cannot delete an item someone is still assigned — recover it first")
    _soft_delete(item, user["email"])
    db.commit()
    return {"ok": True}


class BulkDeleteRequest(BaseModel):
    ids: list[str]


@router.post("/bulk-delete")
def bulk_delete_items(body: BulkDeleteRequest, user: dict = Depends(require_items_delete), db: Session = Depends(get_db)):
    """Soft-delete many items in ONE transaction instead of one request per item
    (the client used to loop, which made deleting 30+ items take 10-20s). Items
    with an active checkout are skipped and reported as `blocked`."""
    ids = [i for i in (body.ids or []) if i]
    if not ids:
        return {"deleted": 0, "blocked": [], "notFound": []}

    # One query each: which of these items still have a live checkout, and which
    # actually exist — instead of N round-trips.
    active_item_ids = {
        row.item_id for row in db.query(ItemCheckout.item_id).filter(
            ItemCheckout.item_id.in_(ids),
            ItemCheckout.status.in_(["pending", "approved", "pending_receipt", "allocated"]),
        ).all()
    }
    # Live permanent assignments block deletion the same way live checkouts do —
    # deleting a held item orphans the assignment once the recycle bin purges.
    assigned_item_ids = {
        row.item_id for row in db.query(ItemAssignment.item_id).filter(
            ItemAssignment.item_id.in_(ids),
            ItemAssignment.status.in_(["pending_acceptance", "active", "return_initiated"]),
        ).all()
    }
    items = db.query(Item).filter(Item.id.in_(ids), or_(Item.deleted_at.is_(None), Item.deleted_at == "")).all()
    found_ids = {it.id for it in items}

    blocked, deleted = [], 0
    for it in items:
        if it.id in active_item_ids:
            blocked.append({"id": it.id, "name": it.name, "reason": "active checkout"})
            continue
        if it.id in assigned_item_ids:
            blocked.append({"id": it.id, "name": it.name, "reason": "live assignment"})
            continue
        _soft_delete(it, user["email"])
        deleted += 1
    db.commit()  # single commit for the whole batch

    return {
        "deleted": deleted,
        "blocked": blocked,
        "notFound": [i for i in ids if i not in found_ids],
    }


@router.get("/deleted")
def list_deleted_items(user: dict = Depends(require_items_delete), db: Session = Depends(get_db)):
    """The recycle bin — soft-deleted items, most-recently-deleted first, so a
    bad 'select all + delete' is recoverable (Ankush). Items older than the
    retention window are purged for good first (lazy cleanup — ISO timestamps are
    all UTC, so a string compare is chronological)."""
    from datetime import timedelta
    cutoff = (datetime.now(timezone.utc) - timedelta(days=_RECYCLE_BIN_DAYS)).isoformat()
    expired = db.query(Item).filter(
        Item.deleted_at.isnot(None), Item.deleted_at != "", Item.deleted_at < cutoff).all()
    if expired:
        for it in expired:
            db.delete(it)
        db.commit()
    rows = db.query(Item).filter(Item.deleted_at.isnot(None), Item.deleted_at != "") \
             .order_by(Item.deleted_at.desc()).limit(2000).all()
    return [_item_to_dict(i) for i in rows]


class RestoreRequest(BaseModel):
    ids: list[str]


def _restore(item: Item) -> None:
    item.deleted_at = item.deleted_by = item.deleted_location = ""


@router.post("/{item_id}/restore")
def restore_item(item_id: str, user: dict = Depends(require_items_delete), db: Session = Depends(get_db)):
    item = db.query(Item).filter(Item.id == item_id).first()
    if not item:
        raise HTTPException(404, "Item not found")
    if not item.deleted_at:
        raise HTTPException(409, "Item is not deleted")
    _restore(item)
    db.commit()
    return _item_to_dict(item)


@router.post("/bulk-restore")
def bulk_restore_items(body: RestoreRequest, user: dict = Depends(require_items_delete), db: Session = Depends(get_db)):
    ids = [i for i in (body.ids or []) if i]
    if not ids:
        return {"restored": 0}
    restored = 0
    for it in db.query(Item).filter(Item.id.in_(ids), Item.deleted_at.isnot(None), Item.deleted_at != "").all():
        _restore(it)
        restored += 1
    db.commit()
    return {"restored": restored}


class BulkUpdateRequest(BaseModel):
    ids:    list[str]
    fields: dict          # only the descriptive fields the user chose to change


# Fields a batch edit may touch. Serial is the static identity and is never editable
# here; status/photo/assignment lifecycle are owned by their own flows.
_BATCH_EDITABLE = {
    "name", "item_type", "make", "model", "year",
    "department", "default_owner", "ownership_type", "location", "op_status",
    "asset_value",
}


@router.post("/bulk-update")
def bulk_update_items(body: BulkUpdateRequest, user: dict = Depends(require_items_admin), db: Session = Depends(get_db)):
    """Apply the SAME field changes to many items in one transaction — the Manage
    tab's Batch Edit. Only whitelisted descriptive fields are written, and only the
    keys the user actually chose to change are present in `fields`."""
    ids = [i for i in (body.ids or []) if i]
    changes = {k: v for k, v in (body.fields or {}).items() if k in _BATCH_EDITABLE}
    if not ids or not changes:
        return {"updated": 0}

    # Normalise the way create/import do, so batch edits stay consistent with them.
    if "item_type" in changes:
        changes["item_type"] = _normalize_type(changes["item_type"], _type_canon(db))
    if "ownership_type" in changes:
        o = (changes["ownership_type"] or "transient").strip().lower()
        changes["ownership_type"] = o if o in ("permanent", "transient") else "transient"
    if "op_status" in changes:
        op = (changes["op_status"] or "").strip()
        if op and op not in _OP_STATUSES:
            raise HTTPException(400, f"Invalid op_status. Must be one of: {', '.join(_OP_STATUSES)}")
        changes["op_status"] = op
    if "name" in changes:
        nm = (changes["name"] or "").strip()
        if not nm:
            raise HTTPException(400, "Name cannot be empty")
        changes["name"] = nm
    for k in ("make", "model", "year", "department", "default_owner", "location"):
        if k in changes:
            changes[k] = (changes[k] or "").strip()
    if "asset_value" in changes:
        try:
            changes["asset_value"] = float(changes["asset_value"] or 0)
        except (ValueError, TypeError):
            changes["asset_value"] = 0.0

    updated = 0
    op_changed = "op_status" in changes
    for it in db.query(Item).filter(Item.id.in_(ids), or_(Item.deleted_at.is_(None), Item.deleted_at == "")).all():
        for k, v in changes.items():
            setattr(it, k, v)
        # A batch can't attach a per-item person, so drop any stale declaration when
        # the op_status is changed in bulk (set the person via single Edit instead).
        if op_changed:
            it.op_status_person_email = ""
            it.op_status_person_name  = ""
        updated += 1
    db.commit()
    return {"updated": updated}


# ── Custom fields ─────────────────────────────────────────────────────────────
# Admin-defined extra fields surfaced in the item Details panel (Ankush). The
# definition lives in item_custom_fields; the per-item value lives in
# items.custom_fields keyed by field_key, so the main table stays lean.

def _slugify_key(label: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "_", (label or "").strip().lower()).strip("_")
    return s or "field"


class CustomFieldIn(BaseModel):
    field_key:       Optional[str]  = ""
    label:           str
    field_type:      Optional[str]  = "text"
    options:         Optional[list] = None
    applies_to_type: Optional[str]  = ""
    sort_order:      Optional[int]  = 0


@router.get("/custom-fields")
def list_custom_fields(user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    rows = db.query(ItemCustomField).order_by(ItemCustomField.sort_order, ItemCustomField.label).all()
    return [_custom_field_to_dict(f) for f in rows]


@router.post("/custom-fields", status_code=201)
def create_custom_field(body: CustomFieldIn, user: dict = Depends(require_items_admin), db: Session = Depends(get_db)):
    label = (body.label or "").strip()
    if not label:
        raise HTTPException(400, "Label cannot be empty")
    ftype = (body.field_type or "text").strip().lower()
    if ftype not in _CUSTOM_FIELD_TYPES:
        raise HTTPException(400, f"Invalid field type. Must be one of: {', '.join(_CUSTOM_FIELD_TYPES)}")
    # Derive a stable key from the label, kept unique across definitions.
    base = (body.field_key or "").strip() or _slugify_key(label)
    existing = {f.field_key for f in db.query(ItemCustomField.field_key).all()}
    key, n = base, 2
    while key in existing:
        key = f"{base}_{n}"; n += 1
    f = ItemCustomField(
        id=str(uuid.uuid4()), field_key=key, label=label, field_type=ftype,
        options=[str(o).strip() for o in (body.options or []) if str(o).strip()] if ftype == "select" else [],
        applies_to_type=(body.applies_to_type or "").strip(),
        sort_order=int(body.sort_order or 0),
        created_by=user["email"], created_at=datetime.now(timezone.utc).isoformat(),
    )
    db.add(f)
    db.commit()
    return _custom_field_to_dict(f)


@router.patch("/custom-fields/{field_id}")
def update_custom_field(field_id: str, body: CustomFieldIn, user: dict = Depends(require_items_admin), db: Session = Depends(get_db)):
    f = db.query(ItemCustomField).filter(ItemCustomField.id == field_id).first()
    if not f:
        raise HTTPException(404, "Custom field not found")
    if body.label is not None and body.label.strip():
        f.label = body.label.strip()
    if body.field_type:
        ftype = body.field_type.strip().lower()
        if ftype not in _CUSTOM_FIELD_TYPES:
            raise HTTPException(400, f"Invalid field type. Must be one of: {', '.join(_CUSTOM_FIELD_TYPES)}")
        f.field_type = ftype
    if body.options is not None:
        f.options = [str(o).strip() for o in body.options if str(o).strip()] if f.field_type == "select" else []
    if body.applies_to_type is not None:
        f.applies_to_type = body.applies_to_type.strip()
    if body.sort_order is not None:
        f.sort_order = int(body.sort_order or 0)
    db.commit()
    return _custom_field_to_dict(f)


@router.delete("/custom-fields/{field_id}")
def delete_custom_field(field_id: str, user: dict = Depends(require_items_delete), db: Session = Depends(get_db)):
    f = db.query(ItemCustomField).filter(ItemCustomField.id == field_id).first()
    if not f:
        raise HTTPException(404, "Custom field not found")
    db.delete(f)  # per-item values stay in items.custom_fields but stop showing — harmless
    db.commit()
    return {"ok": True}


# ── Item types (manager-curated) ──────────────────────────────────────────────
# Managers extend the type list here; a CSV can't invent one (it funnels to Other).
class ItemTypeIn(BaseModel):
    name: str


@router.get("/types")
def list_item_types(user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    return _type_names(db)


@router.post("/types", status_code=201)
def add_item_type(body: ItemTypeIn, user: dict = Depends(require_items_admin), db: Session = Depends(get_db)):
    name = (body.name or "").strip()
    if not name:
        raise HTTPException(400, "Type name cannot be empty")
    if len(name) > 40:
        raise HTTPException(400, "Type name is too long (40 characters max)")
    _seed_types_if_empty(db)
    rows = db.query(ItemType).all()
    # Dedupe by the loose key so office/offices/OFFICE collapse to one type.
    if any(_type_key(r.name) == _type_key(name) for r in rows):
        return _type_names(db)
    nxt = max([r.sort_order for r in rows], default=0) + 1
    db.add(ItemType(name=name, sort_order=nxt, created_by=user["email"], created_at=_now_iso()))
    db.commit()
    return _type_names(db)


@router.delete("/types/{name}")
def delete_item_type(name: str, user: dict = Depends(require_items_admin), db: Session = Depends(get_db)):
    if name.strip().lower() == "other":
        raise HTTPException(400, "“Other” is the fallback type and can't be removed.")
    row = next((r for r in db.query(ItemType).all() if r.name.lower() == name.strip().lower()), None)
    if row:
        db.delete(row)
        db.commit()
    return _type_names(db)


# ── Checkouts ─────────────────────────────────────────────────────────────────

class CheckoutIn(BaseModel):
    id:                  Optional[str] = None  # ignored — server generates
    item_id:             str
    item_name:           str
    item_type:           Optional[str] = ""
    requested_by:        str
    requested_by_email:  str = ""
    raised_by:           str
    department:          str = ""
    days:                int = 1
    reason:              str = ""
    checkout_photo_url:  Optional[str] = ""
    checkout_photo_name: Optional[str] = ""
    order_id:            Optional[str] = ""
    approver_email:      Optional[str] = ""   # manager who should receive the approval notification
    approver_name:       Optional[str] = ""

    def validate_days(self):
        if not (1 <= self.days <= 90):
            raise HTTPException(400, "days must be between 1 and 90")

    def validate_lengths(self):
        if len(self.reason) > 1000:
            raise HTTPException(400, "reason too long (max 1000 chars)")
        if len(self.requested_by) > 200:
            raise HTTPException(400, "requested_by too long")


class CheckoutStatusUpdate(BaseModel):
    status:                    str
    resolved_by:               Optional[str] = ""
    reject_reason:             Optional[str] = ""
    assigned_allocator_email:  Optional[str] = ""
    assigned_allocator_name:   Optional[str] = ""
    allocated_by:              Optional[str] = ""
    checkout_photo_url:        Optional[str] = ""
    checkout_photo_name:       Optional[str] = ""
    return_photo_name:         Optional[str] = ""
    return_photo_url:          Optional[str] = ""
    condition_note:            Optional[str] = ""
    condition:                 Optional[str] = "ok"  # P1-13: ok | damaged | lost (return only)
    handover_photo_by:         Optional[str] = ""   # 'allocator' | 'employee'
    handover_batch:            Optional[bool] = False
    receipt_photo_url:         Optional[str] = ""
    receipt_photo_name:        Optional[str] = ""
    handed_over_at:            Optional[str] = ""


@router.get("/checkouts")
def list_checkouts(user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    q = db.query(ItemCheckout).order_by(ItemCheckout.created_at.desc())
    if not _is_items_manager(user, db):
        q = q.filter(or_(
            ItemCheckout.requested_by_email == user["email"],
            ItemCheckout.assigned_allocator_email == user["email"],
        ))
    return [_checkout_to_dict(c) for c in q.limit(2000).all()]


@router.post("/checkouts", status_code=201)
def create_checkout(body: CheckoutIn, user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    body.validate_days()
    body.validate_lengths()
    if not body.reason.strip():
        raise HTTPException(400, "Reason for checkout is required")

    # Server always generates the checkout ID — client-supplied IDs are ignored
    # to prevent ID injection / collision attacks.
    server_id = f"ICHK-{uuid.uuid4().hex[:8].upper()}-{uuid.uuid4().hex[:8].upper()}"

    # P1-1: lock the item row so two concurrent checkout requests can't both pass the
    # availability / active-checkout check and double-book the same unit. FOR UPDATE
    # serializes them; the loser then sees the winner's committed state below.
    item = db.query(Item).filter(Item.id == body.item_id).with_for_update().first()
    if not item or item.deleted_at:
        raise HTTPException(404, "Item not found")
    if item.ownership_type != "transient":
        raise HTTPException(400, "Only transient items can be checked out")
    if item.status != "available":
        raise HTTPException(409, f'"{item.name}" is no longer available — it may have just been taken by someone else')

    # Verify no active checkout exists for this item (race condition guard)
    active = db.query(ItemCheckout).filter(
        ItemCheckout.item_id == body.item_id,
        ItemCheckout.status.in_(["pending", "approved", "pending_receipt", "allocated"]),
    ).first()
    if active:
        raise HTTPException(409, f'"{item.name}" already has an active checkout request')
    live_assignment = db.query(ItemAssignment).filter(
        ItemAssignment.item_id == body.item_id,
        ItemAssignment.status.in_(["pending_acceptance", "active", "return_initiated"]),
    ).first()
    if live_assignment:
        raise HTTPException(409, f'"{item.name}" is permanently assigned and cannot be checked out')

    now = datetime.now(timezone.utc).isoformat()
    # Managers and above (or an inventory editor+ grant) don't need a separate
    # approval for their own checkouts
    is_manager = _is_items_manager(user, db)
    requester_email = body.requested_by_email.lower()
    user_email = user.get("email", "").lower()
    self_checkout = is_manager and requester_email == user_email
    initial_status = "approved" if self_checkout else "pending"

    order_id = (body.order_id or "").strip()
    row = ItemCheckout(
        id=server_id,
        item_id=body.item_id,
        item_name=body.item_name,
        item_type=body.item_type or "",
        requested_by=body.requested_by,
        requested_by_email=requester_email,
        raised_by=body.raised_by,
        department=body.department,
        days=body.days,
        reason=body.reason,
        status=initial_status,
        created_at=now,
        resolved_at=now if self_checkout else "",
        resolved_by=body.requested_by if self_checkout else "",
        checkout_photo_url=body.checkout_photo_url or "",
        checkout_photo_name=body.checkout_photo_name or "",
        order_id=order_id,
        approver_email=(body.approver_email or "").lower().strip(),
        approver_name=(body.approver_name or "").strip(),
    )
    db.add(row)
    if initial_status == "pending":
        # One notification per order — if this order_id already has a checkout_pending
        # notification, skip to avoid spamming managers with N alerts for one cart.
        # Cart submits POST all items concurrently, so take an advisory lock on the
        # order_id first; otherwise every request sees "no notification yet" and
        # the dedupe check races into N duplicates.
        if order_id and db.get_bind().dialect.name == "postgresql":
            db.execute(sa_text("SELECT pg_advisory_xact_lock(hashtext(:oid))"), {"oid": order_id})
        ref_for_notif = order_id if order_id else server_id
        # Only an UN-actioned notification counts as "already notified" — a
        # re-request rejoining an old order must ping the manager again even
        # though the order's original notification was long since handled.
        already_notified = order_id and db.query(NexusNotification).filter(
            NexusNotification.ref_id == order_id,
            NexusNotification.type == "checkout_pending",
            NexusNotification.actioned == False,
        ).first()
        # Item count + total $ value of the order so far (this row included) —
        # the manager sees what's at stake before approving. autoflush is off,
        # so the current row is counted manually.
        sib_item_ids = [r[0] for r in db.query(ItemCheckout.item_id).filter(
            ItemCheckout.order_id == order_id,
            ItemCheckout.id != server_id,
            ItemCheckout.status == "pending",
        ).all()] if order_id else []
        order_count = len(sib_item_ids) + 1
        val_rows = db.query(Item).filter(Item.id.in_(set(sib_item_ids + [body.item_id]))).all()
        total_value = sum(float(i.asset_value or 0) for i in val_rows)
        notif_body = (f"{body.requested_by} has submitted a Checkout Request — "
                      f"{order_count} item{'s' if order_count != 1 else ''}"
                      + (f", total value **${total_value:,.0f}**" if total_value > 0 else "")
                      + ".\nPlease review, approve or reject.")
        if not already_notified:
            # Targeted: only the manager the employee picked gets the notification.
            # Empty approver (legacy clients / managers raising on behalf) falls back
            # to the all-managers broadcast. The request itself remains visible in
            # every manager's Checkouts tab regardless — anyone can still approve.
            _notify(db, type="checkout_pending", recipient=(body.approver_email or "").lower().strip(),
                    title=f"Checkout Request — {body.requested_by}",
                    body=notif_body,
                    ref_id=ref_for_notif, item_name=body.item_name, requested_by=body.requested_by)
        else:
            # Subsequent cart items: refresh the count/value on the existing
            # notification (the realtime ping pushes the new body to open bells)
            already_notified.body = notif_body
    db.commit()
    _fire_item_event(server_id, initial_status, row.requested_by_email or "")
    return _checkout_to_dict(row)


@router.patch("/checkouts/{checkout_id}")
def update_checkout(checkout_id: str, body: CheckoutStatusUpdate, user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    # P1-2: lock the checkout row itself up front. Previously FOR UPDATE was only taken
    # on the ORDER siblings (below), so a SOLO (no-order) checkout took no lock at all —
    # a concurrent approve + cancel could both read status="pending" and both proceed.
    # Locking the row here serializes those transitions for the solo path too.
    row = db.query(ItemCheckout).filter(ItemCheckout.id == checkout_id).with_for_update().first()
    if not row:
        # Was a 200 {"ok": false} fake-success — clients checking resp.ok showed
        # success on a vanished/stale checkout (Jul 14 audit family).
        raise HTTPException(404, "Checkout not found")

    if row.order_id:
        # Serialize concurrent updates within the same order. "Approve All" /
        # "Hand Over All" / "Return All" fire one PATCH per item in parallel;
        # without this lock each transaction reads its siblings as unchanged
        # and the order-level notification batching below double-fires (one
        # notification per item instead of one per order). FOR UPDATE makes
        # the transactions queue, so each sees the previous one's commit.
        db.query(ItemCheckout).filter(
            ItemCheckout.order_id == row.order_id
        ).with_for_update().all()
        db.refresh(row)

    # Validate photo URLs to prevent fake evidence from external sources
    _validate_photo_url(body.checkout_photo_url, "checkout_photo_url")
    _validate_photo_url(body.receipt_photo_url,  "receipt_photo_url")
    _validate_photo_url(body.return_photo_url,   "return_photo_url")

    items_mgr = _is_items_manager(user, db)   # manager+ OR inventory grant editor+
    if body.status in ("approved", "rejected") and not items_mgr:
        raise HTTPException(403, "Manager or above required to approve or reject checkouts")
    if body.status == "approved" and not (body.assigned_allocator_email or "").strip():
        raise HTTPException(400, "Pick who should allocate this item before approving")
    if body.status == "pending_receipt":
        is_assignee = row.assigned_allocator_email and row.assigned_allocator_email.lower() == user["email"]
        if not is_assignee and not items_mgr:
            raise HTTPException(403, "Only the assigned allocator or a manager can initiate handover")
    if body.status == "allocated":
        is_assignee  = row.assigned_allocator_email and row.assigned_allocator_email.lower() == user["email"]
        is_requester = row.requested_by_email and row.requested_by_email.lower() == user["email"]
        if not is_assignee and not is_requester and not items_mgr:
            raise HTTPException(403, "Only the assigned allocator, the requester, or a manager can confirm handover")
    if body.status == "returned" and user["level"] < 2 and not items_mgr \
            and row.requested_by_email.lower() != user["email"]:
        raise HTTPException(403, "You can only return your own items")
    if body.status == "cancelled":
        # P1-7: the requester can always self-cancel; a manager (or inventory
        # editor+ grant) may also cancel a still-pending/approved checkout on
        # someone's behalf (symmetric with assignment force-cancel). The
        # requester is then told WHO cancelled it (below).
        is_requester = row.requested_by_email.lower() == user["email"].lower()
        if not is_requester and not items_mgr:
            raise HTTPException(403, "You can only cancel your own checkouts")
        if items_mgr and not is_requester and row.status not in ("pending", "approved"):
            raise HTTPException(409, "A manager can only cancel a pending or approved checkout")

    valid_predecessors = _VALID_TRANSITIONS.get(body.status)
    if valid_predecessors is not None and row.status not in valid_predecessors:
        raise HTTPException(409, f"Cannot move a '{row.status}' checkout to '{body.status}'")

    now = datetime.now(timezone.utc).isoformat()
    item = db.query(Item).filter(Item.id == row.item_id).first()

    if body.status in ("approved", "rejected"):
        row.resolved_at = now
        row.resolved_by = body.resolved_by or ""
        if body.status == "approved":
            row.assigned_allocator_email = (body.assigned_allocator_email or "").lower().strip()
            row.assigned_allocator_name  = (body.assigned_allocator_name  or "").strip()
        else:
            row.reject_reason = body.reject_reason or ""

    elif body.status == "cancelled":
        row.resolved_at = now
        # P1-7: record WHO cancelled — the acting user's name if the client didn't
        # supply one (a manager cancelling on behalf otherwise recorded no actor).
        row.resolved_by = body.resolved_by or _title_case_email(user["email"])

    elif body.status == "pending_receipt":
        # Supervisor confirmed physical handover; employee will upload receipt photo
        row.handed_over_at    = now
        row.handover_photo_by = "employee"
        row.handover_batch    = bool(body.handover_batch)
        if body.checkout_photo_url:
            row.checkout_photo_url  = body.checkout_photo_url
            row.checkout_photo_name = body.checkout_photo_name or ""
        row.allocated_by = body.allocated_by or ""

    elif body.status == "allocated":
        coming_from_pending_receipt = row.status == "pending_receipt"
        if not coming_from_pending_receipt:
            # Direct allocator-photo handover: item must still be available
            if item and item.status != "available":
                raise HTTPException(409, "Item is no longer available to allocate")
            row.handed_over_at    = now
            row.handover_photo_by = body.handover_photo_by or "allocator"
            row.handover_batch    = bool(body.handover_batch)
            if body.checkout_photo_url:
                row.checkout_photo_url  = body.checkout_photo_url
                row.checkout_photo_name = body.checkout_photo_name or ""
        else:
            # Employee confirming receipt after supervisor initiated handover
            row.receipt_photo_url  = body.receipt_photo_url  or ""
            row.receipt_photo_name = body.receipt_photo_name or ""
        row.allocated_at = now
        row.allocated_by = body.allocated_by or row.allocated_by or ""
        if item:
            item.status = "checked_out"

    elif body.status == "returned":
        # P1-13: explicit condition enum (ok|damaged|lost) instead of sniffing the
        # free-text note for damage keywords. Only damaged/lost change op_status; the
        # note stays free-text. Defaults to "ok" so pre-enum callers keep working.
        condition = (body.condition or "ok").lower().strip()
        if condition not in _RETURN_CONDITIONS:
            condition = "ok"
        row.returned_at      = now
        row.return_photo_name = body.return_photo_name or ""
        row.return_photo_url  = body.return_photo_url  or ""
        row.condition_note   = body.condition_note    or ""
        if item:
            # Lifecycle status returns to available (derived from the checkout ending);
            # a damaged/lost declaration is recorded on the SEPARATE op_status field.
            item.status = "available"
            if condition == "lost":
                item.op_status = "lost"
                item.op_status_person_email = (row.requested_by_email or "").lower()
                item.op_status_person_name  = row.requested_by or ""
                _notify_op_status_declaration(
                    db, op_status="lost",
                    person_email=item.op_status_person_email,
                    person_name=item.op_status_person_name, item_name=item.name)
            elif condition == "damaged":
                item.op_status = "in_repair"
        # A pending extension is moot once the item is back — clear it and
        # action the managers' extension notification so it leaves their bell.
        if row.extension_status == "pending":
            row.extension_days   = 0
            row.extension_reason = ""
            row.extension_status = ""
            stale_ext = db.query(NexusNotification).filter(
                NexusNotification.type == "extension_pending",
                NexusNotification.actioned == False,
                NexusNotification.ref_id == checkout_id,
            ).first()
            if stale_ext:
                stale_ext.actioned = True

    row.status = body.status

    # Auto-mark the checkout_pending notification as actioned so it clears from
    # all managers' bells even when the approval happens via the Checkouts tab.
    # For order-level approvals: mark actioned once ALL items in the order are
    # no longer pending (handles item-by-item approval from the Checkouts tab).
    if body.status in ("approved", "rejected", "cancelled"):
        ref_ids = [checkout_id]
        if row.order_id:
            ref_ids.append(row.order_id)
            # Check if any sibling items in this order are still pending
            sibling_pending = db.query(ItemCheckout).filter(
                ItemCheckout.order_id == row.order_id,
                ItemCheckout.id != checkout_id,
                ItemCheckout.status == "pending",
            ).count()
            if sibling_pending > 0:
                ref_ids = []  # don't action yet — order not fully resolved
        if ref_ids:
            pending_notif = db.query(NexusNotification).filter(
                NexusNotification.type == "checkout_pending",
                NexusNotification.actioned == False,
                NexusNotification.ref_id.in_(ref_ids),
            ).first()
            if pending_notif:
                pending_notif.actioned = True

    if body.status in ("approved", "rejected"):
        if row.order_id:
            # Tally all siblings' current status (not yet committed — exclude current item)
            siblings = db.query(ItemCheckout).filter(
                ItemCheckout.order_id == row.order_id,
                ItemCheckout.id != checkout_id,
            ).all()
            approved_names = [s.item_name for s in siblings if s.status == "approved"]
            rejected_names = [s.item_name for s in siblings if s.status == "rejected"]
            still_pending  = [s.item_name for s in siblings if s.status == "pending"]

            # Add current item to the right bucket
            if body.status == "approved":
                approved_names.append(row.item_name)
            else:
                rejected_names.append(row.item_name)

            all_resolved = len(still_pending) == 0

            # Build title and body. Neil: say "request", not "order", and put
            # Approved / Not approved on separate lines (bell renders pre-line).
            if all_resolved:
                if approved_names and rejected_names:
                    notif_type = "approved"
                    notif_title = f"Request update: {len(approved_names)} approved, {len(rejected_names)} rejected"
                    parts = []
                    if approved_names:
                        parts.append(f"Approved: **{', '.join(approved_names)}**")
                    if rejected_names:
                        parts.append(f"Not approved: **{', '.join(rejected_names)}**")
                    notif_body = "\n".join(parts)
                elif approved_names:
                    notif_type = "approved"
                    notif_title = f"Request approved: {len(approved_names)} item{'s' if len(approved_names) != 1 else ''}"
                    notif_body = f"All {len(approved_names)} items from your request were approved. Your allocator will hand them over shortly."
                else:
                    notif_type = "rejected"
                    notif_title = f"Request rejected: {len(rejected_names)} item{'s' if len(rejected_names) != 1 else ''}"
                    notif_body = f"Your {len(rejected_names)}-item request was not approved.\nReason: {row.reject_reason or 'No reason given.'}"
            else:
                total = len(approved_names) + len(rejected_names) + len(still_pending)
                notif_type = "approved" if approved_names else "rejected"
                notif_title = f"Request partially processed — {len(approved_names) + len(rejected_names)} of {total} items"
                parts = []
                if approved_names:
                    parts.append(f"Approved: **{', '.join(approved_names)}**")
                if rejected_names:
                    parts.append(f"Not approved: **{', '.join(rejected_names)}**")
                parts.append(f"Still pending: **{', '.join(still_pending)}**")
                notif_body = "\n".join(parts)

            # Update existing order notification if one exists, otherwise create
            existing = db.query(NexusNotification).filter(
                NexusNotification.ref_id == row.order_id,
                NexusNotification.recipient == (row.requested_by_email or "").lower(),
                NexusNotification.type.in_(["approved", "rejected"]),
                NexusNotification.actioned == False,
            ).first()
            if existing:
                existing.type    = notif_type
                existing.title   = notif_title
                existing.body    = notif_body
                existing.read_by = ""  # re-surface as unread
            else:
                _notify(db, type=notif_type, recipient=row.requested_by_email,
                        title=notif_title, body=notif_body,
                        ref_id=row.order_id, item_name=row.item_name, requested_by=row.requested_by)

            # Allocator notification — one updating notification per order, listing
            # every item assigned to them, instead of one ping per item.
            if body.status == "approved" and row.assigned_allocator_email:
                alloc_email = row.assigned_allocator_email.lower()
                assigned_names = [s.item_name for s in siblings
                                  if s.status == "approved" and (s.assigned_allocator_email or "").lower() == alloc_email]
                assigned_names.append(row.item_name)
                if len(assigned_names) == 1:
                    alloc_title = f"Hand over: {row.item_name}"
                    alloc_body  = f"Please hand {row.item_name} over to {row.requested_by}."
                else:
                    alloc_title = f"Hand over {len(assigned_names)} items to {row.requested_by}"
                    alloc_body  = f"Please hand over: {', '.join(assigned_names)}."
                existing_alloc = db.query(NexusNotification).filter(
                    NexusNotification.ref_id == row.order_id,
                    NexusNotification.recipient == alloc_email,
                    NexusNotification.type == "allocate_request",
                    NexusNotification.actioned == False,
                ).first()
                if existing_alloc:
                    existing_alloc.title   = alloc_title
                    existing_alloc.body    = alloc_body
                    existing_alloc.read_by = ""
                else:
                    _notify(db, type="allocate_request", recipient=row.assigned_allocator_email,
                            title=alloc_title, body=alloc_body,
                            ref_id=row.order_id, item_name=row.item_name, requested_by=row.requested_by)
        else:
            # Solo item (no order) — one notification per action
            if body.status == "approved":
                _notify(db, type="approved", recipient=row.requested_by_email,
                        title=f"Checkout approved: {row.item_name}",
                        body=f"Your request for {row.item_name} was approved. {row.assigned_allocator_name or 'Someone'} will hand it over to you.",
                        ref_id=checkout_id, item_name=row.item_name, requested_by=row.requested_by)
                if row.assigned_allocator_email:
                    _notify(db, type="allocate_request", recipient=row.assigned_allocator_email,
                            title=f"Hand over: {row.item_name}",
                            body=f"Please hand {row.item_name} over to {row.requested_by}.",
                            ref_id=checkout_id, item_name=row.item_name, requested_by=row.requested_by)
            else:
                _notify(db, type="rejected", recipient=row.requested_by_email,
                        title=f"Checkout rejected: {row.item_name}",
                        body=f"Your request for {row.item_name} was not approved. Reason: {row.reject_reason or 'No reason given.'}",
                        ref_id=checkout_id, item_name=row.item_name, requested_by=row.requested_by)
    elif body.status == "pending_receipt":
        if row.order_id:
            # One updating "confirm receipt" notification per order. The old
            # last-item-only gate meant handing over PART of an order produced
            # no notification at all — the missed-notification bug from the
            # Jun 10 demo. Same update-in-place pattern as approvals/returns;
            # the FOR UPDATE lock at the top of this endpoint makes it race-safe.
            handed = db.query(ItemCheckout).filter(
                ItemCheckout.order_id == row.order_id,
                ItemCheckout.id != checkout_id,
                ItemCheckout.status == "pending_receipt",
            ).count() + 1  # +1 for current row (autoflush off — not yet visible)
            notif_title = f"Confirm receipt: {handed} item{'s' if handed != 1 else ''}"
            notif_body = (f"{row.item_name} has been handed over. Please confirm receipt and upload a photo." if handed == 1
                          else f"{handed} items from your request have been handed over. Please confirm receipt and upload a photo for each.")
            existing = db.query(NexusNotification).filter(
                NexusNotification.ref_id == row.order_id,
                NexusNotification.recipient == (row.requested_by_email or "").lower(),
                NexusNotification.type == "allocated",
                NexusNotification.actioned == False,
            ).first()
            if existing:
                existing.title   = notif_title
                existing.body    = notif_body
                existing.read_by = ""  # re-surface as unread
            else:
                _notify(db, type="allocated", recipient=row.requested_by_email,
                        title=notif_title, body=notif_body,
                        ref_id=row.order_id, item_name=row.item_name, requested_by=row.requested_by)
        else:
            _notify(db, type="allocated", recipient=row.requested_by_email,
                    title=f"Confirm receipt: {row.item_name}",
                    body=f"{row.item_name} has been handed over to you. Please confirm receipt and upload a photo.",
                    ref_id=checkout_id, item_name=row.item_name, requested_by=row.requested_by)
    elif body.status == "allocated":
        # For order batches: only notify once, when ALL items in the order are allocated
        if row.order_id:
            sibling_not_allocated = db.query(ItemCheckout).filter(
                ItemCheckout.order_id == row.order_id,
                ItemCheckout.id != checkout_id,
                ItemCheckout.status.notin_(["allocated", "returned", "cancelled"]),
            ).count()
            if sibling_not_allocated == 0:
                # All items now allocated — send one consolidated notification
                order_count = db.query(ItemCheckout).filter(
                    ItemCheckout.order_id == row.order_id,
                    ItemCheckout.status == "allocated",
                ).count() + 1  # +1 for current item (not yet committed)
                _notify(db, type="allocated", recipient=row.requested_by_email,
                        title=f"Request confirmed: {order_count} item{'' if order_count == 1 else 's'} with you",
                        body=f"All {order_count} items from your request are confirmed. Please return them within {row.days} day(s).",
                        ref_id=row.order_id, item_name=row.item_name, requested_by=row.requested_by)
            # else: don't send individual notifications mid-batch
        else:
            _notify(db, type="allocated", recipient=row.requested_by_email,
                    title=f"Item confirmed: {row.item_name}",
                    body=f"{row.item_name} checkout is complete. Please return it within {row.days} day(s).",
                    ref_id=checkout_id, item_name=row.item_name, requested_by=row.requested_by)
    elif body.status == "returned":
        # Only notify the allocator — skip if unset to avoid broadcasting to all users
        if row.assigned_allocator_email:
            if row.order_id:
                # One updating notification per order: tally as items come back
                siblings = db.query(ItemCheckout).filter(
                    ItemCheckout.order_id == row.order_id,
                    ItemCheckout.id != checkout_id,
                ).all()
                returned_names = [s.item_name for s in siblings if s.status == "returned"] + [row.item_name]
                still_out      = [s.item_name for s in siblings if s.status in ("approved", "pending_receipt", "allocated")]

                # Neil's format: bold name, then each list on its own line with
                # bold item names (bell renders **bold** markers)
                if still_out:
                    total = len(returned_names) + len(still_out)
                    notif_title = f"Returns in progress: {len(returned_names)} of {total} items back"
                    notif_body  = (f"**{row.requested_by}**\n"
                                   f"Returned: **{', '.join(returned_names)}**\n"
                                   f"Still out: **{', '.join(still_out)}**")
                else:
                    notif_title = f"Request returned: {len(returned_names)} item{'s' if len(returned_names) != 1 else ''}"
                    notif_body  = (f"**{row.requested_by}** returned all {len(returned_names)} items:\n"
                                   f"**{', '.join(returned_names)}**"
                                   + (f"\nCondition: {row.condition_note}" if row.condition_note else ""))

                existing = db.query(NexusNotification).filter(
                    NexusNotification.ref_id == row.order_id,
                    NexusNotification.recipient == (row.assigned_allocator_email or "").lower(),
                    NexusNotification.type == "item_returned",
                    NexusNotification.actioned == False,
                ).first()
                if existing:
                    existing.title   = notif_title
                    existing.body    = notif_body
                    existing.read_by = ""  # re-surface as unread
                else:
                    _notify(db, type="item_returned", recipient=row.assigned_allocator_email,
                            title=notif_title, body=notif_body,
                            ref_id=row.order_id, item_name=row.item_name, requested_by=row.requested_by)
            else:
                _notify(db, type="item_returned", recipient=row.assigned_allocator_email,
                        title=f"Item returned: {row.item_name}",
                        body=f"{row.requested_by} returned {row.item_name}. Condition: {row.condition_note or 'No notes.'}",
                        ref_id=checkout_id, item_name=row.item_name, requested_by=row.requested_by)
    elif body.status == "cancelled":
        # P1-7: when a MANAGER cancels someone else's request, tell the requester who
        # did it — a targeted "cancelled by {manager}" note, NOT a rejection. A plain
        # self-cancel needs no notification.
        if row.requested_by_email and row.requested_by_email.lower() != user["email"].lower():
            actor = _title_case_email(user["email"])
            _notify(db, type="cancelled", recipient=row.requested_by_email,
                    title=f"Checkout cancelled: {row.item_name}",
                    body=f"Your request for {row.item_name} was cancelled by {actor}.",
                    ref_id=checkout_id, item_name=row.item_name, requested_by=row.requested_by)

    db.commit()
    _fire_item_event(checkout_id, row.status, row.requested_by_email or "")
    return _checkout_to_dict(row)


# ── Extension requests ────────────────────────────────────────────────────────

class ExtensionRequest(BaseModel):
    days:   int
    reason: Optional[str] = ""


class ExtensionResolve(BaseModel):
    action: str                      # 'approve' | 'reject'
    note:   Optional[str] = ""


@router.post("/checkouts/{checkout_id}/extension")
def request_extension(checkout_id: str, body: ExtensionRequest, user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """Employee asks for more days on an item they currently hold."""
    # P1-2: lock the checkout row so two concurrent extension requests can't both pass
    # the "already pending" check below — otherwise they double-create the pending
    # extension and fire a duplicate broadcast to managers.
    row = db.query(ItemCheckout).filter(ItemCheckout.id == checkout_id).with_for_update().first()
    if not row:
        raise HTTPException(404, "Checkout not found")
    if row.status != "allocated":
        raise HTTPException(400, "Extensions can only be requested for items in use")
    if (row.requested_by_email or "").lower() != user["email"].lower() and not _is_items_manager(user, db):
        raise HTTPException(403, "Only the person holding the item can request an extension")
    if row.extension_status == "pending":
        raise HTTPException(400, "An extension request is already awaiting approval")
    if not (body.reason or "").strip():
        raise HTTPException(400, "A reason is required to request an extension")
    days = max(1, min(90, int(body.days or 1)))

    row.extension_days   = days
    row.extension_reason = (body.reason or "").strip()
    row.extension_status = "pending"

    # Neil: title leads with the ITEM and the days — the requester is already
    # shown on the card. Reason goes on its own line (bell renders pre-line).
    _notify(db, type="extension_pending", recipient="",
            title=f"Extension Request — {row.item_name} (+{days} day{'s' if days != 1 else ''})",
            body=f"{row.requested_by} requested {days} more day{'s' if days != 1 else ''} for {row.item_name}."
                 + (f"\nReason: {row.extension_reason}" if row.extension_reason else ""),
            ref_id=checkout_id, item_name=row.item_name, requested_by=row.requested_by)

    db.commit()
    _fire_item_event(checkout_id, row.status, row.requested_by_email or "")
    return _checkout_to_dict(row)


@router.post("/checkouts/{checkout_id}/extension/resolve")
def resolve_extension(checkout_id: str, body: ExtensionResolve, user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """Manager approves or rejects a pending extension request."""
    if not _is_items_manager(user, db):
        raise HTTPException(403, "Manager or above required to resolve extensions")
    # Row lock: two managers resolving the same extension concurrently must not
    # both pass the pending check (approve twice = days added twice).
    row = db.query(ItemCheckout).filter(ItemCheckout.id == checkout_id).with_for_update().first()
    if not row:
        raise HTTPException(404, "Checkout not found")
    if row.extension_status != "pending":
        raise HTTPException(400, "No pending extension request on this checkout")

    action = (body.action or "").lower().strip()
    if action not in ("approve", "reject"):
        raise HTTPException(400, "action must be 'approve' or 'reject'")

    ext_days = row.extension_days or 0
    if action == "approve":
        row.days = (row.days or 1) + ext_days
        _notify(db, type="extension_approved", recipient=row.requested_by_email,
                title=f"Extension approved: {row.item_name}",
                body=f"Your extension of {ext_days} day{'s' if ext_days != 1 else ''} for {row.item_name} was approved. New checkout period: {row.days} days.",
                ref_id=checkout_id, item_name=row.item_name, requested_by=row.requested_by)
    else:
        _notify(db, type="extension_declined", recipient=row.requested_by_email,
                title=f"Extension declined: {row.item_name}",
                body=f"Your extension request for {row.item_name} was declined."
                     + (f" Note: {body.note.strip()}" if (body.note or "").strip() else "")
                     + " Please return the item by the original due date.",
                ref_id=checkout_id, item_name=row.item_name, requested_by=row.requested_by)

    row.extension_days   = 0
    row.extension_reason = ""
    row.extension_status = ""

    # Clear the managers' extension_pending notification
    pending_notif = db.query(NexusNotification).filter(
        NexusNotification.type == "extension_pending",
        NexusNotification.actioned == False,
        NexusNotification.ref_id == checkout_id,
    ).first()
    if pending_notif:
        pending_notif.actioned = True

    db.commit()
    _fire_item_event(checkout_id, row.status, row.requested_by_email or "")
    return _checkout_to_dict(row)


# ── Reconcile lifecycle state ─────────────────────────────────────────────────

class ReconcileStateIn(BaseModel):
    # Optionally cancel a stuck pending_receipt checkout (allocator handed over, the
    # employee never confirmed receipt) as part of the same admin action. P1-3.
    cancel_checkout_id: Optional[str] = ""


@router.post("/{item_id}/reconcile-state")
def reconcile_item_state(item_id: str, body: ReconcileStateIn,
                         user: dict = Depends(require_items_admin), db: Session = Depends(get_db)):
    """P1-3: re-derive an item's lifecycle status from its LIVE checkouts/assignments,
    replacing the old raw `status` write (which could strand an item). Optionally
    rescues a stuck `pending_receipt` checkout by cancelling it (the only exit that
    state previously lacked). Manager-guarded (require_items_admin)."""
    item = db.query(Item).filter(Item.id == item_id).with_for_update().first()
    if not item or item.deleted_at:
        raise HTTPException(404, "Item not found")

    cancelled_id = ""
    if (body.cancel_checkout_id or "").strip():
        co = db.query(ItemCheckout).filter(
            ItemCheckout.id == body.cancel_checkout_id.strip(),
            ItemCheckout.item_id == item_id,
        ).with_for_update().first()
        if not co:
            raise HTTPException(404, "Checkout not found on this item")
        if co.status != "pending_receipt":
            raise HTTPException(409, "Only a stuck pending_receipt checkout can be cancelled here")
        co.status      = "cancelled"
        co.resolved_at = _now_iso()
        co.resolved_by = _title_case_email(user["email"])
        cancelled_id   = co.id
        # Tell the requester their unconfirmed handover was cancelled by a manager.
        if co.requested_by_email:
            _notify(db, type="cancelled", recipient=co.requested_by_email,
                    title=f"Checkout cancelled: {co.item_name}",
                    body=f"Your unconfirmed handover of {co.item_name} was cancelled by "
                         f"{_title_case_email(user['email'])} to free the item.",
                    ref_id=co.id, item_name=co.item_name, requested_by=co.requested_by)

    # Derive from live rows. autoflush is off, so a row cancelled just above is still
    # matched by a status filter in SQL — load by item and judge status in Python so
    # the just-cancelled row is correctly seen as no longer live.
    assignments = db.query(ItemAssignment).filter(ItemAssignment.item_id == item_id).all()
    checkouts   = db.query(ItemCheckout).filter(ItemCheckout.item_id == item_id).all()
    live_active_assign = any(a.status == "active" for a in assignments)
    live_allocated_co  = any(c.status == "allocated" for c in checkouts)

    if live_active_assign:
        new_status = "permanently_assigned"
    elif live_allocated_co:
        new_status = "checked_out"
    elif item.status == "retired":
        new_status = "retired"   # a terminal retire stands unless something live overrides it
    else:
        new_status = "available"

    old_status  = item.status
    item.status = new_status
    db.commit()
    return {"ok": True, "item": _item_to_dict(item),
            "previousStatus": old_status, "status": new_status,
            "cancelledCheckoutId": cancelled_id}


# ── AI photo fill ─────────────────────────────────────────────────────────────
# Claude (with web search) finds the manufacturer's product image for items
# missing a photo; we download it into our own Supabase bucket so the catalog
# never hot-links external URLs. These are stock photos — managers can replace
# them with real unit photos via Assign Photos at any time.

_ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
_GOOGLE_CSE_KEY    = os.getenv("GOOGLE_CSE_KEY", "")   # Google Custom Search JSON API key
_GOOGLE_CSE_CX     = os.getenv("GOOGLE_CSE_CX", "")    # Programmable Search Engine id (image search on)
_IMG_CTYPES = {"image/jpeg": "jpg", "image/png": "png", "image/webp": "webp", "image/gif": "gif"}


def _google_image_candidates(query: str, errors: Optional[list] = None) -> list:
    """Direct product-image URLs via the official Google image search API —
    the most reliable source when configured (retail sites bot-wall scrapers).
    Failures are recorded in `errors` so quota exhaustion is visible instead of
    masquerading as "no results"."""
    if not _GOOGLE_CSE_KEY or not _GOOGLE_CSE_CX:
        return []
    try:
        with httpx.Client(timeout=20) as client:
            r = client.get("https://www.googleapis.com/customsearch/v1", params={
                "key": _GOOGLE_CSE_KEY, "cx": _GOOGLE_CSE_CX,
                "q": query, "searchType": "image", "num": 5, "safe": "active",
            })
            if r.status_code != 200:
                if errors is not None:
                    try:
                        msg = r.json().get("error", {}).get("message", "")[:160]
                    except Exception:
                        msg = r.text[:160]
                    errors.append(f"google {r.status_code}: {msg}")
                return []
            return [it.get("link", "") for it in r.json().get("items", []) if it.get("link")]
    except Exception as e:
        if errors is not None:
            errors.append(f"google exc: {str(e)[:120]}")
        return []


def _openverse_image_candidates(query: str) -> list:
    """Keyless CC-image fallback — usually a photo of a similar model rather
    than the exact product render, but beats no photo at all."""
    try:
        with httpx.Client(timeout=20) as client:
            r = client.get("https://api.openverse.org/v1/images/", params={"q": query, "page_size": 5})
            if r.status_code != 200:
                return []
            return [res.get("url", "") for res in r.json().get("results", []) if res.get("url")]
    except Exception:
        return []


def _find_product_page_urls(item) -> list:
    """Ask Claude (web search enabled) for candidate product pages — we then try
    each one's preview image until one validates. Amazon is excluded (bot-walls
    every server-side fetch); retail/manufacturer pages with og:image work."""
    import re
    desc = " ".join(x for x in [item.make, item.model, item.name] if x).strip() or item.name
    payload = {
        "model": "claude-haiku-4-5-20251001",
        "max_tokens": 1024,
        "tools": [{"type": "web_search_20250305", "name": "web_search", "max_uses": 3}],
        "messages": [{
            "role": "user",
            "content": (
                f"Search the web for this product: {desc}. "
                "Give me up to 3 product page URLs that clearly show this product — prefer retailer "
                "product pages (Grainger, Acme Tools, Zoro, Toolstop, eBay listings, CDW, B&H) or the "
                "manufacturer's page. Do NOT use amazon.com links (they block automated access). "
                "Direct image URLs (.jpg/.png/.webp) are even better if you find them. "
                "Reply with ONLY the URLs, one per line, nothing else. If you find nothing, reply NONE."
            ),
        }],
    }
    import time
    data = None
    with httpx.Client(timeout=60) as client:
        for attempt in range(3):
            r = client.post(
                "https://api.anthropic.com/v1/messages",
                headers={"x-api-key": _ANTHROPIC_API_KEY, "anthropic-version": "2023-06-01", "content-type": "application/json"},
                json=payload,
            )
            if r.status_code == 429 and attempt < 2:
                # Tier-1 API keys have tight per-minute limits — wait and retry,
                # but capped so a single item never approaches the HTTP timeout
                wait = min(float(r.headers.get("retry-after", 15)), 30)
                time.sleep(wait)
                continue
            r.raise_for_status()
            data = r.json()
            break
    if data is None:
        return []
    text = "".join(b.get("text", "") for b in data.get("content", []) if b.get("type") == "text").strip()
    urls = [u.rstrip(").,]") for u in re.findall(r"https?://[^\s\"'<>]+", text)]
    # Drop amazon links if Claude ignored the instruction
    return [u for u in urls if "amazon." not in u.lower()][:3]


# Image URLs that are obviously NOT product photos (site chrome) — reject them
_BAD_IMG_HINTS = ("logo", "sprite", "icon", "favicon", "placeholder", "badge", "banner")


def _looks_like_junk_image(url: str) -> bool:
    low = url.lower()
    return any(h in low for h in _BAD_IMG_HINTS) or low.endswith((".svg", ".gif"))


def _fetch_image_to_storage(img_url: str, item_id: str, _depth: int = 0) -> str:
    """Download an image (following one og:image hop if Claude gave a page URL)
    and store it in the item-photos bucket. Returns the public URL or ''."""
    import re
    from urllib.parse import urljoin
    if _depth > 1 or not img_url:
        return ""
    # verify=False ONLY for third-party product sites — many retail/manufacturer
    # sites serve incomplete cert chains that strict validation rejects (browsers
    # AIA-fetch the intermediates, Python doesn't). Risk is negligible here: the
    # payload is a public stock photo, content-type and size validated below,
    # and reviewed by a manager. Our Anthropic/Supabase calls keep strict TLS.
    with httpx.Client(timeout=10, follow_redirects=True, verify=False) as client:
        try:
            resp = client.get(img_url, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36",
                "Accept": "text/html,image/avif,image/webp,image/*,*/*;q=0.8",
            })
            if resp.status_code in (403, 429):
                # Wikimedia (and friends) block fake browser UAs from datacenter
                # IPs but serve honest, descriptive bot UAs per their policy
                resp = client.get(img_url, headers={
                    "User-Agent": "GreensNexusCatalogBot/1.0 (+https://nexus.greensglobal.com; internal asset catalog)",
                    "Accept": "image/*,text/html;q=0.5",
                })
        except httpx.HTTPError:
            return ""  # slow/dead host — fail fast, the caller tries the next candidate
        if resp.status_code != 200:
            return ""
        ctype = resp.headers.get("content-type", "").split(";")[0].strip().lower()
        if ctype.startswith("text/html"):
            # Got a product page instead of an image — collect candidate images in
            # quality order: JSON-LD product image, og:image, twitter:image. Skip
            # obvious site chrome (logos, icons, .svg/.gif).
            html = resp.text
            candidates = []
            for block in re.findall(r'<script[^>]+application/ld\+json[^>]*>(.*?)</script>', html, re.S):
                m = re.search(r'"image"\s*:\s*\[?\s*"(https?://[^"]+)"', block)
                if m:
                    candidates.append(m.group(1))
            for pat in (
                r'<meta[^>]+property=["\']og:image(?::secure_url)?["\'][^>]+content=["\']([^"\']+)',
                r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+property=["\']og:image(?::secure_url)?["\']',
                r'<meta[^>]+name=["\']twitter:image["\'][^>]+content=["\']([^"\']+)',
                r'<link[^>]+rel=["\']image_src["\'][^>]+href=["\']([^"\']+)',
            ):
                m = re.search(pat, html)
                if m:
                    candidates.append(m.group(1))
            for cand in candidates:
                resolved = urljoin(str(resp.url), cand)
                if _looks_like_junk_image(resolved):
                    continue
                got = _fetch_image_to_storage(resolved, item_id, _depth + 1)
                if got:
                    return got
            return ""
        if ctype not in _IMG_CTYPES:
            return ""
        content = resp.content
        if len(content) < 8 * 1024 or len(content) > 8 * 1024 * 1024:
            return ""  # tracking pixel / logo-sized, or unreasonably large
        path = f"item-photos/ai-{item_id}-{uuid.uuid4().hex[:6]}.{_IMG_CTYPES[ctype]}"
    # Upload on a SEPARATE client with strict TLS — verify=False above is only
    # for the third-party retail sites, never for our own infrastructure.
    with httpx.Client(timeout=30) as upload_client:
        up = upload_client.post(
            f"{_SUPABASE_URL}/storage/v1/object/item-photos/{path}",
            headers={
                "Authorization": f"Bearer {_SUPABASE_SERVICE_KEY}",
                "apikey": _SUPABASE_SERVICE_KEY,
                "Content-Type": ctype,
                "x-upsert": "true",
                "cache-control": "max-age=31536000",  # unique paths — browsers cache immutably
            },
            content=content,
        )
        if up.status_code not in (200, 201):
            return ""
        return f"{_SUPABASE_URL}/storage/v1/object/public/item-photos/{path}"


class AutoPhotoRequest(BaseModel):
    item_ids: list[str]
    replace:  bool = False   # True → overwrite existing photos (manager selected specific rows)


@router.post("/auto-photos")
def auto_fill_photos(body: AutoPhotoRequest, user: dict = Depends(require_items_admin), db: Session = Depends(get_db)):
    if not _ANTHROPIC_API_KEY:
        raise HTTPException(503, "AI photo fill is not configured — add ANTHROPIC_API_KEY to the backend app settings")
    if not _SUPABASE_URL or not _SUPABASE_SERVICE_KEY:
        raise HTTPException(503, "Supabase storage is not configured on the backend")

    results = []
    for item_id in body.item_ids[:5]:  # cap per call — the client batches
        item = db.query(Item).filter(Item.id == item_id).first()
        if not item:
            results.append({"item_id": item_id, "status": "not_found"})
            continue
        if item.photo_url and not body.replace:
            results.append({"item_id": item_id, "status": "already_has_photo"})
            continue
        try:
            desc = " ".join(x for x in [item.make, item.model, item.name] if x).strip() or item.name
            # SPEED CONTRACT: with Google configured the pipeline is Google →
            # Openverse, both ~1-2s calls — an item succeeds or fails FAST.
            # Claude web search (rate-limit sleeps, up to a minute per item) is
            # only ever used when no Google key exists at all.
            # Query relaxation: "Make Model Name" can be too specific for the
            # site-restricted index — fall back to the bare item name.
            src_errors = []
            sources = _google_image_candidates(desc, src_errors)
            if not sources and desc != item.name:
                sources = _google_image_candidates(item.name, src_errors)
            if not sources:
                sources = _openverse_image_candidates(desc)
            if not sources and desc != item.name:
                sources = _openverse_image_candidates(item.name)
            if not sources and not _GOOGLE_CSE_KEY:
                try:
                    sources = _find_product_page_urls(item)
                except Exception:
                    sources = []
            if not sources:
                results.append({"item_id": item_id, "status": "no_image", "item_name": item.name,
                                "detail": "; ".join(src_errors[:2]) or "all sources returned zero results"})
                continue
            public_url = ""
            tried = []
            for src in sources[:3]:  # at most 3 download attempts per item
                public_url = _fetch_image_to_storage(src, item.id)
                tried.append(src)
                if public_url:
                    break
            if not public_url:
                results.append({"item_id": item_id, "status": "download_failed", "item_name": item.name,
                                "detail": f"no usable image on: {' | '.join(t[:200] for t in tried)}"})
                continue
            item.photo_url = public_url
            db.commit()
            results.append({"item_id": item_id, "status": "ok", "photo_url": public_url, "item_name": item.name})
        except Exception as e:
            db.rollback()
            results.append({"item_id": item_id, "status": "error", "detail": str(e)[:200], "item_name": item.name})
    return {"results": results}



# -- Permanent assignments ------------------------------------------------------

def _camel(snake: str) -> str:
    parts = snake.split("_")
    return parts[0] + "".join(p.title() for p in parts[1:])


def _assignment_to_dict(a: ItemAssignment) -> dict:
    return {_camel(c.name): (getattr(a, c.name) or "") for c in ItemAssignment.__table__.columns}


class AssignmentCreate(BaseModel):
    assignee_email: str
    assignee_name:  Optional[str] = ""
    # Manager option: activate immediately instead of waiting for the assignee
    # to accept. Permanent assignments only (this model is only used there).
    skip_acceptance: Optional[bool] = False


class AssignmentAccept(BaseModel):
    photo_url:  Optional[str] = ""
    photo_name: Optional[str] = ""
    note:       Optional[str] = ""


class AssignmentReturnInit(BaseModel):
    reason:     str = "normal"     # normal | dead | lost
    photo_url:  Optional[str] = ""
    photo_name: Optional[str] = ""
    note:       Optional[str] = ""


class AssignmentReturnAccept(BaseModel):
    disposition: str = "stock"     # stock | retired


_LIVE_ASSIGN = ["pending_acceptance", "active", "return_initiated"]


def _now_iso():
    return datetime.now(timezone.utc).isoformat()


def _action_notif(db: Session, ntype: str, ref_id: str):
    n = db.query(NexusNotification).filter(
        NexusNotification.type == ntype,
        NexusNotification.ref_id == ref_id,
        NexusNotification.actioned == False,
    ).first()
    if n:
        n.actioned = True


def _activate_assignment_direct(a: "ItemAssignment", item: "Item", now: str):
    """Manager skipped acceptance: flip the assignment straight to active and
    stamp the item pointers exactly the way accept_assignment does."""
    a.status = "active"
    a.accepted_at = now
    a.accept_note = "Assigned directly by manager — acceptance skipped"
    if item:
        item.status = "permanently_assigned"
        item.ownership_type = "permanent"
        item.assigned_to_email, item.assigned_to_name, item.assigned_at = a.assignee_email, a.assignee_name, now
        item.assigned_to_location = ""  # a person now holds it — drop any prior location assignment


def _clear_batch_assign_notif(db: Session, assignee_email: str):
    """P1-4: retire a lingering bulk-assign summary bell for this assignee. The batch
    "N items assigned to you" perm_assign carries the first assignment's id as ref_id;
    _action_notif only clears the single-assign notif that matches the exact assignment
    just handled, so the summary would otherwise stick around until (if ever) the FIRST
    listed item is acted on. Its title starts with the item count (a digit), whereas
    single-assign / reassign notifs start with text ("Item assigned…", "Please return…"),
    so we can safely clear the summary once the assignee acts on ANY item in the batch."""
    email = (assignee_email or "").lower()
    for n in db.query(NexusNotification).filter(
        NexusNotification.type == "perm_assign",
        NexusNotification.recipient == email,
        NexusNotification.actioned == False,
    ).all():
        if n.title[:1].isdigit():
            n.actioned = True


@router.get("/assignments")
def list_assignments(user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    q = db.query(ItemAssignment).order_by(ItemAssignment.created_at.desc())
    if not _is_items_manager(user, db):
        q = q.filter(ItemAssignment.assignee_email == user["email"])
    return [_assignment_to_dict(a) for a in q.limit(1000).all()]


@router.post("/{item_id}/assign", status_code=201)
def assign_item(item_id: str, body: AssignmentCreate, user: dict = Depends(require_items_admin), db: Session = Depends(get_db)):
    # P1-1: lock the item row so two concurrent assigns can't both pass the
    # live-assignment / active-checkout checks below and double-assign the same unit.
    item = db.query(Item).filter(Item.id == item_id).with_for_update().first()
    if not item or item.deleted_at:
        raise HTTPException(404, "Item not found")
    if item.ownership_type != "permanent":
        # Accepting an assignment force-flips ownership to permanent — a silent
        # one-way change that pulls the item out of the checkout catalogue.
        # Require the explicit ownership edit first instead.
        raise HTTPException(400, f'"{item.name}" is a temporary item — it gets checked out, not permanently assigned. '
                                 "Change its ownership to Permanent first if it should live with one person.")
    live = db.query(ItemAssignment).filter(
        ItemAssignment.item_id == item_id, ItemAssignment.status.in_(_LIVE_ASSIGN)).first()
    if live:
        raise HTTPException(409, "Item already has a live assignment - use reassign")
    co = db.query(ItemCheckout).filter(
        ItemCheckout.item_id == item_id,
        ItemCheckout.status.in_(["pending", "approved", "pending_receipt", "allocated"])).first()
    if co:
        raise HTTPException(409, "Item has an active checkout - recover it first")
    a = ItemAssignment(
        id=f"ASG-{uuid.uuid4().hex[:10].upper()}", item_id=item_id, item_name=item.name,
        assignee_email=body.assignee_email.lower().strip(),
        assignee_name=(body.assignee_name or "").strip(),
        assigned_by=_title_case_email(user["email"]), assigned_by_email=user["email"],
        status="pending_acceptance", created_at=_now_iso(),
    )
    db.add(a)
    if body.skip_acceptance:
        _activate_assignment_direct(a, item, _now_iso())
        _notify(db, type="perm_assign", recipient=a.assignee_email,
                title=f"Item assigned to you: {item.name}",
                body=f"{a.assigned_by or 'A manager'} assigned {item.name} to you permanently. It's already active in My Items — nothing to accept.",
                ref_id=a.id, item_name=item.name, requested_by=a.assignee_name)
    else:
        _notify(db, type="perm_assign", recipient=a.assignee_email,
                title=f"Item assigned to you: {item.name}",
                body=f"{a.assigned_by or 'A manager'} assigned {item.name} to you permanently. Please accept it with a photo in My Items.",
                ref_id=a.id, item_name=item.name, requested_by=a.assignee_name)
    db.commit()
    return _assignment_to_dict(a)


@router.post("/{item_id}/reassign", status_code=201)
def reassign_item(item_id: str, body: AssignmentCreate, user: dict = Depends(require_items_admin), db: Session = Depends(get_db)):
    """Start the return flow for the current holder; accepting that return
    auto-creates the next assignment for the new person."""
    item = db.query(Item).filter(Item.id == item_id).first()
    if not item:
        raise HTTPException(404, "Item not found")
    cur = db.query(ItemAssignment).filter(
        ItemAssignment.item_id == item_id, ItemAssignment.status == "active").with_for_update().first()
    if not cur:
        raise HTTPException(409, "No active assignment on this item - use assign")
    # Reassigning to the same person it's already with is a no-op that would
    # pointlessly bounce the item through a return — reject it (Neil).
    if body.assignee_email.lower().strip() == (cur.assignee_email or "").lower().strip():
        raise HTTPException(409, f"{item.name} is already assigned to {cur.assignee_name or cur.assignee_email}.")
    cur.status = "return_initiated"
    cur.return_reason = "reassign"
    cur.return_initiated_at = _now_iso()
    cur.next_assignee_email = body.assignee_email.lower().strip()
    cur.next_assignee_name  = (body.assignee_name or "").strip()
    _notify(db, type="perm_assign", recipient=cur.assignee_email,
            title=f"Please return: {item.name}",
            body=f"{item.name} is being reassigned to {cur.next_assignee_name or cur.next_assignee_email}. Please return it with a photo from My Items.",
            ref_id=cur.id, item_name=item.name, requested_by=cur.assignee_name)
    db.commit()
    return _assignment_to_dict(cur)


class AssignLocationIn(BaseModel):
    location: str = ""


@router.post("/{item_id}/assign-location")
def assign_item_to_location(item_id: str, body: AssignLocationIn, user: dict = Depends(require_items_admin), db: Session = Depends(get_db)):
    """Set WHERE an item lives. Location is just physical placement now (Visesh):
    it's independent of who holds it, so this never touches the person assignment or
    the lifecycle — an item can be at GG Corp AND assigned to a person. Empty clears
    the location. (Kept the legacy assigned_to_location in sync for older views.)"""
    item = db.query(Item).filter(Item.id == item_id).first()
    if not item or item.deleted_at:
        raise HTTPException(404, "Item not found")
    loc = (body.location or "").strip()
    item.location = loc
    item.assigned_to_location = ""  # retire the legacy "assigned to a place" flag
    db.commit()
    return _item_to_dict(item)


def _bulk_assign_blocked(db: Session, ids: list[str]) -> set:
    """Ids that can't be (re)assigned because a live person-assignment or an active
    checkout is in flight — recover those first."""
    blocked = set()
    for a in db.query(ItemAssignment).filter(
            ItemAssignment.item_id.in_(ids), ItemAssignment.status.in_(_LIVE_ASSIGN)).all():
        blocked.add(a.item_id)
    for c in db.query(ItemCheckout).filter(
            ItemCheckout.item_id.in_(ids),
            ItemCheckout.status.in_(["pending", "approved", "pending_receipt", "allocated"])).all():
        blocked.add(c.item_id)
    return blocked


class BulkAssignLocationIn(BaseModel):
    ids:      list[str]
    location: str = ""


@router.post("/bulk-assign-location")
def bulk_assign_to_location(body: BulkAssignLocationIn, user: dict = Depends(require_items_admin), db: Session = Depends(get_db)):
    """Set WHERE many items live at once. Location is physical placement only
    (Visesh) — independent of who holds them — so this just sets `location` and never
    touches the person assignment or lifecycle. Empty clears it."""
    ids = [i for i in (body.ids or []) if i]
    loc = (body.location or "").strip()
    if not ids:
        return {"assigned": 0, "skipped": []}
    assigned = 0
    for it in db.query(Item).filter(Item.id.in_(ids), or_(Item.deleted_at.is_(None), Item.deleted_at == "")).all():
        it.location = loc
        it.assigned_to_location = ""
        assigned += 1
    db.commit()
    return {"assigned": assigned, "skipped": []}


class BulkAssignPersonIn(BaseModel):
    ids:            list[str]
    assignee_email: str
    assignee_name:  str = ""
    skip_acceptance: bool = False


@router.post("/bulk-assign")
def bulk_assign_to_person(body: BulkAssignPersonIn, user: dict = Depends(require_items_admin), db: Session = Depends(get_db)):
    """Assign many items to a PERSON at once. Each item gets its own pending
    acceptance assignment (the assignee accepts with a photo from My Items); the
    assignee is notified once with the total, not per item (batch-notification rule)."""
    ids = [i for i in (body.ids or []) if i]
    email = (body.assignee_email or "").lower().strip()
    name  = (body.assignee_name or "").strip()
    if not ids or not email:
        return {"assigned": 0, "skipped": []}
    blocked = _bulk_assign_blocked(db, ids)
    assigned, names = 0, []
    first_asg_id = ""
    # P1-1: lock the item rows so a concurrent single-assign / checkout can't slip a
    # live claim in between the block scan above and the inserts below.
    for it in db.query(Item).filter(
            Item.id.in_(ids), or_(Item.deleted_at.is_(None), Item.deleted_at == "")).with_for_update().all():
        if it.ownership_type != "permanent":
            blocked.add(it.id)   # temporary items are checked out, never person-assigned
            continue
        if it.id in blocked:
            continue
        asg_id = f"ASG-{uuid.uuid4().hex[:10].upper()}"
        if not first_asg_id:
            first_asg_id = asg_id
        asg = ItemAssignment(
            id=asg_id, item_id=it.id, item_name=it.name,
            assignee_email=email, assignee_name=name,
            assigned_by=_title_case_email(user["email"]), assigned_by_email=user["email"],
            status="pending_acceptance", created_at=_now_iso(),
        )
        db.add(asg)
        if body.skip_acceptance:
            _activate_assignment_direct(asg, it, _now_iso())
        assigned += 1
        names.append(it.name)
    if assigned:
        preview = ", ".join(names[:5]) + ("…" if len(names) > 5 else "")
        # P1-4: give the batch bell a stable ref_id (the first assignment id) so it can
        # be deep-linked AND auto-cleared — with ref_id="" it could never be actioned
        # and lingered forever. Accepting/declining ANY item from the batch clears it
        # (see _clear_batch_assign_notif in accept/decline).
        tail = ("They're already active in My Items — nothing to accept."
                if body.skip_acceptance else "Accept each with a photo in My Items.")
        _notify(db, type="perm_assign", recipient=email,
                title=f"{assigned} item{'s' if assigned != 1 else ''} assigned to you",
                body=f"{_title_case_email(user['email'])} assigned you {assigned} item{'s' if assigned != 1 else ''}: {preview}. {tail}",
                ref_id=first_asg_id, item_name=names[0] if names else "", requested_by=name)
    db.commit()
    return {"assigned": assigned, "skipped": sorted(blocked & set(ids))}


@router.post("/assignments/{assignment_id}/accept")
def accept_assignment(assignment_id: str, body: AssignmentAccept, user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    _validate_photo_url(body.photo_url, "photo_url")
    a = db.query(ItemAssignment).filter(ItemAssignment.id == assignment_id).with_for_update().first()
    if not a:
        raise HTTPException(404, "Assignment not found")
    if a.assignee_email != user["email"]:
        raise HTTPException(403, "Only the assignee can accept this assignment")
    if a.status != "pending_acceptance":
        raise HTTPException(409, "Assignment is not awaiting acceptance")
    # Photo is OPTIONAL on permanent-assignment acceptance (Neil, Jul 17) —
    # checkout handover/receipt/return photos remain mandatory elsewhere.
    a.status = "active"
    a.accept_photo_url, a.accept_photo_name = body.photo_url or "", body.photo_name or ""
    a.accept_note, a.accepted_at = (body.note or "").strip(), _now_iso()
    item = db.query(Item).filter(Item.id == a.item_id).first()
    if item:
        item.status = "permanently_assigned"
        item.ownership_type = "permanent"
        item.assigned_to_email, item.assigned_to_name, item.assigned_at = a.assignee_email, a.assignee_name, a.accepted_at
        item.assigned_to_location = ""  # a person now holds it — drop any prior location assignment
    note_part = f' Condition note: "{a.accept_note}"' if a.accept_note else ""
    _notify(db, type="perm_update", recipient=a.assigned_by_email,
            title=f"Assignment accepted: {a.item_name}",
            body=f"{a.assignee_name or a.assignee_email} accepted {a.item_name}.{note_part}",
            ref_id=a.id, item_name=a.item_name, requested_by=a.assignee_name)
    _action_notif(db, "perm_assign", a.id)
    _clear_batch_assign_notif(db, a.assignee_email)   # P1-4: clear the bulk-assign summary
    db.commit()
    return _assignment_to_dict(a)


@router.post("/assignments/{assignment_id}/decline")
def decline_assignment(assignment_id: str, body: AssignmentReturnInit, user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    a = db.query(ItemAssignment).filter(ItemAssignment.id == assignment_id).with_for_update().first()
    if not a:
        raise HTTPException(404, "Assignment not found")
    if a.assignee_email != user["email"]:
        raise HTTPException(403, "Only the assignee can decline")
    if a.status != "pending_acceptance":
        raise HTTPException(409, "Assignment is not awaiting acceptance")
    a.status = "declined"
    a.return_note = (body.note or "").strip()
    _notify(db, type="perm_update", recipient=a.assigned_by_email,
            title=f"Assignment declined: {a.item_name}",
            body=f"{a.assignee_name or a.assignee_email} declined {a.item_name}." + (f' Reason: "{a.return_note}"' if a.return_note else ""),
            ref_id=a.id, item_name=a.item_name, requested_by=a.assignee_name)
    _action_notif(db, "perm_assign", a.id)
    _clear_batch_assign_notif(db, a.assignee_email)   # P1-4: clear the bulk-assign summary
    db.commit()
    return _assignment_to_dict(a)


@router.post("/assignments/{assignment_id}/initiate-return")
def initiate_assignment_return(assignment_id: str, body: AssignmentReturnInit, user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    _validate_photo_url(body.photo_url, "photo_url")
    a = db.query(ItemAssignment).filter(ItemAssignment.id == assignment_id).with_for_update().first()
    if not a:
        raise HTTPException(404, "Assignment not found")
    if a.assignee_email != user["email"] and not _is_items_manager(user, db):
        raise HTTPException(403, "Only the assignee or a manager can initiate a return")
    if a.status not in ("active", "return_initiated"):
        raise HTTPException(409, "Assignment is not active")
    reason = (body.reason or "normal").lower()
    if reason not in ("normal", "dead", "lost"):
        raise HTTPException(400, "reason must be normal, dead or lost")
    if reason != "lost" and not body.photo_url and a.status == "active":
        raise HTTPException(400, "A return photo is required (unless the item is lost)")
    keep_chain = a.return_reason == "reassign"   # reassignment return keeps its target
    a.status = "return_initiated"
    if not keep_chain:
        a.return_reason = reason
    a.return_photo_url, a.return_photo_name = body.photo_url or "", body.photo_name or ""
    a.return_note = (body.note or "").strip()
    a.return_initiated_at = _now_iso()
    flag = {"dead": "ITEM DEAD - ", "lost": "ITEM LOST - ", "reassign": "Reassignment - "}.get(a.return_reason, "")
    _notify(db, type="perm_return", recipient="",
            title=f"{flag}Return to confirm: {a.item_name}",
            body=f"{a.assignee_name or a.assignee_email} initiated a return of {a.item_name}"
                 + (f' - "{a.return_note}"' if a.return_note else ".")
                 + " Verify and accept it in Checkouts > Assignments.",
            ref_id=a.id, item_name=a.item_name, requested_by=a.assignee_name)
    db.commit()
    return _assignment_to_dict(a)


@router.post("/assignments/{assignment_id}/accept-return")
def accept_assignment_return(assignment_id: str, body: AssignmentReturnAccept, user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    if user["level"] < 2 and not _is_items_manager(user, db):
        raise HTTPException(403, "Supervisor or above required to accept returns")
    a = db.query(ItemAssignment).filter(ItemAssignment.id == assignment_id).with_for_update().first()
    if not a:
        raise HTTPException(404, "Assignment not found")
    if a.status != "return_initiated":
        raise HTTPException(409, "No return awaiting acceptance")
    dispo = body.disposition if body.disposition in ("stock", "retired") else "stock"
    a.status = "closed"
    a.disposition = dispo
    # Auth tokens carry only the email — derive a readable name ("Visesh Lodha",
    # not visesh.lodha@greensglobal.com) for everything user-facing.
    a.return_accepted_by, a.return_accepted_at = _title_case_email(user["email"]), _now_iso()
    item = db.query(Item).filter(Item.id == a.item_id).first()
    if item:
        item.assigned_to_email = item.assigned_to_name = item.assigned_at = ""
        item.status = "retired" if dispo == "retired" else "available"
    _notify(db, type="perm_update", recipient=a.assignee_email,
            title=f"Return accepted: {a.item_name}",
            body=f"Your return of {a.item_name} was accepted by {a.return_accepted_by}. You are no longer responsible for it.",
            ref_id=a.id, item_name=a.item_name, requested_by=a.assignee_name)
    _action_notif(db, "perm_return", a.id)
    # Reassignment chain: spawn the next assignment automatically — but ONLY if the
    # item actually ends up back in stock (available). P1-6: if the accepting manager
    # dispositioned it retired/dead/lost, the promised next assignee would otherwise be
    # silently dropped while the reassign modal claimed it was on its way. Tell both the
    # manager and the promised assignee it was cancelled, and clear next_assignee_* so
    # nothing dangles.
    if a.return_reason == "reassign" and a.next_assignee_email:
        if item and item.status == "available":
            nxt = ItemAssignment(
                id=f"ASG-{uuid.uuid4().hex[:10].upper()}", item_id=a.item_id, item_name=a.item_name,
                assignee_email=a.next_assignee_email, assignee_name=a.next_assignee_name,
                assigned_by=a.return_accepted_by, assigned_by_email=user["email"],
                status="pending_acceptance", created_at=_now_iso(),
            )
            db.add(nxt)
            _notify(db, type="perm_assign", recipient=nxt.assignee_email,
                    title=f"Item assigned to you: {a.item_name}",
                    body=f"{a.item_name} has been reassigned to you. Please accept it with a photo in My Items.",
                    ref_id=nxt.id, item_name=a.item_name, requested_by=nxt.assignee_name)
        else:
            dropped_name  = a.next_assignee_name or a.next_assignee_email
            dropped_email = a.next_assignee_email
            if a.assigned_by_email:
                _notify(db, type="perm_update", recipient=a.assigned_by_email,
                        title=f"Reassignment cancelled: {a.item_name}",
                        body=f"{a.item_name} was accepted back as '{dispo}', so the pending reassignment "
                             f"to {dropped_name} did not go ahead. Assign it again once it is back in service.",
                        ref_id=a.id, item_name=a.item_name, requested_by=a.assignee_name)
            _notify(db, type="perm_update", recipient=dropped_email,
                    title=f"Reassignment cancelled: {a.item_name}",
                    body=f"{a.item_name} was going to be reassigned to you but was taken out of service on return, "
                         "so it will not be coming to you. No action needed.",
                    ref_id=a.id, item_name=a.item_name, requested_by=dropped_name)
            a.next_assignee_email = ""
            a.next_assignee_name  = ""
    db.commit()
    return _assignment_to_dict(a)


def force_return_person(db: Session, email: str, actor_email: str) -> dict:
    """Force-return everything a departing person holds so no equipment stays out
    to a worker who has left. Called by HR offboarding — Items stays the single
    source of truth for these state transitions. Operates in the CALLER's session
    (the caller commits); no per-item notifications, as this is a bulk admin action.

      • Handed-over checkouts (pending_receipt / allocated) -> returned, item freed.
      • Not-yet-handed checkouts (pending / approved)       -> cancelled.
      • Live permanent assignments                          -> closed, item back to stock.

    Returns {"checkouts": n, "assignments": n}."""
    email = (email or "").strip().lower()
    if not email:
        return {"checkouts": 0, "assignments": 0}
    now = _now_iso()
    actor = _title_case_email(actor_email)

    checkouts = db.query(ItemCheckout).filter(
        func.lower(ItemCheckout.requested_by_email) == email,
        ItemCheckout.status.in_(["pending", "approved", "pending_receipt", "allocated"]),
    ).with_for_update().all()
    for c in checkouts:
        handed_over = c.status in ("pending_receipt", "allocated")
        c.status = "returned" if handed_over else "cancelled"
        if handed_over:
            c.returned_at = now
            c.condition_note = c.condition_note or f"Auto-returned on offboarding by {actor}"
            item = db.query(Item).filter(Item.id == c.item_id).first()
            if item and item.status == "checked_out":
                item.status = "available"

    assignments = db.query(ItemAssignment).filter(
        func.lower(ItemAssignment.assignee_email) == email,
        ItemAssignment.status.in_(_LIVE_ASSIGN),
    ).with_for_update().all()
    for a in assignments:
        a.status = "closed"
        a.disposition = "stock"
        a.return_accepted_by, a.return_accepted_at = actor, now
        item = db.query(Item).filter(Item.id == a.item_id).first()
        if item:
            item.assigned_to_email = item.assigned_to_name = item.assigned_at = ""
            item.status = "available"

    return {"checkouts": len(checkouts), "assignments": len(assignments)}


@router.post("/assignments/{assignment_id}/cancel")
def cancel_assignment(assignment_id: str, user: dict = Depends(require_items_admin), db: Session = Depends(get_db)):
    """Manager cancel / force-recover. Pending -> cancelled; active/returning -> closed, item back to stock."""
    a = db.query(ItemAssignment).filter(ItemAssignment.id == assignment_id).with_for_update().first()
    if not a:
        raise HTTPException(404, "Assignment not found")
    if a.status not in _LIVE_ASSIGN:
        raise HTTPException(409, "Assignment is already closed")
    was_pending = a.status == "pending_acceptance"
    a.status = "cancelled" if was_pending else "closed"
    a.disposition = "" if was_pending else "stock"
    # Auth tokens carry only the email — derive a readable name ("Visesh Lodha",
    # not visesh.lodha@greensglobal.com) for everything user-facing.
    a.return_accepted_by, a.return_accepted_at = _title_case_email(user["email"]), _now_iso()
    item = db.query(Item).filter(Item.id == a.item_id).first()
    if item and not was_pending:
        item.assigned_to_email = item.assigned_to_name = item.assigned_at = ""
        item.status = "available"
    _notify(db, type="perm_update", recipient=a.assignee_email,
            title=f"Assignment {'cancelled' if was_pending else 'closed'}: {a.item_name}",
            body=f"{a.return_accepted_by} {'cancelled the pending assignment of' if was_pending else 'force-recovered'} {a.item_name}."
                 + ("" if was_pending else " You are no longer responsible for it."),
            ref_id=a.id, item_name=a.item_name, requested_by=a.assignee_name)
    # P1-6: if this was a reassign-in-flight, a next assignee was promised the item.
    # Force-recover closes the assignment without spawning that next assignment, so tell
    # the manager and the promised assignee it was cancelled, and clear next_assignee_*.
    if a.return_reason == "reassign" and a.next_assignee_email:
        dropped_name  = a.next_assignee_name or a.next_assignee_email
        dropped_email = a.next_assignee_email
        if a.assigned_by_email:
            _notify(db, type="perm_update", recipient=a.assigned_by_email,
                    title=f"Reassignment cancelled: {a.item_name}",
                    body=f"{a.item_name} was force-recovered, so the pending reassignment to "
                         f"{dropped_name} did not go ahead. Assign it again if needed.",
                    ref_id=a.id, item_name=a.item_name, requested_by=a.assignee_name)
        _notify(db, type="perm_update", recipient=dropped_email,
                title=f"Reassignment cancelled: {a.item_name}",
                body=f"{a.item_name} was going to be reassigned to you but was recovered by a manager, "
                     "so it will not be coming to you. No action needed.",
                ref_id=a.id, item_name=a.item_name, requested_by=dropped_name)
        a.next_assignee_email = ""
        a.next_assignee_name  = ""
    _action_notif(db, "perm_assign", a.id)
    _action_notif(db, "perm_return", a.id)
    db.commit()
    return _assignment_to_dict(a)


# ── Allocators / Approvers ────────────────────────────────────────────────────

def _nexus_people_only(db: Session, rows):
    """Restrict role-holder rows to people on the curated Nexus People list
    (nexus_employees) and use their People name. Role grants can exist for any
    M365 account; pickers must only ever offer real Nexus people (Neil, Jul 17)."""
    people = {
        (e.work_email or "").lower(): f"{e.first_name} {e.last_name}".strip()
        for e in db.query(NexusEmployee)
                   .filter(NexusEmployee.status != "offboarded")
                   .filter(NexusEmployee.work_email != "").all()
    }
    out = []
    for r in rows:
        name = people.get((r.email or "").lower())
        if name is None:
            continue
        out.append({"email": r.email, "name": name or r.display_name or _title_case_email(r.email),
                    "role": r.role})
    return sorted(out, key=lambda p: p["name"].lower())


@router.get("/approvers")
def list_approvers(user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """Manager-level users an employee can address their checkout request to.
    Open to all authenticated users — only names/emails/roles are exposed."""
    rows = db.query(NexusRole).filter(NexusRole.role.in_(
        [role for role, level in _ROLE_LEVEL.items() if level >= _ROLE_LEVEL["manager"]]
    )).order_by(NexusRole.email).all()
    return _nexus_people_only(db, rows)


@router.get("/allocators")
def list_allocators(user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    if not _is_items_manager(user, db):
        raise HTTPException(403, "Manager or above required")
    rows = db.query(NexusRole).filter(NexusRole.role.in_(
        [role for role, level in _ROLE_LEVEL.items() if level >= _ROLE_LEVEL["supervisor"]]
    )).order_by(NexusRole.email).all()
    return _nexus_people_only(db, rows)


# ── Report ────────────────────────────────────────────────────────────────────

_REPORT_HEADERS = [
    "Item", "Type", "Make", "Model", "Department", "Location", "Owner",
    "Ownership", "Status", "Requested By", "Days", "Reason",
    "Request Date", "Allocated Date", "Allocated By", "Returned Date", "Condition",
]


def _report_rows(db: Session, *, department, item_type, status, requested_by=None):
    q = db.query(ItemCheckout, Item).outerjoin(
        Item, ItemCheckout.item_id == Item.id
    ).order_by(ItemCheckout.created_at.desc())
    if department:
        q = q.filter(ItemCheckout.department == department)
    if item_type:
        q = q.filter(func.lower(Item.item_type) == item_type.lower().strip())
    if status:
        q = q.filter(ItemCheckout.status == status)
    if requested_by:
        # Comma-separated names — match any (case-insensitive substring per name)
        names = [n.strip() for n in requested_by.split(",") if n.strip()]
        if names:
            q = q.filter(or_(*[ItemCheckout.requested_by.ilike(f"%{n}%") for n in names]))
    rows = []
    for c, item in q.all():
        rows.append([
            c.item_name, item.item_type if item else "", item.make if item else "",
            item.model if item else "", c.department, item.location if item else "",
            item.default_owner if item else "", item.ownership_type if item else "",
            c.status, c.requested_by, c.days, c.reason,
            c.created_at[:10] if c.created_at else "",
            c.allocated_at[:10] if c.allocated_at else "", c.allocated_by or "",
            c.returned_at[:10] if c.returned_at else "", c.condition_note or "",
        ])
    return rows


@router.get("/report")
def export_report(
    format:       str = "excel",
    department:   Optional[str] = None,
    item_type:    Optional[str] = None,
    status:       Optional[str] = None,
    requested_by: Optional[str] = None,
    user: dict = Depends(require_items_admin),
    db: Session = Depends(get_db),
):
    rows  = _report_rows(db, department=department, item_type=item_type, status=status, requested_by=requested_by)
    stamp = datetime.utcnow().strftime("%Y%m%d")

    if format == "pdf":
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
        except ImportError:
            raise HTTPException(500, "reportlab not installed")
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title="Items Checkout Report")
        table = Table([_REPORT_HEADERS] + rows, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A1A2E")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTSIZE",   (0, 0), (-1, -1), 6.5),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5FA")]),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ]))
        doc.build([table])
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=items_report_{stamp}.pdf"})

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")
    wb = Workbook()
    ws = wb.active
    ws.title = "Items Checkout Report"
    hf = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(_REPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hf
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max((len(str(c.value or "")) for c in col), default=0) + 4, 40
        )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=items_report_{stamp}.xlsx"},
    )


# ── Audit log ─────────────────────────────────────────────────────────────────

@router.get("/audit-log")
def items_audit_log(
    q:      Optional[str] = None,
    limit:  int = 100,
    offset: int = 0,
    user: dict = Depends(require_items_admin),
    db: Session = Depends(get_db),
):
    query = db.query(AuditLog).filter(AuditLog.resource_type.in_(["items", "item_checkouts", "inventory-requests"]))
    if q:
        needle = q.strip()
        query = query.filter(or_(
            AuditLog.details.contains(needle),
            AuditLog.user_email.contains(needle.lower()),
            AuditLog.action.contains(needle),
            AuditLog.resource_id.contains(needle),
        ))
    total = query.count()
    rows  = query.order_by(AuditLog.timestamp.desc()).offset(offset).limit(limit).all()
    return {
        "total": total,
        "rows": [
            {"id": r.id, "timestamp": r.timestamp, "user_email": r.user_email,
             "user_role": r.user_role, "action": r.action,
             "resource_id": r.resource_id, "details": r.details,
             "undone_at": r.undone_at or "", "undone_by": r.undone_by or ""}
            for r in rows
        ],
    }


# Columns the audit Undo is allowed to restore (1:1 with the audit field names the
# frontend tracks). Scalar columns only — no relations, no side effects.
# P1-3: lifecycle "status" removed — it is derived from checkouts/assignments, never
# restored as a raw string (an undo could otherwise strand an item's lifecycle state).
_UNDO_ITEM_COLS = {
    "name", "item_type", "make", "model", "year", "department", "location",
    "default_owner", "ownership_type", "serial_number",
    "op_status", "op_status_person_name", "asset_value", "photo_url",
}


class AuditUndoRequest(BaseModel):
    audit_id: int
    fields: Optional[dict] = None   # {audit_field_name: previous_value} for an edit reversal


@router.post("/audit-undo")
def undo_audit_entry(body: AuditUndoRequest, user: dict = Depends(require_items_admin), db: Session = Depends(get_db)):
    """Revert a single item add / edit / delete recorded in the audit log.

    Edits restore the previous values the client reconstructed from the log;
    adds are soft-deleted; deletes are restored. The original entry is marked
    undone and a new canonical entry (carrying a "reason") records the reversal.
    """
    row = db.query(AuditLog).filter(AuditLog.id == body.audit_id).first()
    if not row or row.resource_type != "items":
        raise HTTPException(404, "Audit entry not found")
    if row.undone_at:
        raise HTTPException(409, "This change was already undone")

    action = (row.action or "").lower()
    iid    = row.resource_id or ""
    email  = user["email"]
    now    = datetime.now(timezone.utc).isoformat()
    item   = db.query(Item).filter(Item.id == iid).first() if iid else None

    applied, reason, new_action = {}, "", ""

    if action.startswith("added item"):
        if not item:
            raise HTTPException(404, "That item no longer exists")
        if item.deleted_at:
            raise HTTPException(409, "That item is already deleted")
        active = db.query(ItemCheckout).filter(
            ItemCheckout.item_id == iid,
            ItemCheckout.status.in_(["pending", "approved", "pending_receipt", "allocated"]),
        ).count()
        if active:
            raise HTTPException(409, "Can't undo — the item has an active checkout against it")
        _soft_delete(item, email)
        reason, new_action = "Undo — removed the added item", f"Deleted item {iid}"

    elif action.startswith("deleted item"):
        if not item:
            raise HTTPException(404, "That item no longer exists")
        if not item.deleted_at:
            raise HTTPException(409, "That item is not deleted")
        item.deleted_at = ""
        item.deleted_by = ""
        reason, new_action = "Undo — restored the deleted item", f"Restored item {iid}"

    elif action.startswith("updated item"):
        if not item:
            raise HTTPException(404, "That item no longer exists")
        for k, v in (body.fields or {}).items():
            if k not in _UNDO_ITEM_COLS:
                continue
            if k == "asset_value":
                try:
                    v = float(v or 0)
                except (TypeError, ValueError):
                    v = 0.0
            setattr(item, k, v)
            applied[k] = v
        if not applied:
            raise HTTPException(400, "Nothing to restore for this change")
        # If op_status was reverted off a person-bound state, drop the stale person.
        if "op_status" in applied and applied["op_status"] not in _OP_STATUS_PERSON:
            item.op_status_person_email = ""
            item.op_status_person_name  = ""
        reason, new_action = "Undo — restored previous value(s)", f"Updated item {iid}"

    else:
        raise HTTPException(400, "This kind of change can't be undone")

    row.undone_at = now
    row.undone_by = email

    details = {"path": "/items/audit-undo", "status": 200, "reason": reason, "undo_of": row.id}
    details.update(applied)
    db.add(AuditLog(
        timestamp=now, user_email=email, user_role="",
        action=new_action, resource_type="items", resource_id=iid,
        details=json.dumps(details), ip_address="",
    ))
    db.commit()
    return {"ok": True, "undone_id": row.id}

