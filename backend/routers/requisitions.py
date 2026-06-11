import io
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
import models
from database import get_db
from auth import get_current_user, require_manager, require_administrator, require_level

router = APIRouter(tags=["Requisitions & Hardware Assets"], dependencies=[Depends(get_current_user)])


def _ts():
    return datetime.utcnow().isoformat()


# ── Schemas ──────────────────────────────────────────────────────────────────

class RequisitionCreate(BaseModel):
    id: str
    employee_name: str
    employee_email: str = ""    # beneficiary — on-behalf requests tag THEIR email so it lands in their log
    employee_dept: str
    item: str
    quantity: int = 1
    reason: str = ""
    status: str = "pending_manager"
    supervisor_name: str = ""
    approver_email: str = ""    # manager picked by the requester — only they get the notification


class RequisitionApprove(BaseModel):
    manager_name: str
    # Who will purchase & fulfill this — picked by the manager at approval
    # (mirrors the checkout allocator flow). Optional for backward compat.
    allocator_email: str = ""
    allocator_name:  str = ""


class RequisitionReject(BaseModel):
    manager_name: str
    rejection_reason: str


class RequisitionOrder(BaseModel):
    by_name: str
    note: str = ""           # vendor / expected arrival — shown to the requester


class RequisitionFulfill(BaseModel):
    by_name: str
    note: str = ""
    item_id: str = ""        # items.id when the purchase was added to inventory


class RequisitionAllocate(BaseModel):
    asset_id: str
    supervisor_name: str
    expected_return_date: str = ""


class RequisitionReturn(BaseModel):
    initiated_by: str


class RequisitionConfirmReturn(BaseModel):
    supervisor_name: str
    condition: str = "Available"
    return_photo_name: str = ""
    return_photo_url: str = ""


class RequisitionMarkLost(BaseModel):
    supervisor_name: str
    notes: str = ""


class HardwareAssetCreate(BaseModel):
    id: str
    name: str
    category: str
    serial_number: str = ""
    assigned_to: str = "Unassigned"
    dept: str = ""
    location: str = ""
    status: str = "Available"
    purchased: str = ""
    warranty_end: str = ""


# ── Requisition Routes ────────────────────────────────────────────────────────

@router.get("/requisitions")
def list_requisitions(
    user: dict = Depends(get_current_user),
    db:   Session = Depends(get_db),
):
    # Managers and above see all requisitions; employees/supervisors see only their own
    q = db.query(models.Requisition).order_by(models.Requisition.created_at.desc())
    if user["level"] < 3:
        q = q.filter(
            (models.Requisition.employee_email == user["email"]) |
            (models.Requisition.employee_email == "")  # legacy rows without email
        )

    reqs = q.all()
    result = []
    for r in reqs:
        # For legacy rows (no email), restrict by name match as fallback
        if user["level"] < 3 and not r.employee_email:
            name_part = user["email"].split("@")[0].lower()
            if name_part not in r.employee_name.lower():
                continue
        hist = (
            db.query(models.ApprovalHistory)
            .filter(models.ApprovalHistory.requisition_id == r.id)
            .order_by(models.ApprovalHistory.created_at)
            .all()
        )
        row = {c.name: getattr(r, c.name) for c in r.__table__.columns}
        row["history"] = [
            {"action": h.action, "by": h.action_by, "role": h.action_role,
             "comment": h.comment or "", "date": h.created_at}
            for h in hist
        ]
        result.append(row)
    return result


@router.post("/requisitions", status_code=201)
def create_requisition(
    data: RequisitionCreate,
    user: dict = Depends(get_current_user),
    db:   Session = Depends(get_db),
):
    payload = data.model_dump(exclude={"approver_email"})
    # Beneficiary tagging: an on-behalf request lands in the BENEFICIARY's log,
    # so honour a client-supplied email; default to the verified submitter.
    payload["employee_email"] = (data.employee_email or user["email"]).lower().strip()
    req = models.Requisition(**payload, created_at=_ts(), updated_at=_ts())
    db.add(req)

    # Notification is created HERE, server-side — employees can't write to the
    # notifications API (level gate), so the old frontend addNotification call
    # silently 403'd and managers never heard about employee requisitions.
    reason_snip = (data.reason or "").strip().split("\n")[0][:120]
    _req_notify(db,
        type="req_pending",
        recipient=(data.approver_email or "").lower().strip(),  # targeted; empty = all managers
        title="New Purchase Requisition",
        body=f"{data.employee_name} ({data.employee_dept}) requested {data.quantity}× {data.item}."
             + (f' Reason: "{reason_snip}"' if reason_snip else ""),
        ref_id=data.id, item_name=data.item, requested_by=data.employee_name,
    )
    # On-behalf: tell the beneficiary a request was raised for them
    if payload["employee_email"] and payload["employee_email"] != user["email"].lower():
        _req_notify(db,
            type="req_update",
            recipient=payload["employee_email"],
            title=f"Purchase request raised for you: {data.item}",
            body=f"{user.get('name') or 'A colleague'} submitted a purchase request for {data.quantity}× {data.item} on your behalf. It is pending manager approval.",
            ref_id=data.id, item_name=data.item, requested_by=data.employee_name,
        )

    db.commit()
    db.refresh(req)
    return req


def _req_notify(db: Session, *, type: str, recipient: str, title: str, body: str,
                ref_id: str = "", item_name: str = "", requested_by: str = "") -> None:
    from datetime import timezone as _tz
    import uuid as _uuid
    db.add(models.NexusNotification(
        id=str(_uuid.uuid4()), type=type,
        recipient=(recipient or "").lower().strip(),
        title=title, body=body,
        ref_id=ref_id, item_name=item_name, requested_by=requested_by,
        action="", actioned=False, read_by="",
        created_at=datetime.now(_tz.utc).isoformat(),
    ))


def _action_req_notification(db: Session, req_id: str):
    """Clear the req_pending bell entry once a manager resolves the requisition."""
    notif = db.query(models.NexusNotification).filter(
        models.NexusNotification.type == "req_pending",
        models.NexusNotification.ref_id == req_id,
        models.NexusNotification.actioned == False,
    ).first()
    if notif:
        notif.actioned = True


@router.patch("/requisitions/{req_id}/approve")
def approve_requisition(req_id: str, body: RequisitionApprove, user: dict = Depends(require_manager), db: Session = Depends(get_db)):
    req = db.query(models.Requisition).filter(models.Requisition.id == req_id).first()
    if not req:
        raise HTTPException(404, "Requisition not found")
    req.status = "manager_approved"
    req.manager_name = body.manager_name
    req.manager_approval_date = _ts()
    req.allocator_email = (body.allocator_email or "").lower().strip()
    req.allocator_name  = (body.allocator_name or "").strip()
    req.updated_at = _ts()
    db.add(models.ApprovalHistory(requisition_id=req_id, action="Approved", action_by=body.manager_name, action_role="Manager",
                                  comment=(f"Fulfillment: {req.allocator_name}" if req.allocator_name else ""), created_at=_ts()))
    _action_req_notification(db, req_id)
    if req.employee_email:
        _req_notify(db,
            type="req_approved",
            recipient=req.employee_email,
            title=f"Requisition approved: {req.item}",
            body=f"Your purchase request for {req.quantity}× {req.item} was approved by {body.manager_name}."
                 + (f"\n{req.allocator_name} will purchase it for you." if req.allocator_name else ""),
            ref_id=req_id, item_name=req.item, requested_by=req.employee_name,
        )
    # Targeted work item for the allocator — their bell deep-links to the
    # To Fulfill queue in Purchase Requests.
    if req.allocator_email:
        _req_notify(db,
            type="req_fulfill",
            recipient=req.allocator_email,
            title=f"Purchase to fulfill: {req.quantity}× {req.item}",
            body=f"{body.manager_name} approved {req.employee_name}'s request and assigned the purchase to you."
                 + (f"\nReason: {req.reason}" if (req.reason or "").strip() else ""),
            ref_id=req_id, item_name=req.item, requested_by=req.employee_name,
        )
    db.commit()
    db.refresh(req)
    return req


def _require_fulfiller(req, user):
    """Only the assigned allocator or a manager+ may act on fulfillment."""
    if user["level"] >= 3:
        return
    if (req.allocator_email or "").lower() != user["email"].lower():
        raise HTTPException(403, "Only the assigned fulfiller or a manager can do this")


@router.patch("/requisitions/{req_id}/mark-ordered")
def mark_ordered(req_id: str, body: RequisitionOrder, user: dict = Depends(require_level(2)), db: Session = Depends(get_db)):
    req = db.query(models.Requisition).filter(models.Requisition.id == req_id).first()
    if not req:
        raise HTTPException(404, "Requisition not found")
    if req.status != "manager_approved":
        raise HTTPException(409, f"Cannot mark a '{req.status}' requisition as ordered")
    _require_fulfiller(req, user)
    req.status = "ordered"
    req.ordered_at = _ts()
    req.updated_at = _ts()
    db.add(models.ApprovalHistory(requisition_id=req_id, action="Ordered", action_by=body.by_name, action_role="Fulfiller", comment=body.note or "", created_at=_ts()))
    if req.employee_email:
        _req_notify(db,
            type="req_update",
            recipient=req.employee_email,
            title=f"Ordered: {req.item}",
            body=f"{body.by_name} has placed the order for your {req.quantity}× {req.item}."
                 + (f"\n{body.note}" if (body.note or "").strip() else ""),
            ref_id=req_id, item_name=req.item, requested_by=req.employee_name,
        )
    db.commit()
    db.refresh(req)
    return req


@router.patch("/requisitions/{req_id}/fulfill")
def fulfill_requisition(req_id: str, body: RequisitionFulfill, user: dict = Depends(require_level(2)), db: Session = Depends(get_db)):
    req = db.query(models.Requisition).filter(models.Requisition.id == req_id).first()
    if not req:
        raise HTTPException(404, "Requisition not found")
    if req.status not in ("manager_approved", "ordered"):
        raise HTTPException(409, f"Cannot fulfill a '{req.status}' requisition")
    _require_fulfiller(req, user)
    req.status = "fulfilled"
    req.fulfilled_at = _ts()
    req.fulfillment_note = (body.note or "").strip()
    req.fulfilled_item_id = (body.item_id or "").strip()
    req.updated_at = _ts()
    db.add(models.ApprovalHistory(requisition_id=req_id, action="Fulfilled", action_by=body.by_name, action_role="Fulfiller",
                                  comment=(body.note or "") + (" · added to inventory" if body.item_id else ""), created_at=_ts()))
    # Clear the allocator's req_fulfill work item from the bell
    notif = db.query(models.NexusNotification).filter(
        models.NexusNotification.type == "req_fulfill",
        models.NexusNotification.ref_id == req_id,
        models.NexusNotification.actioned == False,
    ).first()
    if notif:
        notif.actioned = True
    if req.employee_email:
        _req_notify(db,
            type="req_update",
            recipient=req.employee_email,
            title=f"Ready for you: {req.item}",
            body=f"Your {req.quantity}× {req.item} has been purchased{' and added to inventory' if body.item_id else ''}."
                 + (f"\n{req.fulfillment_note}" if req.fulfillment_note else "")
                 + f"\nCoordinate with {body.by_name} to receive it.",
            ref_id=req_id, item_name=req.item, requested_by=req.employee_name,
        )
    db.commit()
    db.refresh(req)
    return req


@router.patch("/requisitions/{req_id}/reject")
def reject_requisition(req_id: str, body: RequisitionReject, user: dict = Depends(require_manager), db: Session = Depends(get_db)):
    req = db.query(models.Requisition).filter(models.Requisition.id == req_id).first()
    if not req:
        raise HTTPException(404, "Requisition not found")
    req.status = "rejected"
    req.manager_name = body.manager_name
    req.rejection_reason = body.rejection_reason
    req.updated_at = _ts()
    db.add(models.ApprovalHistory(requisition_id=req_id, action="Rejected", action_by=body.manager_name, action_role="Manager", comment=body.rejection_reason, created_at=_ts()))
    _action_req_notification(db, req_id)
    if req.employee_email:
        _req_notify(db,
            type="req_rejected",
            recipient=req.employee_email,
            title=f"Requisition rejected: {req.item}",
            body=f"Your purchase request for {req.quantity}× {req.item} was not approved. Reason: {body.rejection_reason or 'No reason given.'}",
            ref_id=req_id, item_name=req.item, requested_by=req.employee_name,
        )
    db.commit()
    db.refresh(req)
    return req


@router.patch("/requisitions/{req_id}/allocate")
def allocate_asset(req_id: str, body: RequisitionAllocate, user: dict = Depends(require_level(2)), db: Session = Depends(get_db)):
    req = db.query(models.Requisition).filter(models.Requisition.id == req_id).first()
    asset = db.query(models.HardwareAsset).filter(models.HardwareAsset.id == body.asset_id).first()
    if not req:
        raise HTTPException(404, "Requisition not found")
    if not asset:
        raise HTTPException(404, "Asset not found")
    if asset.status != "Available":
        raise HTTPException(400, "Asset not available")

    asset.status = "Checked Out"
    asset.assigned_to = req.employee_name
    asset.assigned_req_id = req_id
    asset.last_updated = _ts()[:10]

    req.status = "asset_allocated"
    req.asset_id = body.asset_id
    req.asset_name = asset.name
    req.asset_category = asset.category
    req.asset_serial = asset.serial_number
    req.asset_allocated_date = _ts()
    req.expected_return_date = body.expected_return_date
    req.allocated_by = body.supervisor_name
    req.updated_at = _ts()

    db.add(models.ApprovalHistory(requisition_id=req_id, action="Asset Allocated", action_by=body.supervisor_name, action_role="Supervisor", comment=f"{asset.name} ({body.asset_id})", created_at=_ts()))
    db.commit()
    db.refresh(req)
    return req


@router.patch("/requisitions/{req_id}/initiate-return")
def initiate_return(req_id: str, body: RequisitionReturn, user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    req = db.query(models.Requisition).filter(models.Requisition.id == req_id).first()
    if not req:
        raise HTTPException(404, "Requisition not found")
    if user["level"] < 2 and req.employee_email.lower() != user["email"]:
        raise HTTPException(403, "You can only return your own allocated items")
    if req.asset_id:
        asset = db.query(models.HardwareAsset).filter(models.HardwareAsset.id == req.asset_id).first()
        if asset:
            asset.status = "Return Pending"
            asset.last_updated = _ts()[:10]
    req.status = "return_initiated"
    req.updated_at = _ts()
    db.add(models.ApprovalHistory(requisition_id=req_id, action="Return Initiated", action_by=body.initiated_by, action_role="Employee", created_at=_ts()))
    db.commit()
    db.refresh(req)
    return req


@router.patch("/requisitions/{req_id}/confirm-return")
def confirm_return(req_id: str, body: RequisitionConfirmReturn, user: dict = Depends(require_level(2)), db: Session = Depends(get_db)):
    req = db.query(models.Requisition).filter(models.Requisition.id == req_id).first()
    if not req:
        raise HTTPException(404, "Requisition not found")
    if req.asset_id:
        asset = db.query(models.HardwareAsset).filter(models.HardwareAsset.id == req.asset_id).first()
        if asset:
            asset.status = body.condition
            asset.assigned_to = "Unassigned" if body.condition == "Available" else asset.assigned_to
            asset.assigned_req_id = ""
            asset.last_updated = _ts()[:10]
    req.status = "returned"
    req.actual_return_date = _ts()
    req.return_confirmed_by = body.supervisor_name
    req.return_asset_condition = body.condition
    req.return_photo_name = body.return_photo_name or ""
    req.return_photo_url  = body.return_photo_url  or ""
    req.updated_at = _ts()
    db.add(models.ApprovalHistory(requisition_id=req_id, action="Return Confirmed", action_by=body.supervisor_name, action_role="Supervisor", comment=f"Condition: {body.condition}", created_at=_ts()))
    db.commit()
    db.refresh(req)
    return req


@router.patch("/requisitions/{req_id}/mark-lost")
def mark_lost(req_id: str, body: RequisitionMarkLost, user: dict = Depends(require_level(2)), db: Session = Depends(get_db)):
    req = db.query(models.Requisition).filter(models.Requisition.id == req_id).first()
    if not req:
        raise HTTPException(404, "Requisition not found")
    if req.asset_id:
        asset = db.query(models.HardwareAsset).filter(models.HardwareAsset.id == req.asset_id).first()
        if asset:
            asset.status = "Lost"
            asset.last_updated = _ts()[:10]
    req.status = "asset_lost"
    req.updated_at = _ts()
    db.add(models.ApprovalHistory(requisition_id=req_id, action="Asset Lost", action_by=body.supervisor_name, action_role="Supervisor", comment=body.notes, created_at=_ts()))
    db.commit()
    db.refresh(req)
    return req


@router.get("/requisitions/export/excel")
def export_requisitions_excel(user: dict = Depends(require_manager), db: Session = Depends(get_db)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed — run: pip install openpyxl")

    reqs = db.query(models.Requisition).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Requisitions"

    headers = [
        "Requisition ID", "Employee Name", "Employee Department", "Manager Name",
        "Dept Supervisor", "Item Requested", "Quantity", "Reason", "Request Date",
        "Manager Approval Status", "Manager Approval Date", "Manager Rejection Reason",
        "Asset Allocation Status", "Asset Name", "Asset Category", "Asset Serial / ID",
        "Asset Assigned Date", "Expected Return Date", "Actual Return Date",
        "Return Confirmed By", "Return Asset Condition", "Final Status", "Notes",
    ]

    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    status_label = {
        "pending_manager": "Pending Manager Approval",
        "rejected": "Rejected by Manager",
        "manager_approved": "Manager Approved",
        "asset_allocated": "Asset Allocated",
        "return_initiated": "Return Initiated",
        "returned": "Returned & Closed",
        "asset_lost": "Asset Lost",
    }

    for row_idx, r in enumerate(reqs, 2):
        ws.append([
            r.id, r.employee_name, r.employee_dept, r.manager_name or "",
            r.supervisor_name or "", r.item, r.quantity, r.reason,
            r.created_at[:10] if r.created_at else "",
            "Approved" if r.manager_name and r.status != "rejected" else ("Rejected" if r.status == "rejected" else "Pending"),
            r.manager_approval_date[:10] if r.manager_approval_date else "",
            r.rejection_reason or "",
            "Allocated" if r.status in ("asset_allocated", "return_initiated", "returned", "asset_lost") else "",
            r.asset_name or "", r.asset_category or "", r.asset_serial or "",
            r.asset_allocated_date[:10] if r.asset_allocated_date else "",
            r.expected_return_date or "",
            r.actual_return_date[:10] if r.actual_return_date else "",
            r.return_confirmed_by or "", r.return_asset_condition or "",
            status_label.get(r.status, r.status), "",
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"requisitions_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Hardware Assets ───────────────────────────────────────────────────────────

@router.get("/hardware-assets")
def list_hardware_assets(db: Session = Depends(get_db)):
    return db.query(models.HardwareAsset).all()


@router.post("/hardware-assets", status_code=201)
def create_hardware_asset(data: HardwareAssetCreate, user: dict = Depends(require_administrator), db: Session = Depends(get_db)):
    asset = models.HardwareAsset(**data.model_dump(), last_updated=_ts()[:10])
    db.add(asset)
    db.commit()
    db.refresh(asset)
    return asset


@router.get("/approval-history/{req_id}")
def get_approval_history(req_id: str, user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    if user["level"] < 3:
        req = db.query(models.Requisition).filter(
            models.Requisition.id == req_id,
            models.Requisition.employee_email == user["email"]
        ).first()
        if not req:
            raise HTTPException(403, "Access denied")
    return db.query(models.ApprovalHistory).filter(models.ApprovalHistory.requisition_id == req_id).all()
