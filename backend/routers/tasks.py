"""Task Module - core tasks router (rewritten Jul 2026).

Replaces the old flat demo endpoints. Covers tasks CRUD + subtasks/dependencies/
completion/approval + bulk edits, and the per-task sub-resources: comments,
attachments, activity, plus board sections and custom statuses. Identity is
email-keyed; the serialisers emit the export's runtime shape (assigneeId etc.,
with email used as the person id) so the ported frontend maps cleanly.
Reference implementation: routers/items.py.
"""
import calendar
import io
import re
import time
from datetime import date, datetime, timedelta
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional, Any
import models
from database import get_db
from auth import get_current_user, require_manager, require_any_module_grant
from routers.task_util import (
    now_iso, gen_id, fire_task_event, task_notify, log_activity, email_list,
    is_manager, visible_project_ids, task_is_visible, wall_tasks,
    project_for_task, require_project_role, require_task_role, create_comment,
    task_assignees, set_task_assignees,
)
from task_notify import notify_task_event
from task_files import data_url_to_storage
# Values are stored in the shape each field declares - see that function.
from routers.task_config import coerce_custom_field_values

router = APIRouter(prefix="/tasks", tags=["Tasks"],
                   dependencies=[Depends(get_current_user), Depends(require_any_module_grant("tasks", "tickets"))])


# ── Serialisers (snake → export runtime camelCase; email used as person id) ──
def _nz(v):
    """Empty string → None, so the frontend's `x === null` checks match the export."""
    return v if v not in ("", None) else None


def task_to_dict(t: models.Task) -> dict:
    return {
        "id":               t.id,
        "code":             t.code or "",
        "title":            t.title,
        "description":      t.description or "",
        "type":             t.type or "task",
        "status":           t.status or "not_started",
        "priority":         t.priority or "medium",
        "position":         t.position if t.position is not None else 0.0,
        # assigneeId stays the PRIMARY assignee so every existing reader keeps
        # working unchanged; assigneeIds is the real answer (see models.Task).
        "assigneeId":       _nz(t.assignee_email),
        "assigneeIds":      task_assignees(t),
        "ownerId":          _nz(t.owner_email),
        "followerIds":      t.follower_emails or [],
        "likedByIds":       t.liked_by_emails or [],
        "accessLevel":      t.access_level or "org",
        "projectId":        _nz(t.project_id),
        "projectIds":       [p for p in (t.project_ids or []) if p],
        "sectionId":        _nz(t.section_id),
        "teamId":           _nz(t.team_id),
        "parentTaskId":     _nz(t.parent_task_id),
        "subtaskIds":       t.subtask_ids or [],
        "blockedByIds":     t.blocked_by_ids or [],
        "blockingIds":      t.blocking_ids or [],
        "dependencyTypes":  t.dependency_types if isinstance(t.dependency_types, dict) else {},
        "tags":             t.tags or [],
        "customFieldValues": t.custom_field_values if isinstance(t.custom_field_values, dict) else {},
        "startOn":          _nz(t.start_on),
        "dueOn":            _nz(t.due_on),
        "estimateHours":    t.estimate_hours,
        "actualHours":      t.actual_hours,
        "recurrence":       t.recurrence,
        "isMilestone":      bool(t.is_milestone),
        "approvalStatus":   t.approval_status or "none",
        "completed":        bool(t.completed),
        "completedAt":      _nz(t.completed_at),
        "commentIds":       t.comment_ids or [],
        "attachmentIds":    t.attachment_ids or [],
        "activityIds":      t.activity_ids or [],
        "createdAt":        t.created_at or "",
        "modifiedAt":       t.modified_at or "",
        "syncedWithAsana":  bool(t.synced_with_asana),
    }


def comment_to_dict(c: models.TaskComment) -> dict:
    return {
        "id": c.id, "taskId": c.task_id, "authorId": _nz(c.author_email),
        "body": c.body or "", "createdAt": c.created_at or "",
        "editedAt": _nz(c.edited_at), "pinned": bool(c.pinned),
    }


def attachment_to_dict(a: models.TaskAttachment) -> dict:
    return {
        "id": a.id, "taskId": a.task_id, "name": a.name, "size": a.size or "",
        "kind": a.kind or "other", "dataUrl": _nz(a.url), "url": _nz(a.url),
        "addedAt": a.added_at or "", "addedBy": _nz(a.added_by),
        "commentId": _nz(a.comment_id),
    }


def activity_to_dict(a: models.TaskActivity) -> dict:
    return {
        "id": a.id, "type": a.type or "", "actorId": _nz(a.actor_email),
        "at": a.at or "", "detail": a.detail or "",
        "entityKind": a.entity_kind or "task", "entityId": _nz(a.entity_id),
        "entityCode": _nz(a.entity_code), "entityTitle": _nz(a.entity_title),
    }


def section_to_dict(s: models.TaskSection) -> dict:
    return {"id": s.id, "projectId": _nz(s.project_id), "name": s.name,
            "position": s.position or 0, "createdAt": s.created_at or ""}


def custom_status_to_dict(s: models.TaskCustomStatus) -> dict:
    return {"id": s.id, "label": s.label, "color": s.color or "", "position": s.position or 0,
            "projectIds": [p for p in (s.project_ids or []) if p]}


# ── Task CRUD ────────────────────────────────────────────────────────────────
class TaskCreate(BaseModel):
    id:               Optional[str] = None   # client-provided (optimistic) or server-generated
    code:             Optional[str] = None
    title:            str
    description:      Optional[str] = ""
    type:             Optional[str] = "task"
    status:           Optional[str] = "not_started"
    priority:         Optional[str] = "medium"
    position:         Optional[float] = None   # manual drag-order - see models.Task.position
    assignee_email:   Optional[str] = ""     # legacy single; folded into the list
    assignee_emails:  Optional[list] = None
    owner_email:      Optional[str] = ""
    follower_emails:  Optional[list] = None
    liked_by_emails:  Optional[list] = None
    # None (not "org") so create_task can tell "not specified" apart from an
    # explicit choice - an unspecified task inherits its project's visibility
    # instead of always defaulting to org-wide.
    access_level:     Optional[str] = None
    project_id:       Optional[str] = ""
    project_ids:      Optional[list] = None   # EXTRA projects beyond project_id - see models.Task
    section_id:       Optional[str] = ""
    team_id:          Optional[str] = ""
    parent_task_id:   Optional[str] = ""
    subtask_ids:      Optional[list] = None
    blocked_by_ids:   Optional[list] = None
    blocking_ids:     Optional[list] = None
    dependency_types: Optional[dict] = None
    tags:             Optional[list] = None
    custom_field_values: Optional[dict] = None
    start_on:         Optional[str] = ""
    due_on:           Optional[str] = ""
    estimate_hours:   Optional[float] = None
    actual_hours:     Optional[float] = None
    recurrence:       Optional[dict] = None
    is_milestone:     Optional[bool] = False
    approval_status:  Optional[str] = "none"


class TaskUpdate(BaseModel):
    code:             Optional[str] = None
    title:            Optional[str] = None
    description:      Optional[str] = None
    type:             Optional[str] = None
    status:           Optional[str] = None
    priority:         Optional[str] = None
    position:         Optional[float] = None   # manual drag-order - see models.Task.position
    assignee_email:   Optional[str] = None   # legacy single; folded into the list
    assignee_emails:  Optional[list] = None
    owner_email:      Optional[str] = None
    follower_emails:  Optional[list] = None
    liked_by_emails:  Optional[list] = None
    access_level:     Optional[str] = None
    project_id:       Optional[str] = None
    project_ids:      Optional[list] = None   # EXTRA projects beyond project_id - see models.Task
    section_id:       Optional[str] = None
    team_id:          Optional[str] = None
    parent_task_id:   Optional[str] = None
    subtask_ids:      Optional[list] = None
    blocked_by_ids:   Optional[list] = None
    blocking_ids:     Optional[list] = None
    dependency_types: Optional[dict] = None
    tags:             Optional[list] = None
    custom_field_values: Optional[dict] = None
    start_on:         Optional[str] = None
    due_on:           Optional[str] = None
    estimate_hours:   Optional[float] = None
    actual_hours:     Optional[float] = None
    recurrence:       Optional[dict] = None
    is_milestone:     Optional[bool] = None
    approval_status:  Optional[str] = None
    completed:        Optional[bool] = None


def _next_code(db: Session) -> str:
    n = db.query(models.Task).count() + 1
    return f"TASK-{n:03d}"


def _extra_project_ids(project_ids: Optional[list], project_id: str) -> list:
    """Clean a task's EXTRA project list: strip blanks, drop the primary
    project_id (it's not an "extra"), dedupe while preserving order."""
    seen, out = set(), []
    for p in (project_ids or []):
        p = (p or "").strip()
        if p and p != project_id and p not in seen:
            seen.add(p)
            out.append(p)
    return out


def _follow_as_actor(t, actor: str) -> None:
    """Put `actor` on the task as a collaborator unless they are already on it.

    Whoever puts work on someone else's plate almost always wants the comments
    and the completion (Sagar, Aug 27), and remembering to add yourself is
    exactly the step that gets skipped. Applied when a task is CREATED and
    whenever assignees are newly ADDED - never on an unassign, which is not a
    reason to subscribe anyone to anything.

    No-ops when the actor is an assignee (they are already on it, and a
    duplicate would show their own face twice in the Person cell) or already a
    follower. Deliberately never reached by the Asana sync, which builds rows
    directly in asana_sync.py - the sync worker must not subscribe itself to
    the entire workspace.
    """
    actor = (actor or "").strip().lower()
    if not actor or actor in task_assignees(t):
        return
    followers = list(t.follower_emails or [])
    if actor not in followers:
        t.follower_emails = followers + [actor]


def _assignees_from(data: dict, current: list) -> Optional[list]:
    """What a payload means for assignment, or None if it says nothing.

    Accepts both shapes so nothing has to migrate at once:
      • `assignee_emails` - the real field, wins when present.
      • `assignee_email`  - the legacy single. Treated as "this task is assigned
        to exactly this one person" (or to nobody, when blank), which is what it
        has always meant. It must NOT be merged into the existing list, or
        clearing an assignment the old way would silently leave the others on.
    """
    if "assignee_emails" in data and data.get("assignee_emails") is not None:
        return list(data.get("assignee_emails") or [])
    if "assignee_email" in data and data.get("assignee_email") is not None:
        one = (data.get("assignee_email") or "").strip().lower()
        return [one] if one else []
    return None


def _get_task(db: Session, task_id: str) -> models.Task:
    t = db.query(models.Task).filter(models.Task.id == task_id).first()
    if not t:
        raise HTTPException(404, "Task not found")
    return t


# ── Field validation ─────────────────────────────────────────────────────────
# Everything below rejects input the API used to store verbatim. Found by a QA
# audit (Aug 2026); each case had a real downstream effect rather than being
# merely untidy - see the individual notes.
#
# Only fields PRESENT in the payload are checked, so a task already holding a
# bad value from before this existed can still be patched on other fields
# instead of becoming uneditable.
BUILTIN_STATUSES = {"not_started", "in_progress", "completed", "recurring"}
PRIORITIES = {"urgent", "high", "medium", "low"}
# Deliberately permissive - the job is to catch "not-an-address-at-all", not to
# adjudicate RFC 5322.
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
_MAX_DEPENDENCY_WALK = 10_000   # cycle walks are bounded by `seen`; this is belt-and-braces


def _valid_statuses(db: Session) -> set:
    """Built-ins plus whatever custom statuses the workspace defines. Board and
    list views GROUP by status, so an unknown value silently creates a phantom
    column and misses every statusMeta lookup."""
    custom = {s.id for s in db.query(models.TaskCustomStatus).all()}
    return BUILTIN_STATUSES | custom


def _check_iso_date(value, field: str) -> None:
    """due_on/start_on are plain YYYY-MM-DD strings. A malformed one used to be
    stored happily, and the due-date reminder scan parses with
    date.fromisoformat inside a try/except that CONTINUES on failure - so a
    typo'd date silently disabled reminders for that task forever."""
    if value in ("", None):
        return
    try:
        date.fromisoformat(str(value)[:10])
    except ValueError:
        raise HTTPException(422, f"{field} must be a date in YYYY-MM-DD format (got {value!r}).")


def _walk_reaches(db: Session, start_ids: list, target_id: str, column) -> bool:
    """True if following `column` (a list-of-ids field) from any of start_ids
    reaches target_id. Used for both the parent chain and the dependency graph;
    `seen` makes it terminate even on data that is already cyclic."""
    seen, stack, steps = set(), [i for i in (start_ids or []) if i], 0
    while stack and steps < _MAX_DEPENDENCY_WALK:
        steps += 1
        cur = stack.pop()
        if cur == target_id:
            return True
        if cur in seen:
            continue
        seen.add(cur)
        row = db.query(column).filter(models.Task.id == cur).first()
        nxt = row[0] if row else None
        if isinstance(nxt, list):
            stack.extend(i for i in nxt if i)
        elif nxt:
            stack.append(nxt)
    return False


def validate_task_payload(db: Session, data: dict, task_id: str = "") -> dict:
    """Validate + normalize a create/update/bulk payload in place. Returns the
    same dict so callers can use it directly."""
    if "title" in data and not str(data["title"] or "").strip():
        raise HTTPException(422, "Title is required.")

    if "status" in data and data["status"] is not None:
        allowed = _valid_statuses(db)
        if data["status"] not in allowed:
            raise HTTPException(422, f"Unknown status {data['status']!r}.")

    if "priority" in data and data["priority"] is not None and data["priority"] not in PRIORITIES:
        raise HTTPException(422, f"Unknown priority {data['priority']!r}. "
                                 f"Expected one of: {', '.join(sorted(PRIORITIES))}.")

    for field in ("due_on", "start_on"):
        if field in data:
            _check_iso_date(data[field], field)

    # Negative hours silently REDUCE a person's apparent load in the Workload
    # rollup, which sums estimate/actual per assignee.
    for field in ("estimate_hours", "actual_hours"):
        if field in data and data[field] is not None:
            try:
                if float(data[field]) < 0:
                    raise HTTPException(422, f"{field} cannot be negative.")
            except (TypeError, ValueError):
                raise HTTPException(422, f"{field} must be a number.")

    # Normalize rather than reject: blank/duplicate entries are a UI slip, not
    # something worth failing a save over. Blank tags rendered as empty pills and
    # duplicates appeared twice in the filter list; duplicate collaborators drew
    # the same avatar repeatedly.
    if "tags" in data and isinstance(data["tags"], list):
        seen, out = set(), []
        for tag in data["tags"]:
            v = str(tag or "").strip()
            if v and v.lower() not in seen:
                seen.add(v.lower())
                out.append(v)
        data["tags"] = out

    # assignee_emails rides the same normaliser as the collaborator lists: an
    # assignee is a person too, and a non-address there is worse than a bad
    # follower - it is work nobody can be notified about or find in My Tasks.
    for field in ("assignee_emails", "follower_emails", "liked_by_emails"):
        if field in data and isinstance(data[field], list):
            seen, out = set(), []
            for em in data[field]:
                v = str(em or "").strip().lower()
                if not v or v in seen:
                    continue
                # A collaborator is a person, and every downstream use assumes an
                # address: the notification fan-out mails it and the avatar
                # derives initials from it. A non-address just silently never
                # gets notified.
                if not _EMAIL_RE.match(v):
                    raise HTTPException(422, f"{v!r} is not a valid email address.")
                seen.add(v)
                out.append(v)
            data[field] = out

    # ── structural cycles ────────────────────────────────────────────────
    # A task that is its own parent disappears from the UI entirely: topLevel()
    # drops anything with a parentTaskId, and its "parent" is itself, so it
    # never nests under anything either.
    if task_id and data.get("parent_task_id"):
        parent = data["parent_task_id"]
        if parent == task_id:
            raise HTTPException(422, "A task cannot be its own parent.")
        if _walk_reaches(db, [parent], task_id, models.Task.parent_task_id):
            raise HTTPException(422, "That parent would create a circular subtask chain.")

    # A dependency cycle is worse than untidy: _check_dependency_gate refuses
    # completion while a blocker is open, so two tasks blocking each other can
    # NEVER be completed by anyone.
    if task_id and data.get("blocked_by_ids"):
        ids = [i for i in data["blocked_by_ids"] if i]
        if task_id in ids:
            raise HTTPException(422, "A task cannot block itself.")
        _require_tasks_exist(db, ids)
        if _walk_reaches(db, ids, task_id, models.Task.blocked_by_ids):
            raise HTTPException(422, "That dependency would create a circular chain, "
                                     "leaving both tasks permanently uncompletable.")

    if task_id and data.get("blocking_ids"):
        ids = [i for i in data["blocking_ids"] if i]
        if task_id in ids:
            raise HTTPException(422, "A task cannot block itself.")
        _require_tasks_exist(db, ids)
        if _walk_reaches(db, ids, task_id, models.Task.blocking_ids):
            raise HTTPException(422, "That dependency would create a circular chain.")

    return data


def _require_tasks_exist(db: Session, ids: list) -> None:
    """A dependency on an id that doesn't resolve is invisible: the drawer's
    blockedBy lookup silently drops it and the completion gate ignores it, so
    the task LOOKS blocked in the payload while behaving as if it isn't."""
    found = {r[0] for r in db.query(models.Task.id).filter(models.Task.id.in_(ids)).all()}
    missing = [i for i in ids if i not in found]
    if missing:
        raise HTTPException(422, f"Unknown task id(s): {', '.join(missing[:5])}.")


def _sync_reciprocal_dependencies(db: Session, t: models.Task, data: dict) -> None:
    """Keep the other end of a dependency pointing back.

    blocked_by_ids and blocking_ids are two views of one edge, and only the
    first is authoritative - _check_dependency_gate reads blocked_by_ids, so
    that side alone decides whether a task may start or finish. blocking_ids is
    what the drawer's "Blocking" panel lists.

    The reciprocal used to be written by the browser, as a SECOND request right
    after this one (TaskDetailDrawer's addDep/removeDep). Any other route to the
    same edge - the bulk endpoint, a script, or simply that second call failing
    - left the graph one-sided and the panel under-reporting. Doing it here puts
    both ends in the same transaction, and the client's now-redundant second
    call is idempotent against it.

    Only edges this payload actually changed are touched, so a task that merely
    had its title edited never rewrites its neighbours."""
    for field, mirror in (("blocked_by_ids", "blocking_ids"),
                          ("blocking_ids", "blocked_by_ids")):
        if field not in data:
            continue
        now = {i for i in (getattr(t, field) or []) if i}
        # The pre-patch value is gone from `t` by the time this runs, so the
        # other end is the source of truth for what the edge used to be: every
        # task whose mirror list still names this one.
        was = {row.id for row in db.query(models.Task).all()
               if t.id in (getattr(row, mirror) or [])}
        for other in db.query(models.Task).filter(
                models.Task.id.in_((now | was) - {t.id})).all():
            cur = [i for i in (getattr(other, mirror) or []) if i]
            want = sorted(set(cur) | {t.id}) if other.id in now else [i for i in cur if i != t.id]
            if want != cur:
                setattr(other, mirror, want)
                other.modified_at = now_iso()


def _check_dependency_gate(db: Session, t: models.Task, prev_status: str, prev_completed: bool,
                            new_status: str, new_completed: bool) -> None:
    """Enforce blockedBy relationship types before a status/completion change lands.
    FS/SS gate the task *starting* (leaving not_started); FF/SF gate it *finishing*
    (completed). 'Started' on the blocker means it has left not_started OR is
    completed. Raises 400 with a message naming the still-blocking task."""
    if not t.blocked_by_ids:
        return
    starting_now = prev_status == "not_started" and new_status != "not_started"
    completing_now = (not prev_completed) and new_completed
    if not (starting_now or completing_now):
        return
    dep_types = t.dependency_types or {}
    blockers = {b.id: b for b in db.query(models.Task).filter(models.Task.id.in_(t.blocked_by_ids)).all()}
    for blocker_id in t.blocked_by_ids:
        blocker = blockers.get(blocker_id)
        if not blocker:
            continue
        dep_type = dep_types.get(blocker_id, "FS")
        blocker_started = blocker.status != "not_started" or blocker.completed
        blocker_completed = bool(blocker.completed)
        # Title, never the code: task numbers are not shown anywhere in the
        # module, so "Blocked by TASK-1983" would name something the reader
        # has no way to find.
        name = blocker.title or "another task"
        if dep_type == "FS" and (starting_now or completing_now) and not blocker_completed:
            raise HTTPException(400, f"Blocked by {name}: finish it before starting or completing this task (Finish → Start).")
        if dep_type == "SS" and starting_now and not blocker_started:
            raise HTTPException(400, f"Blocked by {name}: start it before starting this task (Start → Start).")
        if dep_type == "FF" and completing_now and not blocker_completed:
            raise HTTPException(400, f"Blocked by {name}: finish it before completing this task (Finish → Finish).")
        if dep_type == "SF" and completing_now and not blocker_started:
            raise HTTPException(400, f"Blocked by {name}: start it before completing this task (Start → Finish).")


def _asana_push(task_id: str, actor_email: str = "") -> None:
    """Fire-and-forget outbound Asana sync. Fully guarded - must never affect the
    task operation that triggered it (runs in a daemon thread on its own session).

    `actor_email` is the signed-in user who made the change. Asana attributes the
    system stories a write produces ("X changed Priority to Medium") to whoever
    owns the token, so without it every field change reads as the shared sync
    account. This is the boundary that used to drop it."""
    try:
        from asana_sync import on_task_changed
        on_task_changed(task_id, actor_email)
    except Exception:
        pass


def _asana_linked_gids(db: Session, task_ids: list) -> list:
    """Asana gids for tasks about to be deleted, read BEFORE the rows go
    (deleting them takes their AsanaTaskLink rows too). Fully guarded."""
    try:
        from asana_sync import linked_gids
        return linked_gids(db, task_ids)
    except Exception:
        return []


def _asana_queue_delete(db: Session, gids: list, title: str, code: str, actor: str) -> None:
    """Record deletions owed to Asana, in the caller's transaction so the
    tombstone commits with the delete. Fully guarded."""
    try:
        from asana_sync import queue_task_delete
        queue_task_delete(db, gids, title, code, actor)
    except Exception:
        pass


def _asana_push_deleted() -> None:
    """Fire-and-forget drain of the queued deletions. No-op outside the sync
    worker, where "Push all" drains them instead. Fully guarded."""
    try:
        from asana_sync import on_task_deleted
        on_task_deleted()
    except Exception:
        pass


def _asana_push_comment_edit(comment_id: str) -> None:
    """Fire-and-forget outbound push of an edited comment body. Fully guarded -
    the edit is already committed and must never fail on the sync."""
    try:
        from asana_sync import on_comment_edited
        on_comment_edited(comment_id)
    except Exception:
        pass


# ── Completion: `status` and `completed` are one fact in two columns ─────────
# They must never disagree. A task whose status is "completed" while its
# `completed` flag is False renders with an empty circle and no strikethrough
# while sitting in the Completed group - and, worse, task_notify's due scan
# filters on `completed`, so it keeps sending "this task is overdue" about work
# that is finished.
#
# update_task got this right and was the only one that did. create_task hardcoded
# completed=False next to whatever status it was handed, and bulk_update wrote
# `status` through a plain setattr. Both were reachable from the UI: the
# "+ Add task…" row inside a group inherits that group's status, so adding a task
# under the Completed heading created exactly this state, as did the multi-select
# Status dropdown. One rule, three callers, no room to drift.

def _resolve_completed(data: dict, prev_completed: bool) -> bool:
    """What `completed` should become, given a payload and the current value.

    A status set to anything OTHER than completed, with no explicit `completed`
    in the same payload, means the task is being reopened - otherwise the flag
    stuck True and list/board views kept the task parked in Completed."""
    if "status" in data and data["status"] != "completed" and "completed" not in data:
        return False
    return bool(data.get("completed", prev_completed)) or (data.get("status") == "completed")


def _apply_completion(t: models.Task, new_completed: bool, prev_completed: bool) -> None:
    """Land `completed` on the task and drag `status` and `completed_at` with it."""
    if new_completed == prev_completed:
        return
    t.completed = new_completed
    if new_completed:
        t.completed_at = now_iso()
        if t.status != "completed":
            t.status = "completed"
    else:
        t.completed_at = ""
        if t.status == "completed":
            t.status = "not_started"


# ── Recurrence: occurrence generation ────────────────────────────────────────
# recurrence = {freq, interval, dayOfWeek?, dayOfMonth?, until?, count?}
# dayOfWeek is the frontend index (0=Sunday..6=Saturday); until/count are the
# end condition. `count` on an instance means "occurrences remaining, this one
# inclusive" - it is decremented each time the next occurrence is spawned, so a
# series with count=N produces exactly N tasks.
def _add_months(d: date, n: int) -> date:
    m = d.month - 1 + n
    y = d.year + m // 12
    m = m % 12 + 1
    return date(y, m, min(d.day, calendar.monthrange(y, m)[1]))


def _next_due(base_iso: str, rec: dict) -> str:
    """Next due date (YYYY-MM-DD) after `base_iso`, per the recurrence rule."""
    try:
        base = datetime.strptime((base_iso or "")[:10], "%Y-%m-%d").date()
    except ValueError:
        base = date.today()
    freq = rec.get("freq")
    interval = max(1, int(rec.get("interval") or 1))
    if freq == "weekly":
        dow = rec.get("dayOfWeek")
        if dow is None:
            return (base + timedelta(weeks=interval)).isoformat()
        target = (int(dow) - 1) % 7          # frontend Sun=0 → Python Mon=0
        nxt = base + timedelta(days=1)
        while nxt.weekday() != target:
            nxt += timedelta(days=1)
        return (nxt + timedelta(weeks=interval - 1)).isoformat()
    if freq == "monthly":
        nxt = _add_months(base, interval)
        dom = rec.get("dayOfMonth")
        if dom:
            day = min(int(dom), calendar.monthrange(nxt.year, nxt.month)[1])
            nxt = date(nxt.year, nxt.month, day)
        return nxt.isoformat()
    if freq == "yearly":
        return _add_months(base, 12 * interval).isoformat()
    # daily (and any unknown freq) → advance whole days
    return (base + timedelta(days=interval)).isoformat()


def _spawn_next_occurrence(db: Session, t: models.Task, user: dict) -> Optional[models.Task]:
    """When a recurring task is completed, create its next occurrence (unless the
    end condition - `until` date or `count` - is reached). Returns the new task
    or None if the series has ended. Top-level tasks only; subtasks don't recur."""
    rec = t.recurrence
    if not isinstance(rec, dict) or not rec.get("freq") or t.parent_task_id:
        return None

    count = rec.get("count")
    if count is not None and int(count) <= 1:
        return None  # this was the final occurrence

    base_iso = t.due_on or (t.completed_at or now_iso())[:10]
    next_due = _next_due(base_iso, rec)
    until = rec.get("until")
    if until and next_due > until:
        return None  # past the series end date

    new_rec = dict(rec)
    if count is not None:
        new_rec["count"] = int(count) - 1

    now = now_iso()
    nid = gen_id()
    nxt = models.Task(
        id=nid,
        company_id=getattr(t, "company_id", "") or "",   # inherit the parent's company
        code=_next_code(db),
        title=t.title,
        description=t.description or "",
        type=t.type or "task",
        status="recurring",
        priority=t.priority or "medium",
        assignee_email=t.assignee_email or "",
        assignee_emails=list(task_assignees(t)),
        owner_email=t.owner_email or "",
        follower_emails=list(t.follower_emails or []),
        liked_by_emails=[],
        access_level=t.access_level or "org",
        project_id=t.project_id or "",
        section_id=t.section_id or "",
        team_id=t.team_id or "",
        parent_task_id="",
        subtask_ids=[], blocked_by_ids=[], blocking_ids=[], dependency_types={},
        tags=list(t.tags or []),
        custom_field_values=dict(t.custom_field_values or {}),
        start_on="",
        due_on=next_due,
        estimate_hours=t.estimate_hours,
        actual_hours=None,
        recurrence=new_rec,
        is_milestone=bool(t.is_milestone),
        approval_status="none",
        completed=False,
        completed_at="",
        comment_ids=[], attachment_ids=[], activity_ids=[],
        created_at=now, modified_at=now, created_by=user["email"],
    )
    db.add(nxt)
    aid = log_activity(db, type="created", actor_email=user["email"], entity_id=nid,
                       entity_code=nxt.code, entity_title=nxt.title,
                       detail=f"recurring occurrence generated from {t.title}")
    nxt.activity_ids = [aid]
    for _who in task_assignees(nxt):
        if _who == user["email"].lower():
            continue
        task_notify(db, kind="task_assigned", for_email=_who,
                    title="Recurring task due",
                    body=f"{nxt.title} (due {next_due})", task_id=nid,
                    nexus_action={"view": "tasks", "sub": "mine", "label": "View task"})
    return nxt


@router.get("")
def list_tasks(user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    rows = wall_tasks(db, user, db.query(models.Task).all())   # company wall first
    if is_manager(user):
        return [task_to_dict(t) for t in rows]
    # Pass the USER dict, not just the email: the helpers scope external
    # (B2B guest) callers to explicit participation only (Aug 17).
    visible_projects = visible_project_ids(db, user)
    return [task_to_dict(t) for t in rows if task_is_visible(t, user, visible_projects)]


def _csv_set(value: str) -> set:
    """`?status=a,b,c` -> {"a","b","c"}. Blank/absent -> empty set, which every
    filter below reads as "no constraint" rather than "match nothing"."""
    return {v.strip() for v in (value or "").split(",") if v.strip()}


@router.get("/export/excel")
def export_tasks_excel(
    project_id: str = "",
    status: str = "",
    priority: str = "",
    assignee: str = "",
    due_from: str = "",
    due_to: str = "",
    include_completed: bool = True,
    today: str = "",
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Tasks as a real .xlsx - one row per task, collaborators collapsed into a
    single cell so the sheet stays one-task-per-line and pivots cleanly.

    project_id/status/priority/assignee each take a COMMA-SEPARATED list and
    match on ANY of the values (an OR within a filter, AND across filters) -
    the same "pick none to include everything" shape the custom-field project
    picker uses. Every person column is paired with an "Enterprise Id" column
    carrying the work email, so the sheet stays joinable against other systems
    even though people are displayed by name (CLAUDE.md).

    The FILENAME describes what is in the sheet rather than being a fixed
    string - "2026-08-21 - Test Project Task Export.xlsx", "2026-08-21 - Sagar
    Shoundik's Tasks.xlsx" - so a folder of these stays readable months later.
    `today` is the CALLER's local date: the server runs UTC, which names an
    evening export in India after the day it was actually taken.

    Visibility is the SAME rule list_tasks applies (task_is_visible against
    visible_project_ids), so an export can never hand someone a task they
    cannot already open - filters narrow that set, they never widen it.
    Reference implementation for the workbook itself: requisitions.py.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed - run: pip install openpyxl")

    rows = wall_tasks(db, user, db.query(models.Task).all())   # company wall first
    if not is_manager(user):
        visible = visible_project_ids(db, user)
        rows = [t for t in rows if task_is_visible(t, user, visible)]

    # -- Filters (empty = no constraint; several values = match ANY) ----------
    want_projects = _csv_set(project_id)
    want_statuses = _csv_set(status)
    want_priorities = _csv_set(priority)
    want_assignees = {a.lower() for a in _csv_set(assignee)}

    if want_projects:
        rows = [t for t in rows
                if t.project_id in want_projects
                or want_projects & set(t.project_ids or [])]
    if want_statuses:
        rows = [t for t in rows if (t.status or "not_started") in want_statuses]
    if want_priorities:
        rows = [t for t in rows if (t.priority or "medium") in want_priorities]
    if want_assignees:
        rows = [t for t in rows if want_assignees & set(task_assignees(t))]
    if due_from:
        rows = [t for t in rows if t.due_on and t.due_on >= due_from]
    if due_to:
        rows = [t for t in rows if t.due_on and t.due_on <= due_to]
    if not include_completed:
        rows = [t for t in rows if not t.completed]

    # -- Label lookups -------------------------------------------------------
    # Same precedence the frontend's useNameResolver applies, so a sheet and the
    # screen it was exported from never disagree about someone's name: curated
    # People directory first, the older roles directory as the fallback for
    # people no longer in People, then the email itself. People must always show
    # as first + last name, never a raw email (CLAUDE.md) - the paired
    # Enterprise Id column is where the email belongs.
    people = {}
    for r in db.query(models.NexusRole).all():
        if r.email and (r.display_name or "").strip():
            people[r.email.lower()] = r.display_name.strip()
    for e in db.query(models.NexusEmployee).all():
        full = f"{e.first_name or ''} {e.last_name or ''}".strip()
        if e.work_email and full:
            people[e.work_email.lower()] = full

    def person(email):
        if not email:
            return ""
        hit = people.get(email.lower())
        if hit:
            return hit
        local = email.split("@")[0]
        parts = [p for p in re.split(r"[._-]+", local) if p]
        return " ".join(p[:1].upper() + p[1:] for p in parts) or email

    projects = {p.id: p.name for p in db.query(models.TaskProject).all()}
    status_label = {"not_started": "Not Started", "in_progress": "In Progress",
                    "completed": "Completed", "recurring": "Recurring"}
    for s in db.query(models.TaskCustomStatus).all():
        status_label[s.id] = s.label
    priority_label = {"low": "Low", "medium": "Medium", "high": "High", "urgent": "Urgent"}
    titles = {t.id: (t.title or "") for t in db.query(models.Task).all()}

    def us_date(iso):
        """MM/DD/YYYY - US format everywhere in user-facing output (CLAUDE.md)."""
        if not iso:
            return ""
        try:
            return datetime.strptime(iso[:10], "%Y-%m-%d").strftime("%m/%d/%Y")
        except ValueError:
            return iso

    # Grouped by project, then due date - blank dues last rather than first.
    rows.sort(key=lambda t: (projects.get(t.project_id or "", "zzz").lower(),
                             t.due_on or "9999-99-99", (t.title or "").lower()))

    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"

    headers = ["Task ID", "Task", "Parent Task", "Project",
               "Assignee", "Assignee Enterprise Id",
               "Collaborators", "Collaborators Enterprise Id",
               "Start Date", "Due Date", "Priority", "Status", "Completed",
               "Created On", "Modified On"]
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for t in rows:
        extra = [projects.get(p, "") for p in (t.project_ids or []) if p and p != t.project_id]
        project_cell = ", ".join(x for x in [projects.get(t.project_id or "", "")] + extra if x)
        collaborators = [e for e in (t.follower_emails or []) if e]
        ws.append([
            t.code or t.id,
            t.title or "",
            titles.get(t.parent_task_id or "", ""),
            project_cell,
            ", ".join(person(a) or a for a in task_assignees(t)) or "Unassigned",
            ", ".join(task_assignees(t)),
            ", ".join(person(e) for e in collaborators),
            ", ".join(collaborators),
            us_date(t.start_on),
            us_date(t.due_on),
            priority_label.get(t.priority or "medium", t.priority or ""),
            status_label.get(t.status or "not_started", t.status or ""),
            "Yes" if t.completed else "No",
            us_date(t.created_at),
            us_date(t.modified_at),
        ])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(headers)).column_letter}{max(ws.max_row, 1)}"
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # -- Filename: say what is in the sheet ----------------------------------
    # "2026-08-21 - Task Export"                      (no filters)
    # "2026-08-21 - Test Project Task Export"         (one project)
    # "2026-08-21 - Sagar Shoundik's Tasks"           (one person)
    # "2026-08-21 - Test Project - Sagar Shoundik's Tasks"  (both)
    # Several of either collapse to a count ("3 Projects", "4 People's Tasks")
    # rather than running the name past what a file manager will show.
    stamp = today if re.fullmatch(r"\d{4}-\d{2}-\d{2}", today or "") else datetime.utcnow().strftime("%Y-%m-%d")

    picked_projects = [projects[p] for p in sorted(want_projects) if projects.get(p)]
    picked_people = [person(a) for a in sorted(want_assignees)]

    if len(picked_people) == 1:
        who = picked_people[0]
        # "Chris' Tasks", not "Chris's Tasks".
        whose = f"{who}'" if who.endswith("s") else f"{who}'s"
        label = f"{whose} Tasks"
    elif picked_people:
        label = f"{len(picked_people)} People's Tasks"
    else:
        label = "Task Export"

    if len(picked_projects) == 1:
        scope = picked_projects[0]
    elif picked_projects:
        scope = f"{len(picked_projects)} Projects"
    else:
        scope = ""

    if scope and picked_people:
        name = f"{stamp} - {scope} - {label}"
    elif scope:
        name = f"{stamp} - {scope} {label}"
    else:
        name = f"{stamp} - {label}"

    # A project name is free text, so strip what Windows/macOS reject in a
    # filename before it reaches a download folder.
    name = re.sub(r'[\\/:*?"<>|\r\n\t]', " ", name)
    name = re.sub(r"\s+", " ", name).strip()[:120]
    filename = f"{name}.xlsx"

    # Quoted (it has spaces) plus RFC 5987 for names that leave ASCII - an
    # accented person or project name would otherwise arrive mangled.
    ascii_name = filename.encode("ascii", "ignore").decode() or "Task Export.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="{ascii_name}"; '
                 f"filename*=UTF-8''{quote(filename)}"},
    )


@router.get("/search")
def search_everything(q: str = "", limit: int = 6,
                      user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """One query, results grouped by what they are - the shape Asana's own
    search offers, and the shape people expect from a magnifier in a header.

    The header's search used to match MODULE NAMES only, so typing the title of
    a task you were looking at returned nothing at all.

    Visibility is the same rule the task list applies (task_is_visible against
    visible_project_ids), so search can never surface a task someone cannot
    already open. People come from the curated Nexus directory, never a
    M365/GAL-derived list (CLAUDE.md).

    Ranking is deliberately dumb and predictable: a prefix match sorts above a
    match anywhere, then alphabetical. Anything cleverer here is a scoring
    system nobody can explain when it puts the wrong row first."""
    term = (q or "").strip().lower()
    if len(term) < 2:
        return {"tasks": [], "projects": [], "portfolios": [], "teams": [], "people": [], "totals": {}}
    limit = max(1, min(limit, 20))

    def rank(text: str) -> tuple:
        t = (text or "").lower()
        return (0 if t.startswith(term) else 1, t)

    def top(rows, text_of):
        return sorted(rows, key=lambda r: rank(text_of(r)))[:limit]

    manager = is_manager(user)
    visible = None if manager else visible_project_ids(db, user)

    tasks = [t for t in wall_tasks(db, user, db.query(models.Task).all())
             if (term in (t.title or "").lower() or term in (t.code or "").lower())
             and not t.completed
             and (manager or task_is_visible(t, user, visible))]
    project_names = {p.id: p.name for p in db.query(models.TaskProject).all()}
    # Archived projects DO appear here. Archiving hides a project from the
    # active list and from task pickers - it is not deletion, and search is how
    # you get back to finished work. The result carries `archived` so the UI can
    # mark it rather than pass it off as live.
    projects = [p for p in db.query(models.TaskProject).all()
                if term in (p.name or "").lower()
                and (manager or p.id in visible)]
    portfolios = [p for p in db.query(models.TaskPortfolio).all() if term in (p.name or "").lower()]
    teams = [t for t in db.query(models.TaskTeam).all() if term in (t.name or "").lower()]

    people = []
    for e in db.query(models.NexusEmployee).all():
        if (e.status or "") in ("inactive", "offboarded"):
            continue
        name = f"{e.first_name or ''} {e.last_name or ''}".strip()
        if term in name.lower() or term in (e.work_email or "").lower():
            # jobTitle, not "title" - a task's title is also `title`, and one
            # dict key meaning two different things across the same response is
            # how a renderer ends up printing someone's job where their name goes.
            people.append({"email": e.work_email, "name": name or e.work_email,
                           "jobTitle": e.job_title or ""})

    # A subtask carries project_id "" - its project is its parent's. Without
    # the walk, 92 same-titled "Finish all Books for Current Year" subtasks
    # (one per company's Tax Return) rendered as six identical bare lines and
    # read as duplicates of nothing. parentTitle is what tells them apart.
    by_id = {t.id: t for t in db.query(models.Task).all()} if any(t.parent_task_id for t in tasks) else {}

    def shape_task(t):
        parent = by_id.get(t.parent_task_id or "") if t.parent_task_id else None
        proj_id = t.project_id or ""
        cur, depth = t, 0
        while not proj_id and cur is not None and cur.parent_task_id and depth < 20:
            cur = by_id.get(cur.parent_task_id)
            proj_id = (cur.project_id or "") if cur else ""
            depth += 1
        return {"id": t.id, "code": t.code or "", "title": t.title,
                "completed": bool(t.completed), "assigneeId": _nz(t.assignee_email),
                "projectId": _nz(t.project_id),
                "projectName": project_names.get(proj_id, ""),
                "parentTitle": (parent.title if parent else "")}

    return {
        "tasks": [shape_task(t) for t in top(tasks, lambda t: t.title)],
        # Archived projects stay FINDABLE - archiving hides a project from the
        # active list and from task pickers, it does not hide its history. The
        # flag rides along so the result can be marked as archived rather than
        # looking like live work.
        "projects": [{"id": p.id, "name": p.name, "color": p.color or "", "archived": bool(p.archived)}
                     for p in top(projects, lambda p: p.name)],
        "portfolios": [{"id": p.id, "name": p.name} for p in top(portfolios, lambda p: p.name)],
        "teams": [{"id": t.id, "name": t.name, "color": t.color or ""}
                  for t in top(teams, lambda t: t.name)],
        "people": top(people, lambda p: p["name"]),
        # How many matched in total, so a capped list can say "and N more"
        # instead of passing itself off as everything there is.
        "totals": {"tasks": len(tasks), "projects": len(projects), "portfolios": len(portfolios),
                   "teams": len(teams), "people": len(people)},
    }


@router.get("/people/{email}")
def person_profile(email: str, user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """Everything one person's page shows, in one call: who they are, the work
    they hold, and where they belong.

    Four task buckets, matching how people actually ask about a colleague:
    everything they hold (`assigned`), what YOU gave them (`assignedByYou`),
    what they handed out (`created` - they made it, somebody else holds it),
    and where the two of you overlap (`collaboratingWithYou`). Two of those are
    relative to the VIEWER, which is the point: "what did I give this person"
    is a different question from "what are they doing", and a page that only
    answered the second made you go and work the first out by eye.

    All four go through the same visibility rule as the task list, so this page
    can never become a way to read work you could not otherwise open.

    The person comes from the curated Nexus directory (nexus_employees), never
    a M365/GAL list - CLAUDE.md. An address with no directory row still gets a
    page: Asana guests sync in as collaborators long before HR has a record,
    and their work is real even when their profile is thin."""
    email = (email or "").strip().lower()
    if not email:
        raise HTTPException(404, "Person not found")
    emp = db.query(models.NexusEmployee).filter(
        models.NexusEmployee.work_email == email).first()

    manager = is_manager(user)
    visible = None if manager else visible_project_ids(db, user)
    rows = [t for t in wall_tasks(db, user, db.query(models.Task).all())
            if manager or task_is_visible(t, user, visible)]
    project_names = {p.id: p.name for p in db.query(models.TaskProject).all()}

    def shape(t):
        return {"id": t.id, "code": t.code or "", "title": t.title,
                "completed": bool(t.completed), "status": t.status,
                "dueOn": _nz(t.due_on), "assigneeId": _nz(t.assignee_email),
                "projectId": _nz(t.project_id),
                "projectName": project_names.get(t.project_id or "", "")}

    def newest(items, limit=25):
        # Open work first, then nearest due date - a page that opened on
        # something finished last March would bury what the person is on now.
        return [shape(t) for t in sorted(
            items, key=lambda t: (bool(t.completed), t.due_on or "9999", t.title.lower()))[:limit]]

    me = (user.get("email") or "").strip().lower()

    def on_it(t, who):
        """Everyone a task visibly involves - its assignee and its followers."""
        return who and (who in task_assignees(t)
                        or who in email_list(t.follower_emails))

    assigned = [t for t in rows if email in task_assignees(t)]
    assigned_by_you = [t for t in assigned if (t.created_by or "").lower() == me]
    created = [t for t in rows if (t.created_by or "").lower() == email
               and email not in task_assignees(t)]
    # Both of you on the same task. Empty on your OWN page, where "collaborating
    # with you" would otherwise match every task you touch and mean nothing.
    collaborating = ([] if me == email else
                     [t for t in rows if on_it(t, email) and on_it(t, me)])

    today = date.today().isoformat()
    open_assigned = [t for t in assigned if not t.completed]
    projects = [p for p in db.query(models.TaskProject).all()
                if (manager or p.id in visible) and not p.archived
                and (email in email_list(p.member_emails)
                     or (p.owner_email or "").lower() == email
                     or p.id in {t.project_id for t in assigned if t.project_id})]
    teams = [t for t in db.query(models.TaskTeam).all()
             if email in email_list(t.member_emails)]

    return {
        "person": {
            "email": email,
            "name": (f"{emp.first_name or ''} {emp.last_name or ''}".strip() or email) if emp else email,
            "displayName": (emp.display_name or "") if emp else "",
            "jobTitle": (emp.job_title or "") if emp else "",
            "department": (emp.department or "") if emp else "",
            "location": (emp.location or "") if emp else "",
            "photoUrl": (emp.photo_url or "") if emp else "",
            "identityType": (emp.identity_type or "internal") if emp else "external",
            "status": (emp.status or "active") if emp else "",
            "inDirectory": bool(emp),
        },
        "stats": {
            "open": len(open_assigned),
            "completed": len([t for t in assigned if t.completed]),
            "overdue": len([t for t in open_assigned if t.due_on and t.due_on[:10] < today]),
        },
        "tasks": {"assigned": newest(assigned), "assignedByYou": newest(assigned_by_you),
                  "created": newest(created), "collaboratingWithYou": newest(collaborating)},
        "projects": [{"id": p.id, "name": p.name, "color": p.color or ""} for p in projects],
        "teams": [{"id": t.id, "name": t.name, "color": t.color or ""} for t in teams],
    }


@router.get("/delta")
def list_tasks_delta(since: str = "", user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """Incremental fetch for TasksContext's mount load + repeated refresh
    (45s poll, realtime ping) - GET /tasks ships the full ~2,400-task
    workspace every call even when nothing changed. `since=""` (the mount
    case) naturally returns everything with no deletions, so this serves
    both the initial and incremental load through one path.

    `server_time` is captured BEFORE the query runs, not after - same
    reasoning as asana_sync._pull_window's "stamped from when the fetch
    STARTED": a task edited mid-query must fall in the NEXT delta window,
    never be missed because it looked "already covered" by this one."""
    server_time = now_iso()
    q = db.query(models.Task)
    if since:
        q = q.filter(models.Task.modified_at > since)
    rows = wall_tasks(db, user, q.all())   # company wall first
    if not is_manager(user):
        visible_projects = visible_project_ids(db, user)
        rows = [t for t in rows if task_is_visible(t, user, visible_projects)]
    deleted = (db.query(models.TaskDeleteLog).filter(models.TaskDeleteLog.deleted_at > since).all()
              if since else [])
    return {"tasks": [task_to_dict(t) for t in rows],
            "deletedIds": [d.task_id for d in deleted],
            "serverTime": server_time}


@router.post("", status_code=201)
def create_task(body: TaskCreate, background_tasks: BackgroundTasks,
                user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    now = now_iso()
    tid = body.id or gen_id()
    # No explicit access_level -> inherit the project's (so a task in a
    # restricted project doesn't leak org-wide by default); no project either
    # -> falls back to org, same as before.
    project = (db.query(models.TaskProject).filter(models.TaskProject.id == body.project_id).first()
               if body.project_id else None)
    if not project and body.parent_task_id:
        parent = db.query(models.Task).filter(models.Task.id == body.parent_task_id).first()
        if parent:
            project = project_for_task(db, parent)
    require_project_role(db, user, project, "editor")
    # Normalizes tags/followers in place too, so the row is written clean rather
    # than needing a later PATCH to tidy it.
    validate_task_payload(db, body.__dict__, task_id=tid)
    access_level = body.access_level or (project.access_level if project else None) or "org"
    # Company wall: stamp the creator's company so the task is fixed on their side
    # of the wall even if they later change companies (task_util.task_company falls
    # back to deriving it for legacy/sync rows that were never stamped).
    _creator = (db.query(models.NexusEmployee.company)
                .filter(func.lower(models.NexusEmployee.work_email) == user["email"].lower()).first())
    company_id = (_creator.company if _creator else "") or ""
    t = models.Task(
        id=tid,
        company_id=company_id,
        code=body.code or _next_code(db),
        title=body.title,
        description=body.description or "",
        type=body.type or "task",
        status=body.status or "not_started",
        priority=body.priority or "medium",
        # Defaults to "now" (epoch seconds) rather than 0 so a freshly created
        # task lands at the END of manual order, after everything dragged
        # before it - not shuffled to the front of every group.
        position=body.position if body.position is not None else time.time(),
        # Set below via set_task_assignees, which writes both columns.
        assignee_email="",
        owner_email=(body.owner_email or "").strip().lower(),
        follower_emails=body.follower_emails or [],
        liked_by_emails=body.liked_by_emails or [],
        access_level=access_level,
        project_id=body.project_id or "",
        project_ids=_extra_project_ids(body.project_ids, body.project_id or ""),
        section_id=body.section_id or "",
        team_id=body.team_id or "",
        parent_task_id=body.parent_task_id or "",
        subtask_ids=body.subtask_ids or [],
        blocked_by_ids=body.blocked_by_ids or [],
        blocking_ids=body.blocking_ids or [],
        dependency_types=body.dependency_types or {},
        tags=body.tags or [],
        custom_field_values=coerce_custom_field_values(db, body.custom_field_values),
        start_on=body.start_on or "",
        due_on=body.due_on or "",
        estimate_hours=body.estimate_hours,
        actual_hours=body.actual_hours,
        recurrence=body.recurrence,
        is_milestone=bool(body.is_milestone),
        approval_status=body.approval_status or "none",
        # Not hardcoded False any more: a task created straight INTO the
        # Completed group (the "+ Add task…" row inherits its group's status)
        # arrived with status "completed" and this flag off, so it showed no
        # strikethrough and kept drawing overdue emails. TaskCreate has no
        # `completed` field, so status is the only thing that can say so here.
        completed=(body.status == "completed"),
        completed_at=(now if body.status == "completed" else ""),
        comment_ids=[], attachment_ids=[], activity_ids=[],
        created_at=now, modified_at=now, created_by=user["email"],
    )
    set_task_assignees(t, _assignees_from(body.__dict__, []) or [])
    # Whoever raised the task follows it - see _follow_as_actor.
    _follow_as_actor(t, user["email"])
    db.add(t)
    # link into parent
    if t.parent_task_id:
        parent = db.query(models.Task).filter(models.Task.id == t.parent_task_id).first()
        if parent:
            parent.subtask_ids = list(parent.subtask_ids or []) + [tid]
            parent.modified_at = now
    aid = log_activity(db, type="created", actor_email=user["email"], entity_id=tid,
                       entity_code=t.code, entity_title=t.title, detail="created this task")
    t.activity_ids = [aid]
    # Everyone assigned gets told, not just the first - the others would
    # otherwise carry work they were never notified about.
    for who in task_assignees(t):
        if who != user["email"].lower():
            task_notify(db, kind="task_assigned", for_email=who,
                        title="You were assigned a task",
                        body=f"{t.title}", task_id=tid,
                        nexus_action={"view": "tasks", "sub": "mine", "label": "View task"})
    db.commit()
    db.refresh(t)
    fire_task_event(tid, "created")
    _asana_push(tid, user["email"])
    background_tasks.add_task(notify_task_event, tid, "created", user["email"])
    return task_to_dict(t)


_MODIFIED_FIELD_LABELS = {"title": "Title changed", "description": "Description changed",
                         "due_on": "Due date changed", "start_on": "Start date changed",
                         "priority": "Priority changed"}


@router.patch("/{task_id}")
def update_task(task_id: str, upd: TaskUpdate, background_tasks: BackgroundTasks,
                user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    t = _get_task(db, task_id)
    # The assignee counts on their OWN task - see require_task_role.
    require_task_role(db, user, t, "editor")
    data = upd.model_dump(exclude_unset=True)
    validate_task_payload(db, data, task_id=task_id)
    if "custom_field_values" in data:
        data["custom_field_values"] = coerce_custom_field_values(db, data["custom_field_values"])
    prev_assignees = set(task_assignees(t))
    prev_status = t.status
    prev_completed = bool(t.completed)
    prev_followers = set((t.follower_emails or []))
    modified_kinds = [label for field, label in _MODIFIED_FIELD_LABELS.items() if field in data]

    new_status = data.get("status", prev_status)
    # See _resolve_completed: `completed` and `status` are two columns for one
    # user-facing concept, and this rule is now shared with create and bulk so
    # the three cannot disagree about what "done" means.
    new_completed = _resolve_completed(data, prev_completed)
    _check_dependency_gate(db, t, prev_status, prev_completed, new_status, new_completed)

    for field, val in data.items():
        if field == "completed":
            continue  # handled below
        if field in ("assignee_email", "assignee_emails"):
            continue  # both are applied together, below
        if field == "owner_email":
            # strip() as well as lower(): a padded address is stored verbatim and
            # then never equals the same person anywhere else, so the task is
            # assigned to somebody who never sees it in My Tasks.
            val = (val or "").strip().lower()
        setattr(t, field, val)

    # Assignment goes through the setter so the legacy mirror can never drift
    # from the list, whichever field the caller sent.
    wanted = _assignees_from(data, list(prev_assignees))
    if wanted is not None:
        set_task_assignees(t, wanted)

    # Re-clean the extra list whenever either side of the project pair moved -
    # a caller can PATCH project_id alone (e.g. the primary Project picker) or
    # project_ids alone (the "also in" multi-select), and either one can leave
    # the primary duplicated inside the extras or a since-removed blank behind.
    if "project_id" in data or "project_ids" in data:
        t.project_ids = _extra_project_ids(t.project_ids, t.project_id or "")

    # Keeps `completed`, its timestamp and `status` in step regardless of which
    # one the caller actually sent.
    _apply_completion(t, new_completed, prev_completed)

    _sync_reciprocal_dependencies(db, t, data)

    t.modified_at = now_iso()

    # activity + notifications for meaningful changes
    acts = list(t.activity_ids or [])
    if "status" in data and data["status"] != prev_status:
        acts.append(log_activity(db, type="status_changed", actor_email=user["email"],
                                 entity_id=t.id, entity_code=t.code, entity_title=t.title,
                                 detail=f"changed status to {data['status']}"))
    now_assignees = set(task_assignees(t))
    if wanted is not None and now_assignees != prev_assignees:
        acts.append(log_activity(db, type="assignee_changed", actor_email=user["email"],
                                 entity_id=t.id, entity_code=t.code, entity_title=t.title,
                                 detail="reassigned this task"))
        # Assigning somebody LATER follows the task for you too, same as
        # creating it did. Only when people were actually added - taking an
        # assignee off is no reason to subscribe anyone. Skipped when this same
        # PATCH states the collaborator list explicitly: that is the caller
        # saying who is on it, including deciding it is not them.
        if (now_assignees - prev_assignees) and "follower_emails" not in data:
            _follow_as_actor(t, user["email"])
        # Only people newly ADDED are notified. Re-notifying everyone whenever a
        # third assignee joins would mail the original two about a task they
        # already hold.
        for who in (now_assignees - prev_assignees):
            if who != user["email"].lower():
                task_notify(db, kind="task_assigned", for_email=who,
                            title="You were assigned a task",
                            body=f"{t.title}", task_id=t.id,
                            nexus_action={"view": "tasks", "sub": "mine", "label": "View task"})
    if t.completed and not prev_completed:
        acts.append(log_activity(db, type="completed", actor_email=user["email"],
                                 entity_id=t.id, entity_code=t.code, entity_title=t.title,
                                 detail="completed this task"))
    t.activity_ids = acts

    # Completing a recurring task rolls the series forward to the next occurrence.
    spawned = _spawn_next_occurrence(db, t, user) if (t.completed and not prev_completed) else None

    db.commit()
    db.refresh(t)
    fire_task_event(t.id, "updated")
    _asana_push(t.id, user["email"])
    if spawned is not None:
        fire_task_event(spawned.id, "created")
        _asana_push(spawned.id, user["email"])

    if wanted is not None and (set(task_assignees(t)) - prev_assignees):
        background_tasks.add_task(notify_task_event, t.id, "assigned", user["email"])
    if t.completed and not prev_completed:
        background_tasks.add_task(notify_task_event, t.id, "completed", user["email"])
    for f in (set(t.follower_emails or []) - prev_followers):
        background_tasks.add_task(notify_task_event, t.id, "follower_added", user["email"], new_follower=f)
    if modified_kinds:
        background_tasks.add_task(notify_task_event, t.id, "modified", user["email"],
                                  update_kind=", ".join(modified_kinds))
    return task_to_dict(t)


@router.delete("/{task_id}", status_code=204)
def delete_task(task_id: str, background_tasks: BackgroundTasks,
                user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    t = _get_task(db, task_id)
    # The assignee counts on their own task here too. Note this cascade takes
    # the task's SUBTASKS with it (below), which may be assigned to other
    # people - deleting your own task can therefore remove someone else's.
    require_task_role(db, user, t, "editor")
    # Captured before the row is gone - notify_task_event("deleted") runs in
    # the background AFTER this response, by which point db.delete(t) below
    # has already removed it, so there'd be nothing left to look up.
    deleted_snapshot = {
        "code": t.code, "title": t.title, "status": t.status, "priority": t.priority,
        "assignee_email": t.assignee_email or "", "assignee_emails": task_assignees(t),
        "follower_emails": list(t.follower_emails or []),
        "created_by": t.created_by or "", "project_id": t.project_id or "", "due_on": t.due_on or "",
    }
    # detach from parent
    if t.parent_task_id:
        parent = db.query(models.Task).filter(models.Task.id == t.parent_task_id).first()
        if parent:
            parent.subtask_ids = [x for x in (parent.subtask_ids or []) if x != task_id]
    # clear reciprocal dependencies
    for other in db.query(models.Task).filter(
            (models.Task.blocked_by_ids.isnot(None))).all():
        changed = False
        if task_id in (other.blocked_by_ids or []):
            other.blocked_by_ids = [x for x in other.blocked_by_ids if x != task_id]; changed = True
        if task_id in (other.blocking_ids or []):
            other.blocking_ids = [x for x in other.blocking_ids if x != task_id]; changed = True
        if changed:
            other.modified_at = now_iso()
    # delete subtasks
    subs = db.query(models.Task).filter(models.Task.parent_task_id == task_id).all()
    gone_ids = [task_id] + [sub.id for sub in subs]
    # Tombstone every id being deleted, in this same transaction - a delta
    # fetch (GET /tasks/delta) otherwise can't tell "deleted" apart from
    # "unchanged, just not modified in this window". Same reasoning as the
    # Asana pending-delete queue below.
    deleted_stamp = now_iso()
    for gid in gone_ids:
        db.add(models.TaskDeleteLog(id=gen_id(), task_id=gid, deleted_at=deleted_stamp))
    # Asana counterparts of exactly the rows being deleted here, captured while
    # the links still exist - once they're gone nothing can re-derive them, so
    # unlike every other outbound change a lost deletion is lost for good.
    asana_gids = _asana_linked_gids(db, gone_ids)
    for sub in subs:
        db.delete(sub)
    db.query(models.TaskComment).filter(models.TaskComment.task_id == task_id).delete()
    db.query(models.TaskAttachment).filter(models.TaskAttachment.task_id == task_id).delete()
    # A link left pointing at a deleted task blocks that Asana task from ever
    # re-linking (_link_by_asana finds it, the task behind it is gone, and
    # _apply_inbound bails) - so the links go when the tasks do.
    db.query(models.AsanaTaskLink).filter(
        models.AsanaTaskLink.nexus_task_id.in_(gone_ids)).delete(synchronize_session=False)
    log_activity(db, type="deleted", actor_email=user["email"], entity_id=task_id,
                 entity_code=t.code, entity_title=t.title, detail="deleted this task")
    # Queued in THIS transaction so the tombstone lands with the delete. The
    # actual Asana call happens after the commit (a failing one must never roll
    # back the Nexus delete) and, off the sync worker, not at all until someone
    # runs Push all.
    _asana_queue_delete(db, asana_gids, t.title, t.code, user["email"])
    db.delete(t)
    db.commit()
    _asana_push_deleted()
    fire_task_event(task_id, "deleted")
    background_tasks.add_task(notify_task_event, task_id, "deleted", user["email"], snapshot=deleted_snapshot)


class BulkUpdate(BaseModel):
    ids: list[str]
    patch: dict[str, Any]


_BULK_ALLOWED = {"status", "priority", "assignee_email", "assignee_emails", "team_id", "project_id",
                 "completed", "tags", "due_on", "start_on", "is_milestone"}


@router.post("/bulk")
def bulk_update(body: BulkUpdate, user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    patch = {k: v for k, v in (body.patch or {}).items() if k in _BULK_ALLOWED}
    # Same validation update_task applies. Without it bulk was a way around it:
    # an arbitrary status set here still created a phantom board column.
    validate_task_payload(db, patch)
    rows = db.query(models.Task).filter(models.Task.id.in_(body.ids)).all()
    # Same editor requirement update_task enforces per task. Without this, bulk
    # was a straight bypass of it: _BULK_ALLOWED covers assignee, project,
    # completed and due date, so anyone who could merely VIEW a task could
    # reassign or close it by routing through here instead of PATCH.
    # All-or-nothing on purpose - a partial apply would leave the caller with no
    # way to tell which ids were silently skipped.
    for t in rows:
        require_task_role(db, user, t, "editor")
    if "status" in patch or "completed" in patch:
        for t in rows:
            prev_status, prev_completed = t.status, bool(t.completed)
            new_status = patch.get("status", prev_status)
            new_completed = bool(patch.get("completed", prev_completed)) or (patch.get("status") == "completed")
            _check_dependency_gate(db, t, prev_status, prev_completed, new_status, new_completed)
    # Captured before the patch lands so the activity entries below can say what
    # actually changed rather than restating the new value for every row.
    before = {t.id: (t.status, bool(t.completed), tuple(task_assignees(t))) for t in rows}
    for t in rows:
        prev_status, prev_completed, _ = before[t.id]
        # Decided BEFORE the loop writes anything, from the payload and the
        # previous value - the same call update_task makes. `status` used to
        # fall through to the setattr below, so bulk-setting it to "completed"
        # left the flag off (no strikethrough, still emailed as overdue), and
        # bulk-setting it to anything else left a completed task stuck in the
        # Completed bucket. Which of the two won even depended on dict order.
        new_completed = _resolve_completed(patch, prev_completed)
        for k, v in patch.items():
            if k == "completed":
                continue    # handled by _apply_completion below
            if k in ("assignee_email", "assignee_emails"):
                continue    # applied together, below
            setattr(t, k, v)
        bulk_assignees = _assignees_from(patch, list(before[t.id][2]))
        if bulk_assignees is not None:
            set_task_assignees(t, bulk_assignees)
        # A section and a team both belong to ONE project, so carrying either
        # across a move leaves the task pointing at a section the destination
        # does not have - it then renders under a phantom group that no board
        # can show or clear. Done here rather than in the client so no caller
        # can skip it (the drawer's single-task picker clears team_id by hand;
        # bulk moves 31 rows at a time and must not depend on remembering).
        # An explicit team_id in the same patch still wins.
        if "project_id" in patch:
            t.section_id = ""
            if "team_id" not in patch:
                t.team_id = ""
            # Same re-clean update_task does: the primary must never also sit
            # in the "also in" extras list.
            t.project_ids = _extra_project_ids(t.project_ids, t.project_id or "")
        _apply_completion(t, new_completed, prev_completed)
        t.modified_at = now_iso()

    # Activity + notifications. Bulk previously did NEITHER: reassigning fifty
    # tasks told nobody and left no audit trail, while the same edit made one at
    # a time through PATCH did both.
    #
    # Activity is per task (it's an audit trail - one row per thing that
    # changed), but the bell notification is AGGREGATED per person: fifty
    # separate "you were assigned a task" pings for one action is the failure
    # mode CLAUDE.md warns about ("one notification per workflow, never one per
    # item"). Email is deliberately not sent from here at all - see below.
    actor = (user["email"] or "").lower()
    newly_assigned: dict[str, list] = {}
    for t in rows:
        prev_status, prev_completed, prev_assignees = before[t.id]
        acts = list(t.activity_ids or [])
        if "status" in patch and t.status != prev_status:
            acts.append(log_activity(db, type="status_changed", actor_email=user["email"],
                                     entity_id=t.id, entity_code=t.code, entity_title=t.title,
                                     detail=f"changed status to {t.status}"))
        now_assignees = set(task_assignees(t))
        if now_assignees != set(prev_assignees):
            acts.append(log_activity(db, type="assignee_changed", actor_email=user["email"],
                                     entity_id=t.id, entity_code=t.code, entity_title=t.title,
                                     detail="reassigned this task"))
            # Only the people newly added, and still aggregated per person, so a
            # fifty-task reassignment is one ping each rather than fifty.
            for who in (now_assignees - set(prev_assignees)):
                if who != actor:
                    newly_assigned.setdefault(who, []).append(t)
            # Bulk-assigning is assigning: the person who did it follows each
            # task they put on someone else's plate, exactly as PATCH does.
            if (now_assignees - set(prev_assignees)) and "follower_emails" not in patch:
                _follow_as_actor(t, actor)
        if t.completed and not prev_completed:
            acts.append(log_activity(db, type="completed", actor_email=user["email"],
                                     entity_id=t.id, entity_code=t.code, entity_title=t.title,
                                     detail="completed this task"))
        t.activity_ids = acts

    for email, assigned in newly_assigned.items():
        # One task through bulk is the same action as one through PATCH, so it
        # reads identically; more than one collapses into a single ping.
        if len(assigned) == 1:
            one = assigned[0]
            task_notify(db, kind="task_assigned", for_email=email,
                        title="You were assigned a task",
                        body=f"{one.title}", task_id=one.id,
                        nexus_action={"view": "tasks", "sub": "mine", "label": "View task"})
        else:
            task_notify(db, kind="task_assigned", for_email=email,
                        title=f"You were assigned {len(assigned)} tasks",
                        body=", ".join(t.title or "Untitled task" for t in assigned[:5])
                             + (f" and {len(assigned) - 5} more" if len(assigned) > 5 else ""),
                        nexus_action={"view": "tasks", "sub": "mine", "label": "View tasks"})

    db.commit()
    fire_task_event("", "bulk")
    return [task_to_dict(t) for t in rows]


# ── Comments ─────────────────────────────────────────────────────────────────
class CommentCreate(BaseModel):
    body: str
    # No author_email. create_comment() accepts one so the Asana importer and the
    # inbound-email ingester can attribute a backfilled comment to whoever
    # actually wrote it - but both call that function in-process. Exposing it on
    # the HTTP body let any signed-in caller post a comment signed as a colleague,
    # in a thread that is used as a record of who said what. Nothing ever sent it.


class CommentUpdate(BaseModel):
    body: Optional[str] = None
    pinned: Optional[bool] = None


@router.get("/{task_id}/comments")
def list_comments(task_id: str, db: Session = Depends(get_db)):
    # 404 rather than an empty list for a task that isn't there: "no comments"
    # and "no such task" are different answers, and returning the first for the
    # second made a drawer left open on a deleted task look merely empty. POST
    # and PATCH on the same id already 404.
    _get_task(db, task_id)
    rows = db.query(models.TaskComment).filter(models.TaskComment.task_id == task_id).all()
    return [comment_to_dict(c) for c in rows]


@router.post("/{task_id}/comments", status_code=201)
def add_comment(task_id: str, body: CommentCreate, background_tasks: BackgroundTasks, notify: bool = True,
                user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    # `notify=false` (query param) posts silently - used by the Asana importer to
    # backfill historical comments without pinging assignees/followers. Defaults
    # to true, so normal in-app commenting is unchanged.
    t = _get_task(db, task_id)
    require_task_role(db, user, t, "commenter")
    # The comment itself and all six of its side effects live in create_comment
    # (routers/task_util.py) - see that docstring for why this endpoint is only
    # a wrapper. `defer` keeps the notification emails off the response path.
    # author_email is deliberately NOT passed through from the request - it
    # defaults to the actor. See CommentCreate.
    c = create_comment(db, t, actor_email=user["email"],
                       body=body.body or "", notify=notify, defer=background_tasks.add_task)
    return comment_to_dict(c)


@router.patch("/comments/{comment_id}")
def edit_comment(comment_id: str, upd: CommentUpdate, user: dict = Depends(get_current_user),
                 db: Session = Depends(get_db)):
    """Editing the TEXT is author-only; pinning is a curation action any project
    editor may take. Before this, neither was checked at all (the handler took no
    `user`), so any signed-in person could rewrite anyone's comment on any task,
    including in a project they can't otherwise see."""
    c = db.query(models.TaskComment).filter(models.TaskComment.id == comment_id).first()
    if not c:
        raise HTTPException(404, "Comment not found")
    t = db.query(models.Task).filter(models.Task.id == c.task_id).first()
    project = project_for_task(db, t) if t else None
    if upd.body is not None:
        # Managers bypass every other project-role check in this module, but
        # rewriting someone else's words is different in kind from access - so
        # the author check is absolute.
        if (c.author_email or "").lower() != (user["email"] or "").lower():
            raise HTTPException(403, "Only the author can edit a comment.")
        c.body = upd.body
        c.edited_at = now_iso()
    if upd.pinned is not None:
        require_project_role(db, user, project, "editor")
        c.pinned = bool(upd.pinned)
    if t:
        t.modified_at = now_iso()   # so a delta fetch (GET /tasks/delta) picks this task up
    db.commit()
    db.refresh(c)
    fire_task_event(c.task_id, "comment")
    # A corrected comment has to reach Asana too, or the copy most of the
    # workspace reads keeps the wrong text forever. Only on a body change -
    # pinning is a Nexus-only notion with no Asana counterpart.
    if upd.body is not None:
        _asana_push_comment_edit(comment_id)
    return comment_to_dict(c)


@router.delete("/comments/{comment_id}", status_code=204)
def delete_comment(comment_id: str, user: dict = Depends(get_current_user),
                   db: Session = Depends(get_db)):
    """The author, or a manager moderating the thread. Previously unguarded
    entirely - see edit_comment.

    "Project editor" was too weak a bar. On a task with no project, or one in an
    org-level project, every signed-in employee is an editor - so anyone could
    delete a colleague's comment out of a thread that is used as the record of
    who said what, leaving nothing behind to show it happened. edit_comment
    already treats rewriting someone's words as different in kind from access
    and makes the author check absolute; deleting them is not a smaller act.
    Moderation stays possible, but it takes manager+."""
    c = db.query(models.TaskComment).filter(models.TaskComment.id == comment_id).first()
    if not c:
        raise HTTPException(404, "Comment not found")
    t = db.query(models.Task).filter(models.Task.id == c.task_id).first()
    if (c.author_email or "").lower() != (user["email"] or "").lower():
        require_project_role(db, user, project_for_task(db, t) if t else None, "editor")
        if user.get("level", 1) < 3:
            raise HTTPException(403, "Only the author or a manager can delete a comment.")
    if t:
        t.comment_ids = [x for x in (t.comment_ids or []) if x != comment_id]
        t.modified_at = now_iso()   # so a delta fetch (GET /tasks/delta) picks this task up
    db.delete(c)
    db.commit()
    fire_task_event(c.task_id, "comment")


# ── Attachments (bytes live in Supabase storage; `url` points there) ─────────
class AttachmentCreate(BaseModel):
    name: str
    size: Optional[str] = ""
    kind: Optional[str] = "other"
    url: Optional[str] = ""
    comment_id: Optional[str] = ""   # set only when attached while composing a comment


@router.get("/{task_id}/attachments")
def list_attachments(task_id: str, db: Session = Depends(get_db)):
    rows = db.query(models.TaskAttachment).filter(models.TaskAttachment.task_id == task_id).all()
    return [attachment_to_dict(a) for a in rows]


@router.post("/{task_id}/attachments", status_code=201)
def add_attachment(task_id: str, body: AttachmentCreate, user: dict = Depends(get_current_user),
                   db: Session = Depends(get_db)):
    t = _get_task(db, task_id)
    # "commenter", not "editor", and deliberately: a commenter may attach a file
    # while composing a comment (see AttachmentCreate.comment_id), so requiring
    # editor here would 403 exactly the people that flow exists for. Matches
    # add_comment's own threshold.
    require_task_role(db, user, t, "commenter")
    # An attachment tagged with a comment from a DIFFERENT task is incoherent:
    # the comment view groups by comment id within one task, so the file would
    # simply never render, while the row claims an association that isn't real.
    if body.comment_id:
        parent = db.query(models.TaskComment).filter(
            models.TaskComment.id == body.comment_id).first()
        if not parent or parent.task_id != task_id:
            raise HTTPException(422, "comment_id must be a comment on this task.")
    aid = gen_id()
    # Small files arrive from the composer as base64 data: URLs. Store the bytes
    # in Supabase Storage and keep only the link - inline bytes in this column
    # once grew the prod DB to 5.9 GB (see task_files.py). Falls back to the
    # inline URL if storage is unreachable, which is never worse than before.
    url = body.url or ""
    if url.startswith("data:"):
        url = data_url_to_storage(body.name, url) or url
    a = models.TaskAttachment(id=aid, task_id=task_id, name=body.name, size=body.size or "",
                              kind=body.kind or "other", url=url,
                              added_at=now_iso(), added_by=user["email"],
                              comment_id=body.comment_id or "")
    db.add(a)
    t.attachment_ids = list(t.attachment_ids or []) + [aid]
    act = log_activity(db, type="attached", actor_email=user["email"], entity_id=task_id,
                       entity_code=t.code, entity_title=t.title,
                       detail=f'attached "{a.name}"')
    t.activity_ids = list(t.activity_ids or []) + [act]
    t.modified_at = now_iso()   # so a delta fetch (GET /tasks/delta) picks this task up
    db.commit()
    db.refresh(a)
    fire_task_event(task_id, "attachment")
    return attachment_to_dict(a)


@router.delete("/attachments/{attachment_id}", status_code=204)
def delete_attachment(attachment_id: str, user: dict = Depends(get_current_user),
                      db: Session = Depends(get_db)):
    a = db.query(models.TaskAttachment).filter(models.TaskAttachment.id == attachment_id).first()
    if not a:
        raise HTTPException(404, "Attachment not found")
    t = db.query(models.Task).filter(models.Task.id == a.task_id).first()
    # Whoever uploaded it can remove it; otherwise it takes an editor. Removing
    # someone else's evidence from a task shouldn't be open to every viewer.
    if (a.added_by or "").lower() != (user["email"] or "").lower():
        require_project_role(db, user, project_for_task(db, t) if t else None, "editor")
    if t:
        t.attachment_ids = [x for x in (t.attachment_ids or []) if x != attachment_id]
        act = log_activity(db, type="attachment_removed", actor_email=user["email"], entity_id=t.id,
                           entity_code=t.code, entity_title=t.title,
                           detail=f'removed attachment "{a.name}"')
        t.activity_ids = list(t.activity_ids or []) + [act]
        t.modified_at = now_iso()   # so a delta fetch (GET /tasks/delta) picks this task up
    db.delete(a)
    db.commit()
    fire_task_event(a.task_id, "attachment")


# ── Activity ─────────────────────────────────────────────────────────────────
@router.get("/activity")
def global_activity(limit: int = 500, db: Session = Depends(get_db)):
    rows = (db.query(models.TaskActivity)
            .order_by(models.TaskActivity.at.desc()).limit(min(limit, 2000)).all())
    return [activity_to_dict(a) for a in rows]


@router.get("/{task_id}/activity")
def task_activity(task_id: str, db: Session = Depends(get_db)):
    # "Created from Asana" / "Updated from Asana" are the sync's own bookkeeping,
    # logged once per inbound apply. On a task's own timeline they say nothing a
    # reader wants: the very next row is the real story ("changed the due date to
    # Aug 22", by the person who did it), so the marker just doubles the length
    # of the list with a robot's name against every entry. Still written, and
    # still shown in the workspace-wide feed (GET /tasks/activity), where "where
    # did this change come from" is the actual question being asked.
    rows = (db.query(models.TaskActivity)
            .filter(models.TaskActivity.entity_id == task_id,
                    models.TaskActivity.type != "synced_from_asana")
            .order_by(models.TaskActivity.at.desc()).all())
    return [activity_to_dict(a) for a in rows]


# ── Sections & custom statuses (board columns) ───────────────────────────────
class SectionBody(BaseModel):
    id: Optional[str] = None
    project_id: Optional[str] = ""
    # Optional so a PATCH can send only what it changes - a drag-to-reorder
    # carries a position and no name, and a required `name` here rejected it
    # with "name Field required" for a field it was not touching. Create still
    # demands one (guarded in create_section), so a nameless section is
    # impossible either way. Same shape RuleBody already uses.
    name: Optional[str] = None
    position: Optional[int] = 0


@router.get("/meta/sections")
def list_sections(db: Session = Depends(get_db)):
    return [section_to_dict(s) for s in db.query(models.TaskSection).all()]


def _require_section_editor(db: Session, user: dict, project_id: str) -> None:
    """Sections are a project's board columns, so editing them is an editor
    action on that project. A section with no project (project_id "") is
    workspace-level and stays unrestricted, matching how project_role_for
    already treats a task with no project.

    All three section endpoints previously took no `user` at all - anyone
    signed in could rename or delete another project's board columns."""
    if not project_id:
        return
    project = db.query(models.TaskProject).filter(models.TaskProject.id == project_id).first()
    require_project_role(db, user, project, "editor")


@router.post("/meta/sections", status_code=201)
def create_section(body: SectionBody, user: dict = Depends(get_current_user),
                   db: Session = Depends(get_db)):
    _require_section_editor(db, user, body.project_id or "")
    if not (body.name or "").strip():
        raise HTTPException(422, "A section needs a name.")
    s = models.TaskSection(id=body.id or gen_id(), project_id=body.project_id or "",
                           name=body.name.strip(), position=body.position or 0, created_at=now_iso())
    db.add(s)
    db.commit()
    db.refresh(s)
    return section_to_dict(s)


@router.patch("/meta/sections/{section_id}")
def update_section(section_id: str, body: SectionBody, user: dict = Depends(get_current_user),
                   db: Session = Depends(get_db)):
    s = db.query(models.TaskSection).filter(models.TaskSection.id == section_id).first()
    if not s:
        raise HTTPException(404, "Section not found")
    _require_section_editor(db, user, s.project_id or "")
    if body.name:
        s.name = body.name
    if body.position is not None:
        s.position = body.position
    db.commit()
    db.refresh(s)
    return section_to_dict(s)


@router.delete("/meta/sections/{section_id}", status_code=204)
def delete_section(section_id: str, user: dict = Depends(get_current_user),
                   db: Session = Depends(get_db)):
    s = db.query(models.TaskSection).filter(models.TaskSection.id == section_id).first()
    if not s:
        return
    _require_section_editor(db, user, s.project_id or "")
    # Release the tasks that were filed here. Deleting the row alone left them
    # pointing at a section that no longer resolves, so they vanished from any
    # view that groups by a known section - present in the data, invisible in
    # the UI. modified_at is bumped so GET /tasks/delta carries the change.
    orphaned = db.query(models.Task).filter(models.Task.section_id == section_id).all()
    for t in orphaned:
        t.section_id = ""
        t.modified_at = now_iso()
    db.delete(s)
    db.commit()
    for t in orphaned:
        fire_task_event(t.id, "updated")


class CustomStatusBody(BaseModel):
    id: Optional[str] = None
    label: Optional[str] = None
    color: Optional[str] = ""
    position: Optional[int] = 0
    project_ids: Optional[list] = None


@router.post("/meta/custom-statuses/dedupe", dependencies=[Depends(require_manager)])
def dedupe_custom_statuses(db: Session = Depends(get_db)):
    """Collapse same-named statuses onto one row scoped to every project that
    used them.

    Repairs databases seeded before the Asana sync matched on label: Asana's
    "Task Progress" is usually a PER-PROJECT custom field, so each project's
    "Waiting" carried its own option gid and minted its own row.

    Safe to run repeatedly - a no-op once there is nothing to merge. It remaps
    Task.status off every row it deletes, so no task is left pointing at a
    status that no longer exists.
    """
    import asana_sync
    result = asana_sync.dedupe_custom_statuses(db)
    db.commit()
    return result


@router.get("/meta/custom-statuses")
def list_custom_statuses(project_id: str = "", db: Session = Depends(get_db)):
    """`project_id` narrows to the statuses that project actually uses (its own
    plus any global one). Omitted returns every status, which is what Manage
    needs - the board passes it so a stage invented for one project stops
    appearing as a column on every other."""
    rows = db.query(models.TaskCustomStatus).all()
    if project_id:
        rows = [s for s in rows
                if not [p for p in (s.project_ids or []) if p]
                or project_id in (s.project_ids or [])]
    return [custom_status_to_dict(s) for s in rows]


@router.post("/meta/custom-statuses", status_code=201, dependencies=[Depends(require_manager)])
def create_custom_status(body: CustomStatusBody, db: Session = Depends(get_db)):
    s = models.TaskCustomStatus(id=body.id or gen_id(), label=body.label or "",
                                color=body.color or "", position=body.position or 0,
                                project_ids=[p for p in (body.project_ids or []) if p])
    db.add(s)
    db.commit()
    db.refresh(s)
    return custom_status_to_dict(s)


@router.patch("/meta/custom-statuses/{status_id}", dependencies=[Depends(require_manager)])
def update_custom_status(status_id: str, body: CustomStatusBody, db: Session = Depends(get_db)):
    s = db.query(models.TaskCustomStatus).filter(models.TaskCustomStatus.id == status_id).first()
    if not s:
        raise HTTPException(404, "Custom status not found")
    data = body.model_dump(exclude_unset=True, exclude={"id"})
    if "project_ids" in data:
        data["project_ids"] = [p for p in (data["project_ids"] or []) if p]
    for k, v in data.items():
        setattr(s, k, v)
    db.commit()
    db.refresh(s)
    return custom_status_to_dict(s)


@router.delete("/meta/custom-statuses/{status_id}", status_code=204, dependencies=[Depends(require_manager)])
def delete_custom_status(status_id: str, db: Session = Depends(get_db)):
    db.query(models.TaskCustomStatus).filter(models.TaskCustomStatus.id == status_id).delete()
    db.commit()


# ── Task email notification settings (admin) ─────────────────────────────────
from task_notify import get_settings as _get_task_notify_settings, save_settings as _save_task_notify_settings


@router.get("/notify/settings")
def get_task_notify_settings(user: dict = Depends(require_manager), db: Session = Depends(get_db)):
    return _get_task_notify_settings(db)


@router.put("/notify/settings")
def put_task_notify_settings(patch: dict, user: dict = Depends(require_manager), db: Session = Depends(get_db)):
    return _save_task_notify_settings(db, patch, user["email"])


@router.get("/notify/log")
def get_task_notify_log(task_id: str = "", status: str = "", limit: int = 200,
                        user: dict = Depends(require_manager), db: Session = Depends(get_db)):
    q = db.query(models.TaskEmailLog)
    if task_id:
        q = q.filter(models.TaskEmailLog.task_id == task_id)
    if status:
        q = q.filter(models.TaskEmailLog.status == status)
    rows = q.order_by(models.TaskEmailLog.created_at.desc()).limit(min(limit, 500)).all()
    return [{
        "id": r.id, "taskId": r.task_id, "taskCode": r.task_code, "eventType": r.event_type,
        "eventVersion": r.event_version, "recipient": r.recipient, "recipientRole": r.recipient_role,
        "subject": r.subject, "status": r.status, "graphMessageId": r.graph_message_id,
        "conversationId": r.conversation_id, "attempts": r.attempts, "error": r.error,
        "createdAt": r.created_at, "updatedAt": r.updated_at,
    } for r in rows]


# ── Inbound email (replies -> comments) ──────────────────────────────────────
@router.get("/inbound/log")
def get_task_inbound_log(task_id: str = "", status: str = "", limit: int = 200,
                         user: dict = Depends(require_manager), db: Session = Depends(get_db)):
    """What the task mailbox has handed us and what became of it. The answer to
    "I replied and nothing happened" - `reason` says which check refused it."""
    q = db.query(models.TaskInboundEmail)
    if task_id:
        q = q.filter(models.TaskInboundEmail.task_id == task_id)
    if status:
        q = q.filter(models.TaskInboundEmail.status == status)
    rows = q.order_by(models.TaskInboundEmail.processed_at.desc()).limit(min(limit, 500)).all()
    return [{
        "id": r.id, "taskId": r.task_id, "commentId": r.comment_id, "from": r.from_email,
        "subject": r.subject, "status": r.status, "reason": r.reason, "matchedBy": r.matched_by,
        "attachmentCount": r.attachment_count, "receivedAt": r.received_at,
        "processedAt": r.processed_at,
    } for r in rows]


@router.post("/inbound/drain")
def drain_task_inbox(user: dict = Depends(require_manager), db: Session = Depends(get_db)):
    """Run one pass now instead of waiting for the 60s loop - the manual
    counterpart to Asana's Pull, and the only way to exercise this on an
    instance that isn't the sync worker. A plain `def` endpoint, so FastAPI
    runs its blocking Graph calls in a threadpool rather than on the event
    loop."""
    from task_inbound import drain_once
    try:
        return drain_once(db)
    except Exception as e:
        # Surfaced, not swallowed: the first thing this hits on a new mailbox is
        # a missing Mail.ReadWrite grant, and "500" would not say so.
        raise HTTPException(502, f"Could not read the task mailbox: {e}")
